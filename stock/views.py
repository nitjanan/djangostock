from decimal import Decimal
from dis import dis
from multiprocessing import context
import numbers
from pickletools import decimalnl_short
import re
from tkinter import N
from turtle import numinput, position, title
from typing import cast
from unicodedata import decimal, numeric
from urllib import response
from django import forms
from django import http
from django.core import paginator
from django.db.models.fields import NullBooleanField
from django.db.models.query import QuerySet
from django.http import request, HttpResponseRedirect, HttpResponse ,JsonResponse, HttpResponseNotAllowed, StreamingHttpResponse
from django.shortcuts import redirect, render, get_object_or_404
from stock.models import BaseAffiliatedCompany, BaseBranchCompany, BaseDepartment, BaseSparesType, BaseUnit, BaseUrgency, Category, Distributor, Position, Product, Cart, CartItem, Order, OrderItem, PurchaseOrder, PurchaseRequisition, Receive, ReceiveItem, Requisition, RequisitionItem, CrudUser, BaseApproveStatus, UserProfile,PositionBasePermission, PurchaseOrderItem,ComparisonPrice, ComparisonPriceItem, ComparisonPriceDistributor, BasePermission, BaseVisible, BranchCompanyBaseAdress, RateDistributor, BasePOType, BaseCar, BaseRepairType, Invoice, InvoiceItem, BaseExpenseDepartment, ExOESTND, ExOESTNH, ExOEINVH, ExOEINVD, Maintenance, BaseMAType, CarLogbook, UserCarDepartment, BaseJobCarDep, ApproveCarDepartment, PmRoundItem
from stock.forms import SignUpForm, RequisitionForm, RequisitionItemForm, PurchaseRequisitionForm, UserProfileForm, PurchaseOrderForm, PurchaseOrderPriceForm, ComparisonPriceForm, CPDModelForm, CPDForm, CPSelectBidderForm, PurchaseOrderFromComparisonPriceForm, ReceiveForm, ReceivePriceForm, PurchaseOrderReceiptForm, RequisitionMemorandumForm, PurchaseRequisitionAddressCompanyForm, ComparisonPriceAddressCompanyForm, PurchaseOrderAddressCompanyForm, PurchaseOrderCancelForm, RateDistributorForm, PurchaseRequisitionOrganizerForm, MaintenanceForm, CarLogbookForm, RoiCarLogbookForm, CrMaintenanceForm, CPCancelForm
from django.contrib.auth.models import Group,User
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import login, authenticate, logout
from django.core.paginator import Paginator, EmptyPage, InvalidPage
from django.contrib.auth.decorators import login_required, permission_required 
from django.views.decorators.http import require_http_methods
from django.conf import settings
from django.views.generic import ListView, View, TemplateView, DeleteView
from django.template.loader import render_to_string
from django.urls import reverse
from .filters import ComparisonPriceFilter, RequisitionFilter, PurchaseRequisitionFilter, PurchaseOrderFilter,ComparisonPriceFilter, ReceiveFilter, PurchaseOrderItemFilter, RateDistributorFilter, DistributorFilter, InvoidFilter, ExOESTNHFilter, ExOEINVHFilter, MaintenanceFilter, CarLogbookFilter
from .forms import PurchaseOrderItemFormset, PurchaseOrderItemModelFormset, PurchaseOrderItemInlineFormset, CPitemFormset, CPitemInlineFormset, ReceiveItemForm, RequisitionItemModelFormset, ReceiveItemInlineFormset
from django.forms import inlineformset_factory
import stripe, logging, datetime
from django.db.models import Prefetch, Sum
from .resources import ReceiveItemResource, DistributorResource
from tablib import Dataset
from django.db.models import Q
from django.core.cache import cache
import json
from django.core.serializers import serialize
from django.db.models import Count, Avg
import xlwt
from django.db.models import F, Func, Value, CharField,When, Q, Case, ExpressionWrapper
from django.db.models import Sum, IntegerField, Max
from django.views.decorators.cache import cache_control
from decimal import Decimal
import pandas as pd
from django_pandas.io import read_frame
from django.db.models.functions import Round, Concat, Coalesce
from django.contrib import messages
import string
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from django.db.models import Avg
from datetime import date, timedelta
from django.utils.timezone import is_aware, make_naive, make_aware
from django.db import models
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework import generics, viewsets, permissions, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.request import Request
from rest_framework.renderers import TemplateHTMLRenderer
from rest_framework.authentication import SessionAuthentication, BasicAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import api_view
from rest_framework.pagination import PageNumberPagination

from .tokens import create_jwt_pair_for_user
from stock.serializers import SignUpSerializer, PurchaseOrderSerializer, PurchaseOrderItemSerializer, BaseCarSerializer, UserCarDepartmentSerializer
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from django.db.models import OuterRef, Subquery
from django.utils import timezone
import pytz
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
from django.db.models.functions import ExtractMonth, ExtractYear, Concat, Cast
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font, Color, NamedStyle, Side, Border
from openpyxl.utils import get_column_letter
from collections import defaultdict
from itertools import groupby
from operator import itemgetter
from dateutil.rrule import rrule, MONTHLY
import os
from openpyxl.drawing.image import Image as XLImage
import qrcode
from io import BytesIO
from django.core.files import File
from PIL import Image, ImageDraw, ImageOps
from django.views.decorators.csrf import csrf_exempt
from django.core.exceptions import ObjectDoesNotExist

#ใช้ในระบบงาน ค้นหาด้วย code or name
def car_search(request):
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        search_term = request.GET.get('q', '')
        rq_type = request.GET.getlist('rq_type[]') or request.GET.getlist('rq_type')


        cars = BaseCar.objects.filter(Q(name__icontains=search_term) | Q(code__icontains=search_term), (Q(rq_type__in = rq_type)| Q(rq_type = 4))).values('id', 'name', 'code')
        results = []
        for car in cars:
            results.append({
                'id': car['id'],
                'text': f"{car['code']} : {car['name']}"
            })
        return JsonResponse({'results': results})
    return JsonResponse({'results': []})

def findCompanyIn(request):
    code = request.session['company_code']

    #หาหน้าต่างการมองเห็นบริษัททั้งหมดของ user
    user_profile = UserProfile.objects.get(user = request.user.id)
    company_all = BaseBranchCompany.objects.filter(userprofile = user_profile).values('code')

    if code == "ALL":
        company_in = company_all
    else:
        company_in = BaseBranchCompany.objects.filter(code = code).values('code')
    return company_in

def set_border(ws, side=None, blank=True):
    wb = ws._parent
    side = side if side else Side(border_style='thin', color='000000')
    for cell in ws._cells.values():
        cell.border = Border(top=side, bottom=side, left=side, right=side)


#check error id 23-05-2023
def has_only_en(name):
    char_set = string.ascii_letters + string.digits + "-"
    return all((True if x in char_set else False for x in name))

#button New โอน 23-05-2023
def newOld(old_product_id, new_product_id):
    alertStr = None
    alertType = None

    if not has_only_en(new_product_id): #เช็คตัวอักษรภาษาไทยในรหัส
        alertStr = "รหัสสินค้าใหม่ ("+ str(new_product_id) +") มีตัวอักษรภาษาไทย ไม่สามารถ new โอนได้"
        alertType = messages.WARNING
    elif new_product_id.find('O-') != -1 : #เช็ค O- ในรหัส
        alertStr = "รหัสสินค้าใหม่ XX-XO-001 ต้องเป็น XX-X0-001 (ขีดเอ็กซ์โอ ต้องเป็น ขีดเอ็กซ์ศูนย์) ไม่สามารถ new โอนได้"
        alertType = messages.WARNING
    else:
        try:
            #หารหัสใหม่
            have_new_product_id = Product.objects.get(id = new_product_id)
            if have_new_product_id:
                alertStr = "มีรหัสสินค้าใหม่อยู่แล้ว '" +str(new_product_id)+ "' ไม่สามารถ new โอนได้"
                alertType = messages.WARNING
        except Product.DoesNotExist:
            try:
                old_product = Product.objects.get(id = old_product_id)
                #สร้างรหัสสินค้าใหม่
                if old_product and new_product_id:
                    obj = Product.objects.create(
                        id = new_product_id,
                        name = old_product.name + "(*)",
                        unit = old_product.unit,
                        slug = new_product_id.replace('-', "I"),
                        description = old_product.description,
                        price = old_product.price,
                        category = old_product.category,
                        image = old_product.image,
                        stock = old_product.stock,
                        available = old_product.available,
                        created = old_product.created,
                        update = old_product.update,
                        affiliated = old_product.affiliated
                    )
                    obj.save()

                    try:
                        #ย้ายสินค้าในใบขอเบิกไปรหัสสินค้าใหม่
                        r_item = RequisitionItem.objects.filter(product__id = old_product.id)
                        new_product = Product.objects.get(id = new_product_id)
                        for i in r_item:
                            i.product = new_product
                            i.save()
                        #หลังจากย้ายสินค้าไปรหัสใหม่
                        af_item = RequisitionItem.objects.filter(product__id = old_product.id).count()
                        if af_item < 1:
                            old_product.delete()
                            #แก้ชื่อ new_product
                            new_product.name =  new_product.name[:-3]
                            new_product.save()

                            # Create product QR code
                            create_product_qr(new_product)

                            alertStr = "new โอน รหัสสินค้า '"+ str(old_product_id) + "' เป็น '"+ str(new_product_id)+ "' สำเร็จแล้ว"
                            alertType = messages.SUCCESS
                    except RequisitionItem.DoesNotExist:
                        pass
                    
            except Product.DoesNotExist:
                alertStr = "ไม่มีรหัสสินค้าเก่าในระบบ ไม่สามารถ new โอนได้"
                alertType = messages.WARNING
                pass
    return alertType,alertStr

#button Change Exist 23-05-2023
def changeExist(old_product_id, new_product_id):
    alertStr = None
    alertType = None
    try:
        #หารหัสเก่า
        old_product = Product.objects.get(id = old_product_id)
        #หารหัสใหม่
        new_product = Product.objects.get(id = new_product_id)
        if old_product and new_product:
            try:
                #ย้ายสินค้าในใบขอเบิกไปรหัสสินค้าใหม่
                r_item = RequisitionItem.objects.filter(product__id = old_product.id)
                for i in r_item:
                    i.product = new_product
                    i.save()
                #หลังจากย้ายสินค้าไปรหัสใหม่
                af_item = RequisitionItem.objects.filter(product__id = old_product.id).count()
                if af_item < 1:
                    old_product.delete()

                    alertStr = "เปลี่ยนรหัสสินค้าจาก '"+ str(old_product_id) + "' เป็น '"+ str(new_product_id)+ "' สำเร็จแล้ว"
                    alertType = messages.SUCCESS
            except RequisitionItem.DoesNotExist:
                pass

    except Product.DoesNotExist:
        alertStr = "ไม่มีรหัสสินค้าในระบบ ไม่สามารถเปลี่ยนรหัสสินค้าได้"
        alertType = messages.WARNING
        pass
    return alertType,alertStr

#button New โอน 17-06-2024
def newOldDistributor(old_distributor_id, new_distributor_id):
    alertStr = None
    alertType = None

    try:
        #หารหัสใหม่
        have_new_distributor_id = Distributor.objects.get(id = new_distributor_id)
        if have_new_distributor_id:
            alertStr = "มีรหัสผู้จัดจำหน่ายใหม่อยู่แล้ว'" +str(new_distributor_id)+ "' ไม่สามารถ new โอนได้"
            alertType = messages.WARNING
    except Distributor.DoesNotExist:
        try:
            old_distributor = Distributor.objects.get(id = old_distributor_id)

            #สร้างรหัสสินค้าใหม่
            if old_distributor and new_distributor_id:
                obj = Distributor.objects.create(
                    id = new_distributor_id,
                    prefix = old_distributor.prefix,
                    name = old_distributor.name + "(*)",
                    type = old_distributor.type,
                    genre = old_distributor.genre,
                    credit = old_distributor.credit,
                    vat_type = old_distributor.vat_type,
                    discount = old_distributor.discount,
                    credit_limit = old_distributor.credit_limit,
                    account_number = old_distributor.account_number,
                    address = old_distributor.address,
                    tel = old_distributor.tel,
                    payment = old_distributor.payment,
                    contact = old_distributor.contact,
                    affiliated = old_distributor.affiliated,
                    tex = old_distributor.tex,
                    fax = old_distributor.fax,
                    registration_pdf = old_distributor.registration_pdf,
                    created = old_distributor.created
                )
                obj.save()

                try:
                    #ย้ายผู้จัดจำหน่ายในใบเปรียบเทียบไปรหัสผู้จัดจำหน่ายใหม่
                    ###### 1 ######
                    cp = ComparisonPrice.objects.filter(select_bidder = old_distributor)
                    cp.update(select_bidder = new_distributor_id)

                except ComparisonPrice.DoesNotExist:
                    pass

                try:
                    #ย้ายผู้จัดจำหน่ายในใบสั่งซื้อไปรหัสผู้จัดจำหน่ายใหม่
                    ###### 2 ######
                    po = PurchaseOrder.objects.filter(distributor = old_distributor)
                    po.update(distributor = new_distributor_id)
                    
                except PurchaseOrder.DoesNotExist:
                    pass

                try:
                    #ย้ายผู้จัดจำหน่ายในการเปรียบเทียบผู้จัดจำหน่ายไปรหัสผู้จัดจำหน่ายใหม่
                    ###### 3 ######
                    cpd = ComparisonPriceDistributor.objects.filter(distributor = old_distributor)
                    cpd.update(distributor = new_distributor_id)
                    
                except ComparisonPriceDistributor.DoesNotExist:
                    pass

                try:
                    #ย้ายผู้จัดจำหน่ายในประเมินร้านค้าไปรหัสผู้จัดจำหน่ายใหม่
                    ###### 4 ######
                    rd = RateDistributor.objects.filter(distributor = old_distributor)
                    rd.update(distributor = new_distributor_id)
                    
                except RateDistributor.DoesNotExist:
                    pass

                #หลังจากย้ายผู้จัดจำหน่ายไปรหัสใหม่ ลบรหัสเก่า
                af_cp = ComparisonPrice.objects.filter(select_bidder__id = old_distributor.id).count()
                af_po = PurchaseOrder.objects.filter(distributor__id = old_distributor.id).count()
                af_cpd = ComparisonPriceDistributor.objects.filter(distributor__id = old_distributor.id).count()
                af_rd = RateDistributor.objects.filter(distributor__id = old_distributor.id).count()

                new_distributor = Distributor.objects.get(id = new_distributor_id)
                if af_cp < 1 and af_po < 1 and af_cpd < 1 and af_rd < 1:
                    old_distributor.delete()

                    new_distributor.name =  new_distributor.name[:-3]
                    new_distributor.save()

                    alertStr = "new โอน รหัสผู้จัดจำหน่าย '"+ str(old_distributor_id) + "' เป็น '"+ str(new_distributor_id)+ "' สำเร็จแล้ว"
                    alertType = messages.SUCCESS
                    
        except Distributor.DoesNotExist:
            alertStr = "ไม่มีรหัสผู้จัดจำหน่ายเก่าในระบบ ไม่สามารถ new โอนได้"
            alertType = messages.WARNING
            pass
    return alertType,alertStr

def days_between(d1, d2):
    d1 = datetime.datetime.strptime(str(d1), "%Y-%m-%d").date()
    d2 = datetime.datetime.strptime(str(d2), "%Y-%m-%d").date()
    return abs((d2 - d1).days)

def days_between_nagative(d1, d2):
    d1 = datetime.datetime.strptime(str(d1), "%Y-%m-%d").date()
    d2 = datetime.datetime.strptime(str(d2), "%Y-%m-%d").date()
    return (d2 - d1).days

def years_between(d1, d2):
    d1 = datetime.datetime.strptime(str(d1), "%Y-%m-%d").date()
    d2 = datetime.datetime.strptime(str(d2), "%Y-%m-%d").date()
    return abs(d2.year - d1.year)

def convertDateBEtoBC(strDateBE):
    strYear = str(int(strDateBE[:4]) - 543)
    strDateBC = strYear + "-" + strDateBE[5:7] + "-" + strDateBE[8:10]
    return strDateBC

def calculateDurationRate(receive_update, due_receive_update):
    dateDelay = days_between_nagative(receive_update, due_receive_update)
    duration_rate = None

    if dateDelay < -30:
        duration_rate = 1
    elif -30 <= dateDelay <= -8:
        duration_rate = 2
    elif -7 <= dateDelay <= -1:
        duration_rate = 3
    elif dateDelay >= 0:
        duration_rate = 4
    return duration_rate

def calculateTotalRate(price_rate, quantity_rate, duration_rate, service_rate, safety_rate):
    total_rate  = None
    total_rate = (price_rate*5) + (quantity_rate*5) + (duration_rate*5) + (service_rate*5) + (safety_rate*5)
    return total_rate

def calculateGradeRate(total_rate):
    grade_rate = None

    if 90 <= total_rate <= 100:
        grade_rate = 1
    elif 80 <= total_rate <= 89:
        grade_rate = 2
    elif 70 <= total_rate <= 79:
        grade_rate = 3
    elif 60 <= total_rate <= 69:
        grade_rate = 4
    elif total_rate < 60:
        grade_rate = 5

    return grade_rate

# Create your views here.
@login_required(login_url='signIn')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def index(request, category_slug = None):
    try:
        active = request.session['company_code']
    except:
        user_profile = UserProfile.objects.get(user = request.user.id)
        company = BaseBranchCompany.objects.filter(userprofile = user_profile).first()
        active = company.code
        request.session['company'] = company.affiliated.name_sh
        request.session['company_code'] = company.code
    '''
    catalogy show สินค้า
    '''

    products = None
    categories_page = None
    if category_slug != None :
        categories_page = get_object_or_404(Category, slug = category_slug) #เช็ค 404 คือเช็คว่ามีค่ามาไหม
        #products = Product.objects.all().filter(category = categories_page, available=True, affiliated =  company.affiliated) #ดึงข้อมูล Product จาก db ทั้งหมดโดยที่ available=True
        products = Product.objects.all().filter(category = categories_page, available=True)
    else :
        #products = Product.objects.all().filter(available=True, affiliated =  company.affiliated) #ดึงข้อมูล Product จาก db ทั้งหมดโดยที่ available=True
        products = Product.objects.all().filter(available=True)
    
    paginator = Paginator(products,20) #หมายถึงให้แสดงสินค้า 4 ต่อ 1 หน้า
    try:
        page = int(request.GET.get('page', '1')) #กำหนดหมายเลขหน้าแรกเปน str แล้วแปลงเปน int
    except:
        page = 1 #กำหนดหมายเลขหน้าเริ่มแรกเปน 1

    #กำหนดจำนวนชิ้นต่อหน้า
    try:
        productperPage = paginator.page(page)
    except (EmptyPage, InvalidPage):
        productperPage = paginator.page(paginator.num_pages)

    #หา id express
    baseProduct = Product.objects.all().values('id', 'name')

    #หา id express
    baseDistributor = Distributor.objects.all().values('id', 'name')

    # button New โอน
    if request.method == 'POST' and 'btnformNewOld' in request.POST:
        old_product_id = request.POST.get('old_product_id', False)
        new_product_id = request.POST.get('new_product_id', False)
        mess = newOld(old_product_id, new_product_id)
        messages.add_message(request, mess[0], mess[1])
        return redirect('home')
    
    # button Change Exist
    if request.method == 'POST' and 'btnformChangeExist' in request.POST:
        ce_old_product_id = request.POST.get('ce_old_product_id', False)
        ce_new_product_id = request.POST.get('ce_new_product_id', False)
        mess = changeExist(ce_old_product_id, ce_new_product_id)
        messages.add_message(request, mess[0], mess[1])
        return redirect('home')
    
    # button New โอน
    if request.method == 'POST' and 'btnformNewOldDistributor' in request.POST:
        old_distributor_id = request.POST.get('old_distributor_id', False)
        new_distributor_id = request.POST.get('new_distributor_id', False)
        mess = newOldDistributor(old_distributor_id, new_distributor_id)
        messages.add_message(request, mess[0], mess[1])
        return redirect('home')

    #เช็คสิทธิเห็นปุ่ม New โอน หรือ Change Exist
    isEditNewOld = is_edit_new_old(request.user)
    isEditChangeExist = is_edit_change_exist(request.user)
    isEditNewOldDistributor = is_edit_new_old_distributor(request.user)

    '''
    #สร้าง page
    p = Paginator(products, 10)
    page = request.GET.get('page')
    productperPage = p.get_page(page)
    '''


    '''
    first page
    '''

    #หาหน้าต่างการมองเห็นบริษัททั้งหมดของ user
    try:
        company_in = findCompanyIn(request)
    except:
        company_in = BaseBranchCompany.objects.filter(code = active).values('code')

    #ใบขอซื้อ
    try:
        user_profile = UserProfile.objects.get(user_id = request.user.id)

        permiss = BasePermission.objects.filter(codename ='CAAPR')
        #แบบเก่า
        #permiss_pr = PositionBasePermission.objects.filter(position_id = user_profile.position_id, base_permission__codename='CAAPR').prefetch_related(Prefetch('base_permission', queryset=permiss))
        isPermiss_pr = PositionBasePermission.objects.filter(position_id = user_profile.position_id, base_permission__codename='CAAPR', branch_company__code__in = company_in).values('branch_company__code')
    except:
        isPermiss_pr  = False

    try:
        in_company = BaseBranchCompany.objects.filter(userprofile = user_profile, code__in = isPermiss_pr).exists()
    except:
        pass


    #ถ้าเป็นผู้อนุมัติ
    pr_item_ap = None
    if(isPermiss_pr and in_company):
        try:
            #ดึงข้อมูล PurchaseRequisition
            pr_item_ap = PurchaseRequisition.objects.filter(purchase_status = 2, approver_status = 1, approver_user = request.user, branch_company__code__in = isPermiss_pr) #หาสถานะรอดำเนินการของผู้อนุมัติ
            #หาความยาวของ index PurchaseRequisition ที่มี สถานะรอดำเนินการของผู้อนุมัติ
        except PurchaseRequisition.DoesNotExist:
            pass

    new_pr = dict()
    if pr_item_ap:
        for obj in pr_item_ap:
            if obj not in new_pr:
                new_pr[obj] = obj

    '''เปลี่ยน flow ผู้มีสิทธิขอซื้อไม่ต้องกดอนุมัติ
    try:
        #ดึงข้อมูล PurchaseRequisition
        pr_item_pc = PurchaseRequisition.objects.filter(purchase_user = request.user.id, purchase_status = 1, branch_company__code__in = isPermiss_pr) #หาสถานะรอดำเนินการของผู้อนุมัติ
    except PurchaseRequisition.DoesNotExist:
        pr_item_pc = None

    if pr_item_pc:
        for obj in pr_item_pc:
            if obj not in new_pr:
                new_pr[obj] = obj    
    '''

    #ใบสั่งซื้อ
    try:
        user_profile = UserProfile.objects.get(user_id = request.user.id)

        permiss = BasePermission.objects.filter(codename ='CAAPO')
        isPermiss_po = PositionBasePermission.objects.filter(position_id = user_profile.position_id, base_permission__codename='CAAPO', branch_company__code__in = company_in).values('branch_company__code')
    except:
        isPermiss_po  = False

    po_item = None
    #ถ้าเป็นผู้อนุมัติที่มีสิทธิ
    if isPermiss_po:
        try:
            #ดึงข้อมูล PurchaseOrder
            po_item = PurchaseOrder.objects.all().filter(approver_status = 1, amount__isnull = False, amount__gt = 0, approver_user__isnull = True, cp__isnull = True, branch_company__code__in = isPermiss_po) #หาสถานะรอดำเนินการของผู้อนุมัติ
        except PurchaseOrder.DoesNotExist:
            pass

    new_po = dict()
    if po_item:
        for obj in po_item:
            if obj not in new_po:
                new_po[obj] = obj

    cm_po_item = None
    #เคสที่ดึงมาจากใบเปรียบเทียบมีชื่อคนอนุมัติอยู่แล้ว
    if request.user.is_authenticated:
        try:
            #ดึงข้อมูล PurchaseOrder
            cm_po_item = PurchaseOrder.objects.all().filter(approver_status = 1, amount__isnull = False, amount__gt = 0, approver_user = request.user, branch_company__code__in = company_in) #หาสถานะรอดำเนินการของผู้อนุมัติ
        except PurchaseOrder.DoesNotExist:
            cm_po_item = None
    
    if cm_po_item:
        for obj in cm_po_item:
            if obj not in new_po:
                new_po[obj] = obj

    ''' อนุมัติใบเปรียบเทียบแบบเก่าเปลี่ยนเป็นแบบ fix ชื่อ
    #ผู้ตรวจสอบ
    try:
        user_profile = UserProfile.objects.get(user_id = request.user.id)
        permiss = BasePermission.objects.filter(codename__in= ['CAECP1','CAECP2','CAECP3','CAECP4','CAECPD','CAECPA'])
        isPermissAE = PositionBasePermission.objects.filter(position_id = user_profile.position_id, base_permission__codename__in= ['CAECP1','CAECP2','CAECP3','CAECP4','CAECPD','CAECPA']).prefetch_related(Prefetch('base_permission', queryset=permiss)).values('base_permission')
    except:
        isPermissAE = None

    pmAE = []
    if(isPermissAE and in_company):
        for i in isPermissAE:
            obj = BasePermission.objects.get(id = i['base_permission'])
            pmAE.append(obj)

    #ผู้อนุมัติ
    try:
        permiss = BasePermission.objects.filter(codename__in= ['CAACP1','CAACP2','CAACP3','CAACP4','CAACPD','CAACPA'])
        isPermissAA = PositionBasePermission.objects.filter(position_id = user_profile.position_id, base_permission__codename__in= ['CAACP1','CAACP2','CAACP3','CAACP4','CAACPD','CAACPA']).prefetch_related(Prefetch('base_permission', queryset=permiss)).values('base_permission')
    except:
        isPermissAA = None

    pmAA = []
    if(isPermissAA and in_company):
        for i in isPermissAA:
            obj = BasePermission.objects.get(id = i['base_permission'])
            pmAA.append(obj)

    ecm_item = []
    #ถ้าเป็นผู้ตรวจสอบ
    if(isPermissAE and in_company):
        try:
            for ae in pmAE:
                if ae.codename == 'CAECPD':
                    obj = ComparisonPriceDistributor.objects.filter( cp__examiner_status = 1, cp__cm_type = 1, cp__select_bidder_id__isnull = False, cp__branch_company__code__in = company_in).values("cp")
                elif ae.codename == 'CAECPA':
                    obj = ComparisonPriceDistributor.objects.filter( cp__examiner_status = 1, cp__cm_type = 2, cp__select_bidder_id__isnull = False, cp__branch_company__code__in = company_in).values("cp")
                else:
                    obj = ComparisonPriceDistributor.objects.filter( cp__examiner_status = 1, cp__select_bidder_id__isnull = False, is_select = True, amount__range=(ae.ap_amount_min, ae.ap_amount_max), cp__cm_type_id__isnull = True, cp__branch_company__code__in = company_in).values("cp")
                ecm_item.append(obj)
        except ComparisonPriceDistributor.DoesNotExist:
            ecm_item = None

    ecp = []
    new_cm = dict()
    try:
        for item in ecm_item:
            ecp = ComparisonPrice.objects.filter(pk__in = item).distinct()
            if ecp:
                for obj in ecp:
                    if obj not in new_cm:
                        new_cm[obj] = obj
    except:
        pass


    acm_item = []
    #ถ้าเป็นผู้อนุมัติ
    if(isPermissAA and in_company):
        try:
            for aa in pmAA:
                if aa.codename == 'CAACPD':
                    obj = ComparisonPriceDistributor.objects.filter( cp__examiner_status = 2,cp__approver_status = 1, cp__cm_type = 1, cp__select_bidder_id__isnull = False, cp__branch_company__code__in = company_in).values("cp")
                elif aa.codename == 'CAACPA':
                    obj = ComparisonPriceDistributor.objects.filter( cp__examiner_status = 2,cp__approver_status = 1, cp__cm_type = 2, cp__select_bidder_id__isnull = False, cp__branch_company__code__in = company_in).values("cp")
                else:
                    obj = ComparisonPriceDistributor.objects.filter( cp__examiner_status = 2,cp__approver_status = 1, cp__select_bidder_id__isnull = False, is_select = True, amount__range=(aa.ap_amount_min, aa.ap_amount_max), cp__cm_type_id__isnull = True, cp__branch_company__code__in = company_in).values("cp")
                acm_item.append(obj)
        except ComparisonPriceDistributor.DoesNotExist:
            acm_item = None

    acp = []
    try:
        for item in acm_item:
            acp = ComparisonPrice.objects.filter(pk__in = item).distinct()
            if acp:
                for obj in acp:
                    if obj not in new_cm:
                        new_cm[obj] = obj
    except:
        pass

    '''

    #ถ้าเป็นผู้ตรวจสอบ ใบเปรียบเทียบ
    ecm_item = []
    try:
        ecm_item = ComparisonPrice.objects.filter(examiner_status = 1, examiner_user = request.user, branch_company__code__in = company_in)
    except ComparisonPrice.DoesNotExist:
        ecm_item = None

    new_cm = dict()

    try:
        for obj in ecm_item:
            if obj not in new_cm:
                new_cm[obj] = obj
    except:
        pass


    #ถ้าเป็นผู้อนุมัติ ใบเปรียบเทียบ
    acm_item = []
    try:
        acm_item = ComparisonPrice.objects.filter(examiner_status = 2, approver_status = 1, approver_user = request.user, branch_company__code__in = company_in)
    except ComparisonPrice.DoesNotExist:
        acm_item = None
        
    try:
        for obj in acm_item:
            if obj not in new_cm:
                new_cm[obj] = obj
    except:
        pass

    #ถ้าเป็นผู้อนุมัติพิเศษ ใบเปรียบเทียบ
    try:
        user_profile = UserProfile.objects.get(user_id = request.user.id)

        permiss = BasePermission.objects.filter(codename ='CASCP')
        isPermiss_scp = PositionBasePermission.objects.filter(position_id = user_profile.position_id, base_permission__codename='CASCP', branch_company__code__in = company_in).values('branch_company__code', 'base_permission')
    except:
        isPermiss_scp  = None

    pmAA = []
    if(isPermiss_scp):
        for i in isPermiss_scp:
            obj = BasePermission.objects.get(id = i['base_permission'])
            pmAA.append(obj)

    cpd_item = []
    #ถ้าเป็นผู้อนุมัติที่มีสิทธิ
    if isPermiss_scp:
        for aa in isPermiss_scp:
            for pm in pmAA:
                try:
                    #ดึงข้อมูล PurchaseOrder
                    obj = ComparisonPriceDistributor.objects.filter( cp__examiner_status = 2,cp__approver_status = 2, cp__special_approver_status = 1, is_select = True, amount__range=(pm.ap_amount_min, pm.ap_amount_max), cp__cm_type_id__isnull = True, cp__branch_company__code = aa['branch_company__code']).values("cp")
                    cpd_item.append(obj)
                except ComparisonPriceDistributor.DoesNotExist:
                    pass
    
    scp = []
    try:
        for item in cpd_item:
            scp = ComparisonPrice.objects.filter(pk__in = item).distinct()
            if scp:
                for obj in scp:
                    if obj not in new_cm:
                        new_cm[obj] = obj
    except:
        pass

    '''
    scm_item = []
    try:
        scm_item = ComparisonPrice.objects.filter(examiner_status = 2, approver_status = 2, special_approver_status = 1, branch_company__code__in = company_in)
    except ComparisonPrice.DoesNotExist:
        scm_item = None
        
    try:
        for obj in scm_item:
            if obj not in new_cm:
                new_cm[obj] = obj
    except:
        pass
    
    '''
    try:
        user_profile = UserProfile.objects.get(user_id = request.user.id)
        visible_tab = BaseVisible.objects.filter(userprofile = user_profile)
    except:
        visible_tab = None
    
    isStaff = False
    isMobile = False
    isPermission = False

    ###################### Mobile Menu ##################
    #มีสิทธิอนุมัติ pr, cp หรือ po
    try:
        user_profile = UserProfile.objects.get(user_id = request.user.id)
        isPermission = PositionBasePermission.objects.filter(position_id = user_profile.position_id, branch_company__code__in = company_in).exists()
    except UserProfile.DoesNotExist or PositionBasePermission.DoesNotExist:
        pass

    for tab in visible_tab:
        if 'Category' in tab.name:
            isStaff = True
            break
        if 'Mobile' in tab.name:
            isMobile = True
            break

    #สังกัดบริษัทของคนนั้นๆ
    user_in_comp = None
    try:
        ucd = UserCarDepartment.objects.get(user = request.user.id)
        company = BaseBranchCompany.objects.get(code = ucd.in_comp.code)
        user_in_comp = company.code
    except:
        pass

    isApproveMA = False
    try:
        isApproveMA = ApproveCarDepartment.objects.filter(user = request.user.id).exists()
    except ApproveCarDepartment.DoesNotExist:
        pass

    isMobileMenu = False
    if isMobile and not isPermission and not isStaff:
        isMobileMenu = True

    if isStaff:
        return render(request,'index.html', {'products':productperPage , 'category':categories_page, 'baseProduct':baseProduct, 'isEditNewOld': isEditNewOld, 'isEditChangeExist': isEditChangeExist,'isEditNewOldDistributor': isEditNewOldDistributor, active :"active show", 'baseDistributor': baseDistributor, "colorNav":"enableNav"})
    elif isPermission:
        return render(request,'firstPage.html', {'prs':new_pr,'pos':new_po,'cms':new_cm, active :"active show", "colorNav":"enableNav", "isMobile": isMobile})
    elif isMobileMenu:
        return render(request, "mobileApp/menu.html", {"disableTab":"disableTab","colorNav":"disableNav", "user_in_comp": user_in_comp, "isApproveMA": isApproveMA, "isMobileMenu": isMobileMenu})


def productPage(request, category_slug, product_slug):
    try:
        product = Product.objects.get(category__slug = category_slug, slug = product_slug)
    except Exception as e:
        raise e
    return render(request,'product.html', {'product':product})

def _cart_id(request):
    cart = request.session.session_key
    if not cart:
        cart = request.session.create()
    return cart


@login_required(login_url='signIn') #ถ้า addCart แล้วยังไม่ได้ login ก็ให้ไปหน้า signIn
def addCart(request, product_id):
    #ส่งรหัสสินค้ามา
    #ดึงสินค้าตามรหัสที่ส่งมา
    product = Product.objects.get(id=product_id) #หา Product จาก id ที่ส่งมา
    #สร้างตะกร้าสินค้า
    try:
        cart = Cart.objects.get(cart_id=_cart_id(request)) #หาตะกร้าสินค้า
    except Cart.DoesNotExist:
        cart = Cart.objects.create(cart_id=_cart_id(request)) #ถ้าไม่มีตะกร้าสินค้าให้สร้างใหม่
        cart.save()

    #ซื้อสินค้าอันนี้คือกดเพิ่มสินค้า
    try:
        #ซื้อรายการสินค้าซ้ำ
        cart_item = CartItem.objects.get(product=product,cart=cart) #หาว่ามี cart_item หรือยัง
        if cart_item.quantity < cart_item.product.stock :
            #เพิ่มจำนวนสินค้า
            cart_item.quantity += 1
            #บันทึกลงฐานข้อมูล
            cart_item.save()
    except CartItem.DoesNotExist :
        #หา CartItem คือสั่งรายการครั้งแรก
        #สร้าง CartItem และ บันทึกลงฐานข้อมูล
        cart_item = CartItem.objects.create(
            product = product,
            cart = cart,
            quantity = 1,
        )
        cart_item.save()
    return redirect('cartdetail')


def cartdetail(request):
    total = 0 #ผลรวมสินค้าทั้งหมดในตะกร้า
    counter = 0 #จำนวนสินค้าทั้งหมด
    cart_items = None 

    try:
        cart = Cart.objects.get(cart_id = _cart_id(request)) #ดึงตะกร้า
        cart_items = CartItem.objects.filter(cart = cart, active = True) #ดึงข้อมูลสินค้าในตะกร้า
        for item in cart_items:
            total+=(item.product.price*item.quantity) #ผลรวมของสินค้าทุกชิ้น
            counter+=item.quantity #ผลรวมของจำนวนสินค้าทุกชิ้น
    except Exception as e:
        pass #ไม่ต้องทำไร

    stripe.api_key = settings.SECRET_KEY
    strip_total = int(total*100) #เช่น 800 strip มองเป็น 8 เลยต้อง  *100
    description = "Payment Online"
    data_key = settings.PUBLIC_KEY

    if request.method == "POST":
        try:
            #ชื่อจากใน POST เอามาจากปริ้นคำสั่ง print(request.POST) แล้ว link ฟิลด์กับ order ที่หน้า admin
            token = request.POST['stripeToken']
            email = request.POST['stripeEmail']
            name = request.POST['stripeBillingName']
            address = request.POST['stripeBillingAddressLine1']
            city = request.POST['stripeBillingAddressCity']
            postcode = request.POST['stripeShippingAddressZip']

            # print(request.POST)

            customer = stripe.Customer.create(
                email = email,
                source = token,
            )

            charge = stripe.Charge.create(
                amount = strip_total,
                currency = "thb",
                description = description,
                customer = customer.id
            )

            #บันทึกข้อมูลใบสั่งซื้อ (ดึงข้อมูลจาก request.POST ไป save ใน db order)
            order = Order.objects.create(
                name = name,
                address = address,
                city = city,
                postcode = postcode,
                total = total,
                email = email,
                token = token,
            )
            order.save()

            #บันทึกรายการสั่งซื้อ
            #ใน cart_items มีสินค้าเช่น รองเท้า, กระเป๋า เลยต้องวนลูปทุกตัว
            for item in cart_items:
                order_item = OrderItem.objects.create(
                    product = item.product.name,
                    quantity = item.quantity,
                    price = item.product.price,
                    order = order,
                )
                order_item.save()

                #ลดจำนวน stock หลังจากที่จ่ายเงินแล้ว
                product = Product.objects.get(id = item.product.id) #หา product จาก product id
                product.stock = int(item.product.stock - item.quantity) #จำนวนสินค้าในสต็อก = ในสต็อก - ที่ซื้อ
                product.save() #บันทึก

                item.delete() #ล้างตะกร้าสินค้า
            return redirect('thankyou')

        except stripe.error.CardError as e:
            return False, e

    return render(request,'cartdetail.html',
    dict(cart_items=cart_items, total = total, counter = counter,
    data_key = data_key, strip_total = strip_total, description = description))

def removeCart(request, product_id):
    #หา cart ที่เลือกอยู๋ A
    cart = Cart.objects.get(cart_id = _cart_id(request))
    #หาสินค้าที่เลือกจะลบ 2
    product = get_object_or_404(Product,id = product_id)

    #ลบรายการที่มี product และ cart ตรงกับที่เลือกอยู๋
    cartItem = CartItem.objects.get(product = product, cart = cart)
    #ลบรายการสินค้า 2 ที่อยู่ในตะกร้า A
    cartItem.delete()
    return redirect('cartdetail')

def signUpView(request):
    #ถ้าส่งข้อมูลจาก form signup มา
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        profile_form = UserProfileForm(request.POST)
        if form.is_valid() and profile_form.is_valid():
            #บันทึกข้อมูล User
            user = form.save()

            profile = profile_form.save(commit=False)
            profile.user = user

            profile.save()
            #บันทึก Group Customer
            #ดึง username จากฟอร์มมาใช้
            username = form.cleaned_data.get('username')
            #ดึง user จาก db มาใช้
            signUpUser = User.objects.get(username = username)
            #จัด Group
            #customer_group = Group.objects.get(name = "Customer")
            #customer_group.user_set.add(signUpUser)
            return redirect('signIn')

    else:
        form = SignUpForm()
        profile_form = UserProfileForm()
    context = {'form':form, 'profile_form':profile_form}
    return render(request, "signup.html", context)

def signInView(request):
    if request.method == 'POST':
        form = AuthenticationForm(data = request.POST)
        if form.is_valid():
            username = request.POST['username']
            password = request.POST['password']
            user = authenticate(username=username,password=password)
            #ถ้าล็อกอินสำเร็จไปหน้า home else ให้ไปสมัครใหม่
            if user is not None:
                login(request, user)
                try:
                    user_profile = UserProfile.objects.get(user = request.user.id)
                    company = BaseBranchCompany.objects.filter(userprofile = user_profile).first()
                except:
                    company = None
                request.session['company_code'] = company.code
                request.session['company'] = company.name
                return redirect('home')
            else:
                return redirect('signUp')
    else:
        form = AuthenticationForm()
        company = BaseBranchCompany.objects.first()
        request.session['company_code'] = company.code
        request.session['company'] = company.name

    return render(request, 'signin.html', {'form':form,})

def signOutView(request):
    logout(request)
    return redirect('signIn')

def search(request):
    active = request.session['company_code']

    name = request.GET['title']
    products = Product.objects.filter(Q(name__icontains=name) | Q(id__icontains=name))
    return render(request,'index.html', {'products':products,active :"active show","colorNav":"enableNav"})

def orderHistory(request):
    #ดึง order ทั้งหมดของ user คนนี้
    if request.user.is_authenticated:
        email = str(request.user.email) #หา order จาก email เพราะ  email ไม่มีทางซ้ำกันอยู่แล้ว
        orders = Order.objects.filter(email=email)#ดึง order ทั้งหมด
    return render(request, 'orders.html', {'orders':orders})

def viewOrder(request, order_id):
    #ดึง order ที่เลือกของ user คนนี้
    if request.user.is_authenticated:
        email = str(request.user.email) #หา order จาก email เพราะ  email ไม่มีทางซ้ำกันอยู่แล้ว
        order = Order.objects.get(email=email, id=order_id) #ดึง order ที่เลือก
        order_items = OrderItem.objects.filter(order= order) #หา OrderItem
    return render(request, 'viewOrder.html', {'order':order, 'order_items': order_items})

def thankyou(request):
    return render(request, 'thankyou.html')

def requisitionPageMode(mode):
    if mode == 1:
        return 'requisitions_page'
    elif mode == 2:
        return 'ap_pr_page'
    elif mode == 4:
        return 'h_requisitions_page'
    elif mode == 5:
        return 'h_i_pr_page'

def requisitionShowMode(mode):
    if mode == 1:
        return 'requisitions_show'
    elif mode == 2:
        return 'ap_pr_show'
    elif mode == 4:
        return 'h_requisitions_show'
    elif mode == 5:
        return 'h_i_pr_show'

def PRPageMode(mode):
    if mode == 1:
        return 'pr_page'
    elif mode == 2:
        return 'ap_pr_page'
    elif mode == 3:
        return 'rc_page'
    elif mode == 4:
        return 'h_pr_page'
    elif mode == 5:
        return 'h_i_pr_page'

def PRShowMode(mode):
    if mode == 1:
        return 'pr_show'
    elif mode == 2:
        return 'ap_pr_show'
    elif mode == 3:
        return 'rc_show'
    elif mode == 4:
        return 'h_pr_show'
    elif mode == 5:
        return 'h_i_pr_show'

def CPPageMode(mode):
    if mode == 1:
        return 'cp_page'
    elif mode == 2:
        return 'ap_cp_page'
    elif mode == 3:
        return 'rc_page'
    elif mode == 4:
        return 'h_cp_page'
    elif mode == 5:
        return 'h_i_cp_page'

def CPShowMode(mode):
    if mode == 1:
        return 'cp_show'
    elif mode == 2:
        return 'ap_cp_show'
    elif mode == 3:
        return 'rc_show'
    elif mode == 4:
        return 'h_cp_show'
    elif mode == 5:
        return 'h_i_cp_show'

def POPageMode(mode):
    if mode == 1:
        return 'po_page'
    elif mode == 2:
        return 'ap_po_page'
    elif mode == 3:
        return 'rc_page'
    elif mode == 4:
        return 'h_po_page'
    elif mode == 5:
        return 'h_i_po_page'

def POShowMode(mode):
    if mode == 1:
        return 'po_show'
    elif mode == 2:
        return 'ap_po_show'
    elif mode == 3:
        return 'rc_show'
    elif mode == 4:
        return 'h_po_show'
    elif mode == 5:
        return 'h_i_po_show'

def IVPageMode(mode):
    if mode == 1:
        return 'iv_page'
    elif mode == 4:
        return 'h_ex_iv_page'

def IVShowMode(mode):
    if mode == 1:
        return 'iv_show'
    elif mode == 4:
        return 'h_ex_iv_show'
    
def MAPageMode(mode):
    if mode == 1:
        return 'ma_page'
    elif mode == 4:
        return 'h_ma_page'

def MAShowMode(mode):
    if mode == 1:
        return 'ma_show'
    elif mode == 4:
        return 'h_ma_show'

@login_required(login_url='signIn')
@permission_required('auth.view_user', login_url='requisition')
def requisitionAll(request):
    data = Requisition.objects.all()
    myFilter = RequisitionFilter(request.GET, queryset = data)
    data = myFilter.qs
    return render(request,'requisitionAll.html',{'requisitions':data, 'filter':myFilter})

def requisitionItemAll(request, requisition_id):
    #ดึงใบเบิกของ
    data = Requisition.objects.get(id = requisition_id)
    #ดึงสินค้าในใบเบิกของ
    items = RequisitionItem.objects.filter(requisition_id=requisition_id)

    # form สินค้าในใบเบิกของ
    form = RequisitionItemForm(request.POST or None, initial={'requisition_id': requisition_id})
    if form.is_valid():
        form.save()
        return redirect('createRequisitionItem', requisition_id = requisition_id)

def findStatusApprove(request, id, status):
    try:
        user = User.objects.get(id = id)
        if(status != 'รออนุมัติ'):
            approve_name = user.first_name + " " + user.last_name
        else:
            approve_name = ""
    except:
        approve_name = ""
    return approve_name

#ใช้ได้เหมือนกันกับ  CrudView 
def crudRequisitionItemView(request, requisition_id):
    users = RequisitionItem.objects.filter(requisition_id = requisition_id)
    requisition = Requisition.objects.get(id = requisition_id)

    user_login = User.objects.get(id = request.user.id)

    #หาชื่อหัวหน้างาน
    chief_approve_name = findStatusApprove(request, requisition.chief_approve_user_id, requisition.chief_approve_status)
    #หาชื่อเจ้าหน้าที่พัสดุ
    supplies_approve_name = findStatusApprove(request, requisition.supplies_approve_user_id, requisition.supplies_approve_status)
    
    context = {
        'users': users,
        'requisition': requisition,
        'chief_approve_name': chief_approve_name,
        'supplies_approve_name': supplies_approve_name,
    }
    if request.method == 'POST':
        status = request.POST['status'] or None
        obj = Requisition.objects.get(id = requisition_id)
        if(user_login.has_perm('auth.change_user')):
            obj.chief_approve_status = status
            obj.chief_approve_user_id = request.user.id
        if(user_login.has_perm('auth.add_user')):
            obj.supplies_approve_status = status
            obj.supplies_approve_user_id = request.user.id
        obj.save()
        return redirect('crudRequisitionItemView', requisition_id = requisition_id)
    
    return render(request, "requisitionItemDetail.html", context)
    
@login_required(login_url='signIn')
def requisition(request):
    active = request.session['company_code']

    month = datetime.datetime.now().month
    year = datetime.datetime.now().year
    '''
    data = Requisition.objects.filter(Q(created__year__gte = year,
                              created__month__gte= month,
                              created__year__lte = year,
                              created__month__lte = month) | Q(purchase_requisition_id__isnull = True))   
    '''
    data = Requisition.objects.filter(Q(invoice_id__isnull = True) & Q(purchase_requisition_id__isnull = True) & Q(is_edit = True), branch_company__code = active)

    #กรองข้อมูล
    myFilter = RequisitionFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)

    context = {
            'requisitions':dataPage,
            'filter':myFilter,
            'requisitions_page': "tab-active",
            'requisitions_show': "show",
            active :"active show",
            "colorNav":"enableNav"
    }

    return render(request,'requisition/viewRequisition.html', context)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def createRequisition(request):
    active = request.session['company_code']
    company = BaseBranchCompany.objects.get(code = active)
    #requestName = User.objects.all()
    #chiefName = User.objects.filter(groups__name='หัวหน้างาน')
    
    try:
        form = RequisitionForm(request, request.POST or None, initial={'branch_company': company})
        if form.is_valid():
                form = RequisitionForm(request, request.POST or None, request.FILES)
                new_contact = form.save(commit=False)
                new_contact.supplies_approve_user_name_id = request.user.id

                try:
                    userProfile = UserProfile.objects.get(user = new_contact.name)
                    new_contact.section = userProfile.department
                except UserProfile.DoesNotExist:
                    pass
                new_contact.save()

                #อัพเดทประเภทการซ่อม ในใบแจ้งซ่อม
                try:
                    mas = Maintenance.objects.filter(id = new_contact.ma_id)
                    for i in mas:
                        i.repair_type = new_contact.repair_type
                        i.save()
                except Maintenance.DoesNotExist:
                    pass
                
                # บันทึก ManyToMany relationships
                form.save_m2m()
                return HttpResponseRedirect(reverse('crud_ajax', args=(new_contact.pk,)))
    except ValueError:
        pass

    context = {
        'form':form,
        #'requestName':requestName,
        #'chiefName':chiefName,
        'requisitions_page': "tab-active",
        'requisitions_show': "show",
        active :"active show",
        "disableTab":"disableTab",
        "colorNav":"disableNav"
    }
    return render(request, "requisition/createRequisition.html", context)

def removeRequisition(request, id):
    requisition = Requisition.objects.get(id = id)
    #ลบ item ใน Requisition ด้วย
    items = RequisitionItem.objects.filter(requisition_id = id)
    items.delete()
    #ลบ Requisition ทีหลัง
    requisition.delete()
    return redirect('requisition')

def editRequisition(request, id):
    requisition = Requisition.objects.get(id=id)
    form = RequisitionForm(request, instance=requisition)

    if request.method == 'POST':
        form = RequisitionForm(request, request.POST, request.FILES, instance=requisition)
        if form.is_valid():
            new_contact = form.save()
            try :
                obj = PurchaseRequisition.objects.get(requisition = new_contact.pk)
                obj.purchase_user_id = requisition.chief_approve_user_name
                obj.save()
            except PurchaseRequisition.DoesNotExist :
                pass
            return redirect('requisition')

    context = {
        'form':form,
        'requisitions_page': "tab-active",
        'requisitions_show': "show",
    }
    return render(request, "requisition/createRequisition.html", context)

def createRequisitionItem(request, requisition_id):
    template_name = 'requisitionItem/createRequisitionItem.html'
    heading_message = 'Model Formset Demo'

    active = request.session['company_code']
    bc = BaseBranchCompany.objects.get(code = active)

    requisition = Requisition.objects.get(id=requisition_id)

    formset = RequisitionItemModelFormset(request.POST or None, queryset=RequisitionItem.objects.none())
    if request.method == 'POST':
        formset = RequisitionItemModelFormset(request.POST or None)
        if formset.is_valid():
            # save po item
            for form in formset:
                # only save if name is present
                if form.cleaned_data.get('product_name'):
                    obj = form.save(commit=False)
                    obj.requisition_id = requisition.id
                    obj.requisit = requisition
                    obj.quantity_pr = int(obj.quantity) - int(obj.quantity_take)
                    obj.save()
            #redirect นอก loop
            return redirect('viewPO')

    context = {
        'requisition':requisition,
        'po_page': "tab-active",
        'po_show': "show",
        'bc': bc,
        'formset': formset,
        'heading': heading_message,
    }
    return render(request, template_name, context)

def removeRequisitionItem(request, item_id):
    item = RequisitionItem.objects.get(id = item_id)
    requisition = Requisition.objects.get(id = item.requisition_id)
    item.delete()
    return redirect('createRequisitionItem', requisition_id = requisition.id)

def editRequisitionItem(request, item_id):
    item = get_object_or_404(RequisitionItem, id = item_id)
    requisit = Requisition.objects.get(id = item.requisition_id)
    item_all = RequisitionItem.objects.filter(requisition_id = item.requisition_id)
    form = RequisitionItemForm(instance=item )

    if request.method == 'POST':
        form = RequisitionItemForm(request.POST, instance=item )
        return redirect('createRequisitionItem', requisition_id = requisit.id)
    else:
        form = RequisitionItemForm(instance=item)
        return redirect('createRequisitionItem', requisition_id = requisit.id)

def createInvoiceAndItem(rqs, obj):
    #ถ้ามี invoice_code และเป็นหน่วยงานภายใน (agency.id = 1) ให้สร้าง invoice
    if rqs.branch_company.invoice_code and float(obj.quantity_take) > 0 and rqs.agency.id == 1:

        have_iv_id = Invoice.objects.filter(requisit = rqs).first()
        if not have_iv_id:
            # ไม่ Invoice create invoice
            iv = Invoice.objects.create(
                bring_name = rqs.name,
                payer_name = rqs.supplies_approve_user_name,
                car = rqs.car,
                note = rqs.note,
                expense_dept = rqs.expense_dept,
                branch_company = rqs.branch_company,
                requisit = rqs
            )
            iv.save()
            # set have_iv_id
            have_iv_id = iv

        #create invoice item
        iv_item = InvoiceItem.objects.create(
            iv = have_iv_id,
            item = obj,
            quantity = obj.quantity_take,
            unit = obj.unit
        )
        iv_item.save()

def updateInvoice(rqs):
    #update invoid
    try:
        iv = Invoice.objects.get(requisit = rqs)
        iv.bring_name = rqs.name
        iv.payer_name = rqs.supplies_approve_user_name
        iv.car = rqs.car
        iv.note = rqs.note
        iv.expense_dept = rqs.expense_dept
        iv.save()
    except Invoice.DoesNotExist:
        pass

def updateInvoiceAndItem(rqs, obj):
    #update invoid
    updateInvoice(rqs)

    iv = Invoice.objects.get(requisit = rqs)
    #update invoid item
    if float(obj.quantity_take) > 0:
        try:
            #มีแล้ว update
            iv_item = InvoiceItem.objects.get(item = obj)
            iv_item.quantity = obj.quantity_take
            iv_item.unit = obj.unit
            iv_item.save()
        except InvoiceItem.DoesNotExist:
            #ยังไม่มีให้ create
            iv_item = InvoiceItem.objects.create(
                iv = iv,
                item = obj,
                quantity = obj.quantity_take,
                unit = obj.unit
            )
            iv_item.save()
    else:
        #obj.quantity_take = 0 ลบ item
        try:
          item =  InvoiceItem.objects.get(item = obj)
          item.delete()
        except InvoiceItem.DoesNotExist:
            pass

def CUInvoiceAndItem(request, rq_id):
    iv = None
    rqs = Requisition.objects.get(id = rq_id)
    have_iv_id = Invoice.objects.filter(requisit = rq_id).first()

    have_item = RequisitionItem.objects.filter(requisit = rqs).exists()#มีสินค้าในใบขอเบิกไหม
    have_take_item = RequisitionItem.objects.filter(requisit = rqs, quantity_take__gt = 0).exists()#มีจำนวนที่จ่าย
    have_pr_item = RequisitionItem.objects.filter(requisit = rqs, quantity_pr__gt = 0).exists()#มีจำนวนที่ซื้อ

    #ถ้ามีจำนวนที่ต้องจ่าย และ invoice_code และเป็นหน่วยงานภายใน (agency.id = 1) ให้สร้าง invoice
    if have_take_item and rqs.branch_company.invoice_code and rqs.agency.id == 1:
        if have_iv_id:
            updateInvoice(rqs)
        else:
            #สร้าง invioce
            iv = Invoice.objects.create(
                bring_name = rqs.name,
                payer_name = rqs.supplies_approve_user_name,
                car = rqs.car,
                note = rqs.note,
                expense_dept = rqs.expense_dept,
                branch_company = rqs.branch_company,
                requisit = rqs
            )
            iv.save()

            #สร้าง invioce item
            r_items = RequisitionItem.objects.filter(requisit = rqs, quantity_take__gt = 0)
            for r_i in r_items:
                iv_item = InvoiceItem.objects.create(
                    iv = iv,
                    item = r_i,
                    quantity = r_i.quantity_take,
                    unit = r_i.unit
                )
                iv_item.save()

            #เก็บ invoice id ใน requisition
            rqs.invoice_id = iv.pk
            rqs.iv_ref_no = iv.ref_no

            if not have_pr_item:#ถ้าไม่มีจำนวนที่ต้องซื้อ ปิดใบเบิกเลย
                rqs.is_edit = False

            rqs.save()

    elif have_item and rqs.branch_company.invoice_code and rqs.agency.id == 2:#กรณีหน่วยงานภายนอก ไม่มีการออกใบจ่าย
        #ถ้าเป็นหน่วยงานภายนอกให้ลบ iv
        if have_iv_id:
          #ลบ iv item ก่อน
          items =  InvoiceItem.objects.filter(iv = have_iv_id)
          items.delete()

          #set invoice_id ใน requisition เป็น None
          rqs.invoice_id = None
          rqs.iv_ref_no = None

          #ลบ iv หลังจากนั้น
          have_iv_id.delete()
        
        if not have_pr_item:
            rqs.is_edit = False #ปิดใบเบิกเลย กรณีหน่วยงานภายนอกและไม่มีจำนวนที่ซื้อ
        rqs.save()

    if not have_item:#ถ้าไม่มีสินค้าในใบขอเบิก
        messages.error(request, "ไม่สามารถสร้างใบขอซื้อหรือใบจ่ายสินค้าภายในได้ เนื่องจากไม่มีสินค้าในใบขอเบิก")
        return redirect('preparePR')
    if have_pr_item: #มีจำนวนที่ต้องซื้อ เข้า createPR
        return HttpResponseRedirect(reverse('createPR', args=(rq_id,)))
    elif iv: #มีการจ่ายภายใน
        messages.success(request, "สร้างใบจ่ายสินค้าภายใน " + str(iv.ref_no)+ " เรียบร้อยแล้ว")
        return redirect('preparePR')
    else: #กรณีหน่วยงานภายนอก ไม่มีการออกใบขอซื้อหรือใบจ่าย
        messages.warning(request, "ใบขอเบิกนี้เป็นหน่วยงานภายนอกและไม่มีการออกใบขอซื้อ ข้อมูลเป็นประวัติใบขอเบิกเรียบร้อยแล้ว")
        return redirect('preparePR')
# class CrudView(PermissionRequiredMixin, TemplateView):
#    raise_exception = True 
#    permission_required = 'stock.cruduser.change_cruduser'

class CrudView(TemplateView):
    template_name = 'requisition/crud.html'
    def get_context_data(self, *args, **kwargs):
        active = self.request.session['company_code']

        bc = BaseBranchCompany.objects.get(code = active)
        context = super().get_context_data(*args,**kwargs)
        users = RequisitionItem.objects.filter(requisition_id = self.kwargs['requisition_id'])
        context['users'] = users
        requisition = Requisition.objects.get(id=self.kwargs['requisition_id'])
        context['requisition'] = requisition
        try:
            pr = PurchaseRequisition.objects.get(id = requisition.purchase_requisition_id)
        except:
            pr = ""

        #หากมีการเบิกสินค้านี้ ของทะเบียนรถหรือเครื่องจักรเดียวกัน ภายใน 30, 90, 365 
        distinct_products = {
            product : created 
            for product, created in RequisitionItem.objects.filter(
                ~Q(requisition_id = self.kwargs['requisition_id']),
                Q(requisit__rq_type__id=1) | Q(requisit__rq_type__id=2),
                requisit__created__lt=requisition.created,
                requisit__car=requisition.car,
                requisit__branch_company = requisition.branch_company
            ).values_list('product', 'requisit__created')
        }

        #set alert_level text ภายใน 30, 90, 365
        for item in users:
            created = distinct_products.get(item.product.id)
            if created:
                diff_days = (item.requisit.created - created).days
                if diff_days <= 30:
                    item.alert_level = "***"
                elif diff_days <= 90:
                    item.alert_level = "**"
                elif diff_days <= 365:
                    item.alert_level = "*"
                else:
                    item.alert_level = ""
            
        context['pr'] = pr
        context['baseUrgency'] = BaseUrgency.objects.all()
        context['baseUnit'] = BaseUnit.objects.all()
        context['baseProduct'] = Product.objects.all().values('id', 'name')
        context['requisitions_page'] = "tab-active"
        context['requisitions_show'] = "show"
        context['bc'] = bc
        context[active] = "active show"
        context['disableTab'] = "disableTab"
        context['colorNav'] = "disableNav"

        return context

class CreateCrudUser(View):

    def  get(self, request, *args, **kwargs):
        name1 = request.GET.get('name', None)
        description1 = request.GET.get('description', None)
        quantity1 = request.GET.get('quantity', None)
        quantityTake1 = request.GET.get('quantity_take', None)
        machine1 = request.GET.get('machine', None)
        desireddate1 = request.GET.get('desired_date', None) or None
        unit1 = request.GET.get('unit', None)
        urgency1 = request.GET.get('urgency', None)
        product1 = request.GET.get('product', None)
        
        try:
            product_item = get_object_or_404(Product, id=product1)
        except:
            product_item = None

        try:
            if(not quantityTake1):
                quantityTake1 = 0
            quantityPQ  = float(quantity1) - float(quantityTake1)
        except:
            quantityPQ = 0

        rqs = Requisition.objects.get(id=kwargs['requisition_id'])

        if not description1:
            description1 = rqs.repair_type.name
        if not machine1:
            machine1 = rqs.car.name + rqs.car.code #ใช้ในระบบงาน version ใหม่ดึงจาก car.name + car.code อันเก่าดึงจาก machine 10/10/2024
        if not desireddate1:
            desireddate1 = rqs.desired_date
        if not urgency1:
            urgency1 = rqs.urgency.id

        #08-07-2024 repair_type -> description, machine -> car
        obj = RequisitionItem.objects.create(
            requisition_id = kwargs['requisition_id'],
            product_name = name1,
            unit = unit1,
            quantity = quantity1,
            quantity_take = quantityTake1,
            quantity_pr = quantityPQ,
            requisit = rqs,
            product = product_item,
            description = description1,
            machine = machine1,
            desired_date = desireddate1,
            urgency = urgency1,
        )

        ''' 24/10/2024 เปลี่ยน flow ออกใบจ่ายสินค้า จากการกดสร้างใบขอซื้อและใบจ่ายสินค้าภายในแทน
        ############# create invoice ##############
        createInvoiceAndItem(rqs, obj)
        '''

        user = {'id':obj.id,'name':obj.product_name,'description':obj.description,
        'quantity':obj.quantity,'quantity_take':obj.quantity_take,'machine':obj.machine,'desired_date':obj.desired_date,
        'unit':obj.unit, 'urgency': obj.urgency}

        data = {
            'user': user
        }
        return JsonResponse(data)

class UpdateCrudUser(View):

    def  get(self, request, *args, **kwargs):
        id1 = request.GET.get('id', None)
        name1 = request.GET.get('name', None)
        description1 = request.GET.get('description', None)
        quantity1 = request.GET.get('quantity', None)
        quantityTake1 = request.GET.get('quantity_take', None)
        machine1 = request.GET.get('machine', None)
        desireddate1 = request.GET.get('desired_date', None)
        unit1 = request.GET.get('unit', None)
        urgency1 = request.GET.get('urgency', None)
        product1 = request.GET.get('product', None)

        try:
            product_item = get_object_or_404(Product, id=product1)
        except:
            product_item = None

        try:
            quantityPQ  = float(quantity1) - float(quantityTake1)
        except:
            quantityPQ = 0

        rqs = Requisition.objects.get(id=kwargs['requisition_id'])

        #08-07-2024 repair_type -> description, machine -> car
        obj = RequisitionItem.objects.get(id=id1)
        obj.requisition_id = kwargs['requisition_id']
        obj.product_name = name1
        obj.quantity = quantity1
        obj.quantity_take = quantityTake1
        obj.quantity_pr = quantityPQ
        obj.unit = unit1
        obj.product = product_item

        if description1:
            obj.description = description1
        else:
            obj.description = rqs.repair_type.name

        if machine1:
            obj.machine = machine1
        else:
            obj.machine = rqs.car.name + rqs.car.code #ใช้ในระบบงาน version ใหม่ดึงจาก car.name + car.code อันเก่าดึงจาก machine 10/10/2024

        if desireddate1:
            obj.desired_date = desireddate1
        else:
            obj.desired_date = rqs.desired_date
            
        if urgency1:
            obj.urgency = urgency1
        else:
            obj.urgency = rqs.urgency.id
        obj.save()

        ''' 24/10/2024 เปลี่ยน flow ออกใบจ่ายสินค้า จากการกดสร้างใบขอซื้อและใบจ่ายสินค้าภายในแทน
        ############# update invoice ##############
        have_iv = Invoice.objects.filter(requisit = rqs).exists()
        if have_iv:
            updateInvoiceAndItem(rqs, obj)
        else:
            CUInvoiceAndItem(rqs)
        '''

        user = {'id':obj.id,'name':obj.product_name,'description':obj.description,
        'quantity':obj.quantity,'quantity_take':obj.quantity_take,'machine':obj.machine,'desired_date':obj.desired_date,
        'unit':obj.unit, 'urgency':obj.urgency}

        data = {
            'user': user
        }
        return JsonResponse(data)

class DeleteCrudUser(View):
    def  get(self, request):
        id1 = request.GET.get('id', None)
        RequisitionItem.objects.get(id=id1).delete()
        data = {
            'deleted': True
        }
        return JsonResponse(data)

def editAllRequisition(request, requisition_id):
    active = request.session['company_code']
    bc = BaseBranchCompany.objects.get(code = active)
    users = RequisitionItem.objects.filter(requisition_id = requisition_id)
    requisition = Requisition.objects.get(id=requisition_id)
    try:
        pr = PurchaseRequisition.objects.get(id = requisition.purchase_requisition_id)
    except:
        pr = ""
            
    baseUrgency = BaseUrgency.objects.all()
    baseUnit = BaseUnit.objects.all()
    baseProduct = Product.objects.all().values('id', 'name')

    #หากมีการเบิกสินค้านี้ ของทะเบียนรถหรือเครื่องจักรเดียวกัน ภายใน 30, 90, 365
    distinct_products = {
        product : created 
        for product, created in RequisitionItem.objects.filter(
            ~Q(requisition_id=requisition_id),
            Q(requisit__rq_type__id=1) | Q(requisit__rq_type__id=2),
            requisit__created__lt=requisition.created,
            requisit__car=requisition.car,
            requisit__branch_company = requisition.branch_company
        ).values_list('product', 'requisit__created')
    }

    #set alert_level text ภายใน 30, 90, 365
    for item in users:
        created = distinct_products.get(item.product.id)
        if created:
            diff_days = (item.requisit.created - created).days
            if diff_days <= 30:
                item.alert_level = "***"
            elif diff_days <= 90:
                item.alert_level = "**"
            elif diff_days <= 365:
                item.alert_level = "*"
            else:
                item.alert_level = ""


    #form save
    form = RequisitionForm(request, instance=requisition)

    if request.method == 'POST':
        form = RequisitionForm(request, request.POST, request.FILES , instance=requisition)
        if form.is_valid():
            new_contact =  form.save()
            r_item = RequisitionItem.objects.filter(requisit = new_contact.pk)
            for i in r_item:
                i.description = new_contact.repair_type.name
                i.machine = new_contact.car.name
                i.desired_date = new_contact.desired_date
                i.urgency = new_contact.urgency.id
                i.save()
            try :
                obj = PurchaseRequisition.objects.get(requisition = new_contact.pk)
                obj.purchase_user_id = requisition.chief_approve_user_name
                obj.save()
            except PurchaseRequisition.DoesNotExist :
                pass

            #อัพเดทประเภทการซ่อม ในใบแจ้งซ่อม
            try:
                mas = Maintenance.objects.filter(id = new_contact.ma_id)
                for i in mas:
                    i.repair_type = new_contact.repair_type
                    i.save()
            except Maintenance.DoesNotExist:
                pass

            ''' 24/10/2024 เปลี่ยน flow ออกใบจ่ายสินค้า จากการกดสร้างใบขอซื้อและใบจ่ายสินค้าภายในแทน
            #crate or update invoid
            CUInvoiceAndItem(requisition)            
            '''

            return redirect('requisition')

    context = {
        'form':form,
        'users': users,
        'requisition': requisition,
        #'chiefName' : chiefName,
        'pr': pr,
        'baseUrgency': baseUrgency,
        'baseUnit': baseUnit,
        'baseProduct': baseProduct,
        #'requestName':requestName,
        'bc':bc,
        'requisitions_page':"tab-active",
        'requisitions_show':"show",
        active :"active show",
		"disableTab":"disableTab",
		"colorNav":"disableNav"
    }

    return render(request, "requisition/editAllRequisition.html", context)

def showRequisition(request, requisition_id, mode):
    active = request.session['company_code']
    bc = BaseBranchCompany.objects.get(code = active)

    items = RequisitionItem.objects.filter(requisition_id = requisition_id)
    requisition = Requisition.objects.get(id=requisition_id)
    try:
        pr = PurchaseRequisition.objects.get(id = requisition.purchase_requisition_id)
    except:
        pr = ""

    #หากมีการเบิกสินค้านี้ ของทะเบียนรถหรือเครื่องจักรเดียวกัน ภายใน 30, 90, 365
    distinct_products = {
        product : created 
        for product, created in RequisitionItem.objects.filter(
            ~Q(requisition_id=requisition_id),
            Q(requisit__rq_type__id=1) | Q(requisit__rq_type__id=2),
            requisit__created__lt=requisition.created,
            requisit__car=requisition.car,
            requisit__branch_company = requisition.branch_company
        ).values_list('product', 'requisit__created')
    }

    #set alert_level text ภายใน 30, 90, 365
    for item in items:
        created = distinct_products.get(item.product.id)
        if created:
            diff_days = (item.requisit.created - created).days
            if diff_days <= 30:
                item.alert_level = "***"
            elif diff_days <= 90:
                item.alert_level = "**"
            elif diff_days <= 365:
                item.alert_level = "*"
            else:
                item.alert_level = ""
            
    baseUrgency = BaseUrgency.objects.all()
    baseUnit = BaseUnit.objects.all()

    page = requisitionPageMode(mode)
    show = requisitionShowMode(mode)

    context = {
        'items': items,
        'requisition': requisition,
        'pr': pr,
        'baseUrgency': baseUrgency,
        'baseUnit': baseUnit,
        'bc':bc,
        page:"tab-active",
        show:"show",
        active :"active show",
		"disableTab":"disableTab",
		"colorNav":"disableNav"
    }

    return render(request, "requisition/showRequisition.html", context)


@login_required(login_url='signIn')
def viewPR(request):
    active = request.session['company_code']

    #ถ้า user login เป็นจัดซื้อ
    isPurchasing = is_purchasing(request.user)

    #ถ้าเป็นจัดซื้อให้ดึงมาเฉพาะที่ ผู้ขอซื้อ กับ ผู้อนุมัติ อนุมัติแล้ว
    if isPurchasing:
        data = PurchaseRequisition.objects.filter(purchase_status_id = 2, approver_status_id = 2, organizer = request.user , is_complete = 0, branch_company__code = active)
        #requisit = PurchaseRequisition.objects.filter(purchase_status_id = 2, approver_status_id = 2, organizer = request.user, is_complete = 0).values("requisition")
        #ri_not_used = RequisitionItem.objects.filter(is_used = False, quantity_pr__gt=0, requisit__in = requisit).defer('requisition_id', 'product_name', 'urgency')
    else:
        data = PurchaseRequisition.objects.filter(Q(purchase_status_id = 1) | Q(approver_status_id = 1) & ~Q(purchase_status_id = 3), is_complete = 0 , branch_company__code = active)
        #requisit = PurchaseRequisition.objects.filter(Q(purchase_status_id = 1) | Q(approver_status_id = 1) & ~Q(purchase_status_id = 3), is_complete = 0).values("requisition")
        #ri_not_used = RequisitionItem.objects.filter(is_used = False, quantity_pr__gt=0, requisit__in = requisit).defer('requisition_id', 'product_name', 'urgency')

    #กรองข้อมูล
    myFilter = PurchaseRequisitionFilter(request.GET, queryset = data)
    data = myFilter.qs

    baseUrgency = BaseUrgency.objects.all()
    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)

    context = {
                'prs':dataPage,
                'filter':myFilter,
                'isPurchasing':isPurchasing,
                'baseUrgency':baseUrgency,
                'pr_page': "tab-active",
                'pr_show': "show",
                active :"active show",
                "colorNav":"enableNav"
              }

    return render(request,'purchaseRequisition/viewPR.html',context)

def preparePR(request):
    active = request.session['company_code']
    requisitions = Requisition.objects.filter(Q(invoice_id__isnull = True) & Q(purchase_requisition_id__isnull = True) & Q(is_edit = True), branch_company__code = active)

    #สร้าง page
    paginator = Paginator(requisitions,10) #หมายถึงให้แสดงสินค้า 4 ต่อ 1 หน้า
    try:
        page = int(request.GET.get('page', '1')) #กำหนดหมายเลขหน้าแรกเปน str แล้วแปลงเปน int
    except:
        page = 1 #กำหนดหมายเลขหน้าเริ่มแรกเปน 1

    #กำหนดจำนวนชิ้นต่อหน้า
    try:
        dataPage = paginator.page(page)
    except (EmptyPage, InvalidPage):
        dataPage = paginator.page(paginator.num_pages)

    context = {
        'requisitions':dataPage,
        'pr_page': "tab-active",
        'pr_show': "show",
        active :"active show",
        "colorNav":"enableNav"
    }
    return render(request,'purchaseRequisition/preparePR.html', context)

def is_purchasing(user):
    return user.groups.filter(name='จัดซื้อ').exists()

def is_supplies(user):
    return user.groups.filter(name='พัสดุ').exists()

def is_edit_po_id(user):
    return user.groups.filter(name='แก้ไขรหัสใบสั่งซื้อ').exists()

def is_edit_approver_user_po(user):
    return user.groups.filter(name='แก้ไขผู้อนุมัติใบสั่งซื้อ').exists()

#ถ้ามีสิทธิดูรายงานของบริษัททั้งหมด
def is_view_report_all(user):
    return user.groups.filter(name='ดูรายงานของบริษัททั้งหมด').exists()

def is_edit_new_old(user):
    return user.groups.filter(name='NewOld').exists()

def is_edit_change_exist(user):
    return user.groups.filter(name='ChangeExist').exists()

def is_edit_new_old_distributor(user):
    return user.groups.filter(name='NewOldDistributor').exists()

def is_del_iv(user):
    return user.groups.filter(name='DeleteInvoice').exists()

def is_edit_pr(user):
    return user.groups.filter(name='EditPR').exists()

def is_change_car_ma(user):
    return user.groups.filter(name='ChangeCarMA').exists()

def is_re_pr(user):
    return user.groups.filter(name='rePR').exists()

def is_special_approver_cp(cp_id):
    bp = BasePermission.objects.get(codename='CASCP')
    return ComparisonPriceDistributor.objects.filter(cp = cp_id, is_select = True, amount__range=(bp.ap_amount_min, bp.ap_amount_max), cp__cm_type_id__isnull = True).exists()

def searchExaminerAndApproverUser(request):
    if 'select_bidder_id' in request.GET and 'cp_id' in request.GET:
        select_bidder_id = request.GET.get('select_bidder_id')
        cp_id = request.GET.get('cp_id')
        cm_type =  request.GET.get('cm_type')

        cpd = ComparisonPriceDistributor.objects.get(distributor__id = select_bidder_id, cp = cp_id)

        user_examiner =  findExaminerUserComparisonPrice(request, cpd.id, cm_type)
        user_approve = findApproveUserComparisonPrice(request, cpd.id, cm_type)
        
    data = {
        'user_examiner_list': list(user_examiner),
        'user_approve_list': list(user_approve),
    }
    return JsonResponse(data)

def searchRepairType(request):
    if 'rq_type' in request.GET:
        rq_type = request.GET.get('rq_type')

        base_repair_type = BaseRepairType.objects.filter(Q(rq_type = rq_type)| Q(rq_type = 4)).values('id', 'name')
        
    data = {
        'repair_type_list': list(base_repair_type),
    }
    return JsonResponse(data)

def searchExpenseDept(request):
    if 'agency' in request.GET:
        agency = request.GET.get('agency')

        base_expense_dep =  BaseExpenseDepartment.objects.filter(agency = agency).values('id', 'name')
        
    data = {
        'expense_dept_list': list(base_expense_dep),
    }
    return JsonResponse(data)

def setDataProductExpress(request):
    active = request.session['company_code']

    if 'id_product' in request.GET:
        id_product = request.GET.get('id_product')

        have_product = None
        exp_name = None
        exp_unit = None
        exp_alert = None
        
        try:
            have_product =  Product.objects.filter(id = id_product).exists()
            exp_product =  Product.objects.get(id = id_product)
            exp_name = exp_product.name
            if exp_product.unit:
                exp_unit = exp_product.unit.id
        except Product.DoesNotExist:
            exp_alert = "ไม่มีรหัสสินค้า '"+ str(id_product) +"' ในระบบ"

        try:
            if have_product:
                last_rq = RequisitionItem.objects.filter(requisit__branch_company__code = active, product__id = id_product).order_by('-id')[0]
                
                formatted_quantity = str(int(last_rq.quantity)) if last_rq.quantity == int(last_rq.quantity) else str(last_rq.quantity)
                exp_alert = "เบิกรายการนี้ล่าสุด วันที่ " + str(datetime.datetime.strptime(str(last_rq.created), "%Y-%m-%d").strftime("%d/%m/%Y")) + " จำนวน " + str(formatted_quantity) + " " + last_rq.unit
        except:
            exp_alert = None
        
    data = {
        'have_product': have_product,
        'exp_name': exp_name,
        'exp_unit': exp_unit,
        'exp_alert': exp_alert,
    }
    return JsonResponse(data)

def findExaminerUserComparisonPrice(request, cpd_id, cm_type):
    active = request.session['company_code']

    try:
        company_in = findCompanyIn(request)
    except:
        company_in = BaseBranchCompany.objects.filter(code = active).values('code')

    cpd = ComparisonPriceDistributor.objects.get(id = cpd_id)
    if cm_type == '1' or cm_type == '2':
        if cm_type == '1':
            permiss = BasePermission.objects.get(codename = 'CAECPD')
        elif cm_type == '2':
            permiss = BasePermission.objects.get(codename = 'CAECPA')
        position = PositionBasePermission.objects.filter(base_permission = permiss.id, branch_company__code__in = company_in).values('position_id')
    else:
        permiss = BasePermission.objects.filter(ap_amount_min__lte = cpd.amount , ap_amount_max__gte = cpd.amount, codename__in = ['CAECP1','CAECP2','CAECP3','CAECP4']).values('id')
        position = PositionBasePermission.objects.filter(base_permission__in = permiss, branch_company__code__in = company_in).values('position_id')
    user = UserProfile.objects.filter(position__in = position, branch_company__code = active).values('user__id', 'user__first_name','user__last_name')

    return user

def findApproveUserComparisonPrice(request, cpd_id, cm_type):
    active = request.session['company_code']

    try:
        company_in = findCompanyIn(request)
    except:
        company_in = BaseBranchCompany.objects.filter(code = active).values('code')

    cpd = ComparisonPriceDistributor.objects.get(id = cpd_id)
    if cm_type == '1' or cm_type == '2':
        if cm_type == '1':
            permiss = BasePermission.objects.get(codename = 'CAACPD')
        elif cm_type == '2':
            permiss = BasePermission.objects.get(codename = 'CAACPA')
        position = PositionBasePermission.objects.filter(base_permission = permiss.id, branch_company__code__in = company_in).values('position_id')
    else: 
        permiss = BasePermission.objects.filter(ap_amount_min__lte = cpd.amount , ap_amount_max__gte = cpd.amount, codename__in= ['CAACP1','CAACP2','CAACP3','CAACP4']).values('id')
        position = PositionBasePermission.objects.filter(base_permission__in = permiss, branch_company__code__in = company_in).values('position_id')
    user = UserProfile.objects.filter(position__in = position, branch_company__code = active).values('user__id', 'user__first_name','user__last_name')

    return user

def calculateSumQuantityCPItem(r_item, cp):
    quantity_cp_item = None
    #ค้นหาจำนวนสินค้าในใบเปรียบเทียบเพื่อดูว่าดึงสินค้าจากใบขอซื้อไปใช้หมดหรือยัง
    cp_item_query = ComparisonPriceItem.objects.filter(item = r_item, cp = cp)
    quantity_cp_item = cp_item_query.aggregate(Sum('quantity'))
    sum_cp_item = quantity_cp_item['quantity__sum'] if cp_item_query else Decimal(0.00)
    return sum_cp_item

def calculateSumQuantityPOItem(r_item, po):
    quantity_po_item = None
    #ค้นหาจำนวนสินค้าในใบสั่งซื้อเพื่อดูว่าดึงสินค้าจากใบขอซื้อไปใช้หมดหรือยัง
    po_item_query = PurchaseOrderItem.objects.filter(item = r_item, po = po, po__cp__isnull = True, po__is_cancel = False)
    quantity_po_item = po_item_query.aggregate(Sum('quantity'))
    sum_po_item = quantity_po_item['quantity__sum'] if po_item_query else Decimal(0.00)
    return sum_po_item


def calculateSumQuantityCPAndPOItem(r_item):
    quantity_cp_item = None
    quantity_po_item = None

    #ค้นหาจำนวนสินค้าในใบเปรียบเทียบเพื่อดูว่าดึงสินค้าจากใบขอซื้อไปใช้หมดหรือยัง
    cp_item_query = (
        ComparisonPriceItem.objects.filter(item=r_item)
        .values('cp')
        .annotate(max_quantity=Max('quantity'))#Take the maximum quantity for each 'cp'
    )
    total_sum = cp_item_query.aggregate(total_sum=Sum('max_quantity'))['total_sum']
    sum_cp_item = total_sum if total_sum else Decimal(0.00)

    #ค้นหาจำนวนสินค้าในใบสั่งซื้อเพื่อดูว่าดึงสินค้าจากใบขอซื้อไปใช้หมดหรือยัง
    po_item_query = PurchaseOrderItem.objects.filter(item = r_item, po__cp__isnull = True, po__is_cancel = False).values('po').distinct()
    quantity_po_item = po_item_query.aggregate(Sum('quantity'))
    sum_po_item = quantity_po_item['quantity__sum'] if po_item_query else Decimal(0.00)

    return sum_cp_item + sum_po_item

def get_bool(str):
    if str == 'True':
        bol  = True
    elif str == 'False':
        bol =  False
    return bol

def createPR(request, requisition_id):
    active = request.session['company_code']
    bc = BaseBranchCompany.objects.get(code = active)

    company = BaseBranchCompany.objects.get(code = active)

    items= RequisitionItem.objects.filter(requisition_id = requisition_id, quantity_pr__gt=0)
    try:
        requisition = Requisition.objects.get(id=requisition_id, purchase_requisition_id__isnull = True)
    except:
        return redirect('preparePR')

    baseUrgency = BaseUrgency.objects.all() #ระดับความเร่งด่วน
    baseUnit = BaseUnit.objects.all()
    #หา id express
    baseProduct = Product.objects.all().values('id', 'name')

    #ถ้า user login เป็นจัดซื้อ
    isPurchasing = is_purchasing(request.user)
        
    #หาจำนวนสินค้าทั้งหมด
    quantityTotal = 0
    for item in items:
        quantityTotal += item.quantity_pr

    #หาว่ามีใบจ่ายสินค้าภายในหรือไม่
    try:
        iv = Invoice.objects.get(requisit__id = requisition_id)
    except Invoice.DoesNotExist:
        iv = None

    '''24-07-2024 เปลี่ยนเป็นดึงจาก หมายเหตุ/เหตุผล ของใบขอเบิกแทน
    try:
        first_items = RequisitionItem.objects.filter(requisition_id = requisition_id, quantity_pr__gt=0).first()
        description = first_items.description
    except:
        description = ""
    '''
        
    #form
    form = PurchaseRequisitionForm(request, request.POST or None, initial={'note': requisition.note, 'branch_company': company, 'organizer': requisition.organizer,})
    if form.is_valid():
        #save id express
        bool = saveIdExpressPR(request)
        if bool is None:
            #save PR
            new_contact = form.save()

            #save requisition ใน purchase_requisition_id
            new_contact.requisition = requisition
            new_contact.purchase_status_id = 1 #กำหนดค่าเริ่มต้น
            new_contact.purchase_user_id = requisition.chief_approve_user_name
            # set ให้ผู้ขอซื้ออนุมัติมาเลย
            new_contact.purchase_status_id = 2
            new_contact.purchase_update = datetime.datetime.now()

            new_contact.approver_status_id = 1 #กำหนดค่าเริ่มต้น
            #พัสดุ
            new_contact.stockman_user_id = request.user.id
            new_contact.stockman_update = datetime.datetime.now()

            new_contact.save()

            #save purchase_requisition_id ใน requisition
            obj = Requisition.objects.get(id = requisition_id)
            obj.purchase_requisition_id = new_contact.pk
            obj.organizer = new_contact.organizer
            obj.pr_ref_no = new_contact.ref_no
            obj.save()

            messages.info(request, "สร้างใบขอซื้อ " + str(new_contact.ref_no) + " เรียบร้อยแล้ว")
            if iv:
                messages.success(request, "สร้างใบจ่ายสินค้าภายใน " + str(iv.ref_no)+ " เรียบร้อยแล้ว")
            return HttpResponseRedirect(reverse('viewPR'))
        else:
            return HttpResponseRedirect(reverse('createPR', args=(requisition_id,)))

    #ที่อยู่และหัวบริษัท
    company = BranchCompanyBaseAdress.objects.filter(branch_company__code = active ).first()

    #context
    context = {
        'items':items,
        'requisition':requisition,
        'quantityTotal': quantityTotal,
        'baseUrgency': baseUrgency,
        'baseUnit': baseUnit,
        'form':form,
        'baseProduct':baseProduct,
        'isPurchasing':isPurchasing,
        'company':company,
        'bc':bc,
        'pr_page': "tab-active",
        'create_mode': True,
        'pr_show': "show",
        active :"active show",
		"disableTab":"disableTab",
		"colorNav":"disableNav"
    }
    return render(request,'purchaseRequisition/createPR.html', context)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def createCMorPO(request, pr_id):
    active = request.session['company_code']
    bc = BaseBranchCompany.objects.get(code = active)

    try:
        company = BaseBranchCompany.objects.get(code = request.session['company_code'])
    except:
        return redirect('signOut')

    pr = PurchaseRequisition.objects.get(id=pr_id)

    items= RequisitionItem.objects.filter(requisition_id = pr.requisition.id, quantity_pr__gt=0)
    
    #หาจำนวนของที่ซื้อไปแล้ว
    itemUseQuantity = []
    for i in items:
        quantity = calculateSumQuantityCPAndPOItem(i)
        itemUseQuantity.append(quantity)

    requisition = Requisition.objects.get(id=pr.requisition.id)
    baseUrgency = BaseUrgency.objects.all()
    baseUnit = BaseUnit.objects.all()
    #หา id express
    baseProduct = Product.objects.all().values('id', 'name')

    #หาจำนวนสินค้าทั้งหมด
    quantityTotal = 0
    for item in items:
        quantityTotal += item.quantity_pr
    
    listItem = request.POST.getlist('choices')

    if request.method=='POST' and 'btnformCM' in request.POST:
        cp = ComparisonPrice.objects.create(
            organizer = request.user,
            approver_status_id = 1,
            examiner_status_id = 1,
            branch_company = pr.branch_company,
            address_company = pr.address_company,
            note = pr.note,
        )
        cp.save()

        cpd = ComparisonPriceDistributor.objects.create(
            cp_id = cp.id,
            vat_type_id = 0
        )
        cpd.save()
        for i in listItem:
            #บอกสถานะว่าใช้ไปใน CM หรือ PR แล้ว
            ri = RequisitionItem.objects.get(id = i)
            sum_po_cp = calculateSumQuantityCPAndPOItem(ri)
            remain = ri.quantity_pr - sum_po_cp #คำนวนหาจำนวนสินค้าในใบเปรียบเทียบและใบสั่งซื้อที่ทำไปแล้ว 24-09-2022

            try:
               b_unit =  BaseUnit.objects.filter(name = ri.unit).first()
            except BaseUnit.DoesNotExist:
               b_unit = ri.product.unit

            #สร้าง ComparisonPriceItem ใหม่
            item = ComparisonPriceItem.objects.create(
                item_id = i,
                bidder_id = cpd.id,
                cp = cp.id,
                quantity = remain,
                unit = b_unit
            )
            item.save()

            ri.is_used = True#save ว่าใช้ไปแล้ว
            ri.quantity_used = calculateSumQuantityCPAndPOItem(ri)#save จำนวนสินค้าที่ใช้ไปแล้ว
            ri.save()
            
        #check สินค้าว่าดึงไปทำหมดไหม
        checkRequisitionItemAllUsed(pr.requisition)
        return HttpResponseRedirect(reverse('editComparePricePOItemFromPR', args=(cp.id, cpd.id)))
    
    if request.method=='POST' and 'btnformPO' in request.POST:
        po = PurchaseOrder.objects.create(
            stockman_user = request.user,
            approver_status_id = 1,
            vat_type_id = 0,
            pr = pr,
            branch_company = pr.branch_company,
            address_company = pr.address_company,
        )
        po.save()

        for i in listItem:
            #บอกสถานะว่าใช้ไปใน CM หรือ PR แล้ว
            ri = RequisitionItem.objects.get(id = i)
            sum_po_cp = calculateSumQuantityCPAndPOItem(ri)#คำนวนหาจำนวนสินค้าในใบเปรียบเทียบและใบสั่งซื้อที่ทำไปแล้ว 24-09-2022
            remain = ri.quantity_pr - sum_po_cp

            try:
               b_unit =  BaseUnit.objects.filter(name = ri.unit).first()
            except BaseUnit.DoesNotExist:
               b_unit = ri.product.unit 
            
            #สร้าง PurchaseOrderItem ใหม่
            item = PurchaseOrderItem.objects.create(
                item_id = i,
                quantity = remain,
                po_id = po.id,
                unit = b_unit
            )
            item.save()

            ri.is_used = True#save ว่าใช้ไปแล้ว
            ri.quantity_used = calculateSumQuantityCPAndPOItem(ri)#save จำนวนสินค้าที่ใช้ไปแล้ว
            ri.save()

        #check สินค้าว่าดึงไปทำหมดไหม
        checkRequisitionItemAllUsed(pr.requisition)
        return HttpResponseRedirect(reverse('editPOFromPR', args=(po.id,)))

    context = {
        'items':items,
        'itemUseQuantity': itemUseQuantity,
        'requisition':requisition,
        'quantityTotal': quantityTotal,
        'baseUrgency': baseUrgency,
        'baseProduct':baseProduct,        
        'pr': pr,
        'bc': bc,
        'pr_page': "tab-active",
        'pr_show': "show",
        active :"active show",
		"disableTab":"disableTab",
		"colorNav":"disableNav"
    }
    return render(request,'purchaseRequisition/createCMorPO.html',context)

def checkRequisitionItemAllUsed(requisit_id):
    #เช็คว่าดึงรายการในใบขอเบิกไปใช้หมดหรือยัง
    try:
        check_item = RequisitionItem.objects.filter(requisit = requisit_id, is_used = False).distinct()
        try:
            pr = PurchaseRequisition.objects.get(requisition = requisit_id)
            if not check_item:
                pr.is_complete = True
            else:
                pr.is_complete = False
            pr.save()
        except:
            pass
    except RequisitionItem.DoesNotExist:
        pass

def removePR(request, pr_id):
    pr = PurchaseRequisition.objects.get(id = pr_id)
    requisition = Requisition.objects.get(purchase_requisition_id = pr_id)
    #set อ้างอิง pr = None
    requisition.purchase_requisition_id = None
    requisition.pr_ref_no = None

    #ลบใบจ่ายสินค้าภายใน auto
    try:
        iv = Invoice.objects.get(requisit = requisition)
        iv_items = InvoiceItem.objects.filter(iv = iv)
        iv_items.delete()
        iv.delete()

        #set อ้างอิง iv = None
        requisition.invoice_id = None
        requisition.iv_ref_no = None

    except Invoice.DoesNotExist:
        pass
    
    requisition.is_edit = True
    requisition.save()
    
    pr.delete()
    return redirect('viewPR')

def editPR(request, pr_id):
    active = request.session['company_code']
    pr = PurchaseRequisition.objects.get(id=pr_id)
    bc = BaseBranchCompany.objects.get(code = pr.branch_company)
    form = PurchaseRequisitionForm(request, instance=pr)

    items= RequisitionItem.objects.filter(requisition_id = pr.requisition.id, quantity_pr__gt=0)
    requisition = Requisition.objects.get(id=pr.requisition.id)
    baseUrgency = BaseUrgency.objects.all()
    baseUnit = BaseUnit.objects.all()
    #หา id express
    baseProduct = Product.objects.all().values('id', 'name')

    #หาจำนวนสินค้าทั้งหมด
    quantityTotal = 0
    for item in items:
        quantityTotal += item.quantity_pr

    if request.method == 'POST':
        form = PurchaseRequisitionForm(request, request.POST, instance=pr)
        if form.is_valid():
            new_contact = form.save()
            #save id express
            saveIdExpressPR(request)
            
            return redirect('viewPR')

    context = {
        'items':items,
        'requisition':requisition,
        'quantityTotal': quantityTotal,
        'baseUrgency': baseUrgency,
        'baseUnit': baseUnit,
        'form':form,
        'baseProduct':baseProduct,
        'pr': pr,
        'bc': bc,
        'create_mode': False,
        'pr_page': "tab-active",
        'pr_show': "show",
        active :"active show",
		"disableTab":"disableTab",
		"colorNav":"disableNav"
    }
    return render(request, "purchaseRequisition/createPR.html", context)

def showPR(request, pr_id, mode):
    active = request.session['company_code']

    pr = PurchaseRequisition.objects.get(id=pr_id)
    bc = BaseBranchCompany.objects.get(code = pr.branch_company)

    items= RequisitionItem.objects.filter(requisition_id = pr.requisition.id, quantity_pr__gt=0)
    requisition = Requisition.objects.get(id=pr.requisition.id)
    baseUrgency = BaseUrgency.objects.all()
    baseUnit = BaseUnit.objects.all()
    #หา id express
    baseProduct = Product.objects.all().values('id', 'name')

    #หาจำนวนสินค้าทั้งหมด
    quantityTotal = 0
    for item in items:
        quantityTotal += item.quantity_pr

    page = PRPageMode(mode)
    show = PRShowMode(mode)

    #ถ้า user login เป็นจัดซื้อ
    isPurchasing = is_purchasing(request.user)
    #ถ้า user login เป็นจัดซื้อ
    isSupplies = is_supplies(request.user)
    #ถ้า user มีสิทธิ edit pr
    isEditPR = is_edit_pr(request.user)

    is_staff = False
    if isPurchasing or isSupplies:
        is_staff = True

    # re approver แสดงเฉพาะหน้า history complete และ history incomplete เท่านั้น
    isReApprove = False
    if (mode == 4 or mode == 5) and (isPurchasing or isSupplies):
        isReApprove = True

    # re pr นำรายการใบขอซื้อกลับมาทำใหม่ โดยไม่จำเป็นต้องทำการอนุมัติใหม่
    '''
    isRePr = False
    if (mode == 4 or mode == 5) and isPurchasing and pr.organizer.id == request.user.id and pr.approver_status.id == 2:
        isRePr = True    
    '''

    isRePr = is_re_pr(request.user)

    #ที่อยู่และหัวบริษัท
    company = BranchCompanyBaseAdress.objects.filter(branch_company__code = active).first()
    form = RequisitionMemorandumForm(instance=requisition)
    pr_form = PurchaseRequisitionAddressCompanyForm(instance=pr)
    pro_form = PurchaseRequisitionOrganizerForm(request, instance=pr)

    if request.method == 'POST' and 'btnformR' in request.POST:
        form = RequisitionMemorandumForm(request.POST, request.FILES, instance=requisition)
        if form.is_valid():
            form.save()
            return redirect('viewPRHistory')

    if request.method == 'POST' and 'btnformPR' in request.POST:
        pr_form = PurchaseRequisitionAddressCompanyForm(request.POST, request.FILES, instance=pr)
        if pr_form.is_valid():
            pr_form.save()
            return redirect('showPR', pr_id = pr.id , mode = mode)
    
    if request.method == 'POST' and 'btnformPRO' in request.POST:
        pro_form = PurchaseRequisitionOrganizerForm(request, request.POST, request.FILES, instance=pr)
        if pro_form.is_valid():
            pro_form.save()
            return redirect('viewPRHistory')

    context = {
            'items':items,
            'requisition':requisition,
            'quantityTotal': quantityTotal,
            'baseUrgency': baseUrgency,
            'baseUnit': baseUnit,
            'baseProduct':baseProduct,
            'pr': pr,
            'isPurchasing': isPurchasing,
            'isSupplies': isSupplies,
            'is_staff': is_staff,
            'isEditPR': isEditPR,
            'bc':bc,
            'create_mode': False,
            page: "tab-active",
            show: "show",
            'form':form,
            'pr_form': pr_form,
            'pro_form': pro_form,
            'isReApprove': isReApprove,
            'isRePr':isRePr,
            active :"active show",
			"disableTab":"disableTab",
			"colorNav":"disableNav"
    }

    return render(request, "purchaseRequisition/showPR.html", context)

@require_http_methods(["POST"])
def saveIdExpressPR(request):
    try:
        list_id = request.POST.getlist('id') or None
        if list_id:
            list_id = list(map(int, list_id))
            list_product = request.POST.getlist('product') or None
            try:
                list_product = list(map(int, list_product))
            except ValueError:
                pass
            for i, val_id in enumerate(list_id):
                for j, val_product in enumerate(list_product):
                    if i == j:
                        item = get_object_or_404(RequisitionItem, id=val_id)
                        if val_product != "":
                            try:
                                product_item = get_object_or_404(Product, id=val_product)
                                item.product = product_item
                            except:
                                return False
                            item.save()
                    else:
                        pass
    except stripe.error.CardError as e:
        return False, e


@login_required(login_url='signIn')
def viewPRApprove(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)
    data = PurchaseRequisition.objects.filter(Q(purchase_status_id = 1)| Q(approver_status_id = 1) & ~Q(purchase_status_id = 3), branch_company__code__in = company_in)

    #กรองข้อมูล
    myFilter = PurchaseRequisitionFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)

    context = {
                'prs':dataPage,
                'filter':myFilter,
                'ap_pr_page': "tab-active",
                'ap_pr_show': "show",
                active :"active show",
                "colorNav":"enableNav"
              }

    return render(request,'purchaseRequisitionApprove/viewPRApprove.html',context)

def editPRApprove(request, pr_id, isFromHome):
    active = request.session['company_code']

    try:
        company_in = findCompanyIn(request)
    except:
        company_in = BaseBranchCompany.objects.filter(code = active).values('code')

    pr = PurchaseRequisition.objects.get(id=pr_id)
    bc = BaseBranchCompany.objects.get(code = pr.branch_company)

    items= RequisitionItem.objects.filter(requisition_id = pr.requisition.id, quantity_pr__gt=0)
    requisition = Requisition.objects.get(id=pr.requisition.id)
    baseUrgency = BaseUrgency.objects.all()
    baseUnit = BaseUnit.objects.all()

    #หาจำนวนสินค้าทั้งหมด
    quantityTotal = 0
    for item in items:
        quantityTotal += item.quantity_pr

    #get permission with position login
    try:
        user_profile = UserProfile.objects.get(user_id = request.user.id)

        permiss = BasePermission.objects.filter(codename ='CAAPR')
        permiss_pr = PositionBasePermission.objects.filter(position_id = user_profile.position_id, base_permission__codename='CAAPR', branch_company__code__in = company_in).values('branch_company__code')
    except:
        permiss_pr  = None

    try:
        in_company = BaseBranchCompany.objects.filter(userprofile = user_profile, code__in = permiss_pr).exists()
    except:
        pass


    #ถ้าเป็นผู้อนุมัติ
    isPermiss = False
    if(permiss_pr and in_company):
        try:
            #ดึงข้อมูล PurchaseRequisition
            isPermiss = PurchaseRequisition.objects.filter(id = pr_id, approver_user = request.user, branch_company__code__in = permiss_pr).exists() #หาสถานะรอดำเนินการของผู้อนุมัติ
            #หาความยาวของ index PurchaseRequisition ที่มี สถานะรอดำเนินการของผู้อนุมัติ
        except PurchaseRequisition.DoesNotExist:
            isPermiss = False


    context = {
        'items':items,
        'requisition':requisition,
        'quantityTotal': quantityTotal,
        'baseUrgency': baseUrgency,
        'baseUnit': baseUnit,
        'pr':pr,
        'isPermiss':isPermiss,
        'isFromHome':isFromHome,
        'bc':bc,
        'ap_pr_page': "tab-active",
        'ap_pr_show': "show",
        active :"active show",
		"disableTab":"disableTab",
		"colorNav":"disableNav"
    }

    if request.method == 'POST':
        post_status = request.POST['status'] or None
        status = BaseApproveStatus.objects.get(name = post_status)

        obj = PurchaseRequisition.objects.get(id = pr_id)
        #ผู้ขอซื้อ
        if(obj.purchase_user_id == request.user.id):
            obj.purchase_status = status
            obj.purchase_update = datetime.datetime.now()
        #ผู้อนุมัติ
        if(isPermiss):
            obj.approver_status = status
            obj.approver_user_id = request.user.id
            obj.approver_update = datetime.datetime.now()
        #หากอนุมัติแล้วแก้ใบขอเบิกไม่ได้
        if obj.purchase_status_id != 1 or obj.approver_status_id != 1:
            try:
                r = Requisition.objects.get(purchase_requisition_id = obj.id)
                r.is_edit = False
                r.save()
            except Requisition.DoesNotExist: #ยกเลิกเนื่องจาก ออกใบขอซื้อเบิล 22/10/2024
                obj.approver_status_id = 4
                tmp_note = obj.note
                obj.note = "(ยกเลิกจากระบบ)" + tmp_note
        obj.save()
        return redirect('editPRApprove', pr_id = pr_id, isFromHome = isFromHome)

    return render(request, "purchaseRequisitionApprove/editPRApprove.html", context)

@login_required(login_url='signIn')
def viewPO(request):
    active = request.session['company_code']
    data = PurchaseOrder.objects.filter(Q(approver_status_id = 1) | Q(is_receive = False) & ~Q(approver_status_id = 3), is_cancel = False, branch_company__code = active)

    #กรองข้อมูล
    myFilter = PurchaseOrderFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)

    prs = PurchaseOrderItem.objects.filter(po__in=data).values('item__requisit__pr_ref_no','item__requisit__purchase_requisition_id','po')

    context = {
        'pos':dataPage,
        'filter':myFilter,
        'prs': prs,
        'po_page': "tab-active",
        'po_show': "show",
        active :"active show",
        "colorNav":"enableNav"
    }
    return render(request, "purchaseOrder/viewPO.html", context)

def preparePO(request):
    return render(request, 'purchaseOrder/preparePO.html')

def createPO(request):
    try:
        company = BaseBranchCompany.objects.get(code = request.session['company_code'])
    except:
        return redirect('signOut')

    # set ค่าเริ่มต้นของเจ้าหน้าที่พัสดุ และสเตตัสการอนุมัติเป็น รอดำเนินการ
    form = PurchaseOrderForm(request.POST or None)
    if form.is_valid():
        form = PurchaseOrderForm(request.POST or None, request.FILES)
        new_contact = form.save(commit=False)
        new_contact.stockman_user = request.user
        new_contact.approver_status_id = 1
        new_contact.branch_company = company
        new_contact.save()
        return HttpResponseRedirect(reverse('createPOItem', args=(new_contact.pk,)))

    context = {
        'form':form,
        'po_page': "tab-active",
        'po_show': "show",
    }

    return render(request, "purchaseOrder/createPO.html", context)

def editPO(request, po_id):
    po = PurchaseOrder.objects.get(id=po_id)
    form = PurchaseOrderForm(instance=po)
    if request.method == 'POST':
        form = PurchaseOrderForm(request.POST, instance=po)
        if form.is_valid():
            form.save()
            return redirect('viewPO')

    context = {
        'form':form,
        'po_page': "tab-active",
        'po_show': "show",
    }

    return render(request, "purchaseOrder/createPO.html", context)

def editPOFromPR(request, po_id):
    active = request.session['company_code']
    company = BaseBranchCompany.objects.get(code = request.session['company_code'])
    #distributorList = Distributor.objects.filter(affiliated = company.affiliated)
    #distributorList = Distributor.objects.all().values('id','name','credit__id','vat_type__id')

    #ถ้า user login แก้ไขผู้อนุมัติบสั่งซื้อได้
    isEditApproverUserPO = is_edit_approver_user_po(request.user)

    po = PurchaseOrder.objects.get(id=po_id)
    form = PurchaseOrderForm(instance=po)
    form_rate = RateDistributorForm()

    if request.method == 'POST':
        form = PurchaseOrderForm(request.POST, request.FILES, instance=po)
        form_rate = RateDistributorForm(request.POST or None)
        if form.is_valid() and form_rate.is_valid():
            form.save()
            rate_contact = form_rate.save(commit=False)
            rate_contact.po = po
            rate_contact.organizer_user = po.stockman_user
            if rate_contact.price_rate and rate_contact.quantity_rate and rate_contact.service_rate and rate_contact.safety_rate:
                rate_contact.save()
            return redirect('editPOItem', po_id = po_id, isFromPR = 'True', isReApprove = 'False')
        
    #'distributorList': distributorList,
    context = {
        'form':form,
        'form_rate':form_rate,
        'isEditApproverUserPO': isEditApproverUserPO,
        'po_page': "tab-active",
        'po_show': "show",
        active :"active show",
		"disableTab":"disableTab",
		"colorNav":"disableNav"
    }

    return render(request, "purchaseOrder/createPO.html", context)

def editPOFromComparison(request, po_id):
    po = PurchaseOrder.objects.get(id=po_id)
    form = PurchaseOrderFromComparisonPriceForm(instance=po)
    if request.method == 'POST':
        form = PurchaseOrderFromComparisonPriceForm(request.POST, instance=po)
        if form.is_valid():
            form.save()
            return redirect('viewPO')

    context = {
        'form':form,
        'po_page': "tab-active",
        'po_show': "show",
    }

    return render(request, "purchaseOrder/createPO.html", context)

def removePO(request, po_id):
    po = PurchaseOrder.objects.get(id = po_id)
    #ลบ item ใน PurchaseOrder ด้วย
    items = PurchaseOrderItem.objects.filter(po = po)
    for obj in items:
        try:
            d_item = RequisitionItem.objects.get(id = obj.item.id)
            #คำนวนหาจำนวนสินค้าในใบเปรียบเทียบและใบสั่งซื้อที่ทำไปแล้ว 24-09-2022
            po_sum_quantity = calculateSumQuantityPOItem(obj.item, po)
            sum_po_cp = calculateSumQuantityCPAndPOItem(obj.item)
            remain = d_item.quantity_pr - (sum_po_cp - po_sum_quantity)

            #save จำนวนสินค้าที่ใช้ไปแล้ว
            d_item.quantity_used = sum_po_cp - po_sum_quantity

            if remain <= 0:
                d_item.is_used = True
                d_item.save()
                #เช็คว่าดึงรายการในใบขอเบิกไปใช้หมดหรือยัง
                check_item = RequisitionItem.objects.filter(requisit = d_item.requisit, is_used = False).distinct()
                if check_item:
                    try:
                        pr = PurchaseRequisition.objects.get(requisition = d_item.requisit)
                        pr.is_complete = True
                        pr.save()
                    except:
                        pass
            else:
                d_item.is_used = False
                d_item.save()
                pr = PurchaseRequisition.objects.get(requisition = d_item.requisit)
                pr.is_complete = False
                pr.save()
        except RequisitionItem.DoesNotExist:
            pass

    items.delete()

    #ล้าง po_ref_no ที่ผูกไว้
    try:
        cp = ComparisonPrice.objects.get(po_ref_no = po.ref_no)
        cp.po_ref_no = ""
        cp.save()
    except (ComparisonPrice.DoesNotExist, AttributeError):
        pass

    #ลบรายการประเมินร้านค้าที่ผูกไว้
    try:
        rd = RateDistributor.objects.filter(po = po)
        for r in rd:
            r.delete()
    except (RateDistributor.DoesNotExist, AttributeError):
        pass

    #ลบ PurchaseOrder ทีหลัง
    po.delete()
    return redirect('viewPO')

def showPO(request, po_id, mode):
    active = request.session['company_code']

    base_po_type = BasePOType.objects.all()
    po = PurchaseOrder.objects.get(id = po_id)
    bc = BaseBranchCompany.objects.get(code = po.branch_company)

    items = PurchaseOrderItem.objects.filter(po = po)

    new_pr_id = dict()
    if items:
        for obj in items:
            if obj.item.requisit.purchase_requisition_id not in new_pr_id:
                new_pr_id[obj.item.requisit.purchase_requisition_id] = obj

    new_pr = dict()
    for id in new_pr_id:
        try:
            pr = PurchaseRequisition.objects.get(id = id)
            new_pr[pr] = pr
        except PurchaseRequisition.DoesNotExist:
            pass
    
    page = POPageMode(mode)
    show = POShowMode(mode)

    form = PurchaseOrderReceiptForm(instance=po)
    if request.method == 'POST' and 'btnformPOr' in request.POST:
        form = PurchaseOrderReceiptForm(request.POST, request.FILES, instance=po)
        if form.is_valid():
            form.save()
            return redirect('viewPOHistory')

    poa_form = PurchaseOrderAddressCompanyForm(instance=po)
    if request.method == 'POST' and 'btnformPOa' in request.POST:
        poa_form = PurchaseOrderAddressCompanyForm(request.POST, request.FILES, instance=po)
        if poa_form.is_valid():
            poa_form.save()
            return redirect('showPO', po_id = po_id , mode = mode)

    cancel_form = PurchaseOrderCancelForm(instance=po)
    if request.method == 'POST' and 'btnformCancelPO' in request.POST:
        cancel_form = PurchaseOrderCancelForm(request.POST, instance=po)
        if cancel_form.is_valid():
            cc = cancel_form.save(commit=False)
            if cc.cancel_reason is not None:
                cc.is_cancel = True
                cc.save()
                rePOItemIsUse(po_id)#หากยกเลิกรายการ เช็ค item และนำใบขอซื้อกลับมาทำใหม่
            return redirect('showPO', po_id = po_id , mode = mode)

    #ถ้า user login เป็นจัดซื้อ
    isPurchasing = is_purchasing(request.user)
    #ถ้า user login เป็นจัดซื้อ
    isSupplies = is_supplies(request.user)
    isUploadeReceipt = False
    if isPurchasing or isSupplies:
        isUploadeReceipt = True

    context = {
            'po':po,
            'items':items,
            'new_pr':new_pr,
            'form': form,
            'isUploadeReceipt':isUploadeReceipt,
            'isPurchasing':isPurchasing,
            'mode': mode,
            'poa_form':poa_form,
            'cancel_form':cancel_form,
            'bc':bc,
            'base_po_type':base_po_type,
            page: "tab-active",
            show: "show",
            active :"active show",
			"disableTab":"disableTab",
			"colorNav":"disableNav"
    }
    return render(request, 'purchaseOrder/showPO.html',context)

def rePOItemIsUse(po_id):
    items = PurchaseOrderItem.objects.filter(po = po_id)
    for obj in items:
        try:
            d_item = RequisitionItem.objects.get(id = obj.item.id)
            #คำนวนหาจำนวนสินค้าในใบเปรียบเทียบและใบสั่งซื้อที่ทำไปแล้ว 24-09-2022
            po_sum_quantity = calculateSumQuantityPOItem(obj.item, po_id)
            sum_po_cp = calculateSumQuantityCPAndPOItem(obj.item)
            remain = d_item.quantity_pr - (sum_po_cp - po_sum_quantity)

            #save จำนวนสินค้าที่ใช้ไปแล้ว
            d_item.quantity_used = sum_po_cp - po_sum_quantity

            if remain <= 0:
                d_item.is_used = True
                d_item.save()
                #เช็คว่าดึงรายการในใบขอเบิกไปใช้หมดหรือยัง
                check_item = RequisitionItem.objects.filter(requisit = d_item.requisit, is_used = False).distinct()
                if check_item:
                    try:
                        pr = PurchaseRequisition.objects.get(requisition = d_item.requisit)
                        pr.is_complete = True
                        pr.save()
                    except:
                        pass
            else:
                d_item.is_used = False
                d_item.save()
                pr = PurchaseRequisition.objects.get(requisition = d_item.requisit)
                pr.is_complete = False
                pr.save()
        except RequisitionItem.DoesNotExist:
            pass

def createPOItem(request, po_id):
    template_name = 'purchaseOrderItem/createPOItem.html'
    heading_message = 'Model Formset Demo'

    active = request.session['company_code']
    bc = BaseBranchCompany.objects.get(code = active)

    #ดึง item ที่ทำใบ po แล้ว
    itemList = RequisitionItem.objects.filter(requisit__purchase_requisition_id__isnull = False, is_receive = False, product__isnull = False)

    po_data = PurchaseOrder.objects.get(id=po_id)
    if request.method == 'GET':
        formset = PurchaseOrderItemModelFormset(queryset=PurchaseOrderItem.objects.none())
        price_form = PurchaseOrderPriceForm()
    elif request.method == 'POST':
        formset = PurchaseOrderItemModelFormset(request.POST)
        price_form = PurchaseOrderPriceForm(request.POST, instance=po_data)
        if formset.is_valid() and price_form.is_valid():
            # save ราคาใบ po
            price = price_form.save(commit=False)
            if not price.discount:
                price.discount = 0.00
            if not price.freight:
                price.freight = 0.00
            price.save()

            # save po item
            for form in formset:
                # only save if name is present
                if form.cleaned_data.get('item'):
                    obj = form.save(commit=False)
                    obj.po = po_data
                    obj.save()
            #redirect นอก loop
            return redirect('viewPO')

    po = PurchaseOrder.objects.get(id = po_id)
    context = {
        'po':po,
        'po_page': "tab-active",
        'po_show': "show",
        'bc':bc,
        'formset': formset,
        'price_form': price_form,
        'itemList':itemList,
        'heading': heading_message,
    }
    return render(request, template_name, context)

def editPOItem(request, po_id, isFromPR, isReApprove):
    active = request.session['company_code']
    bc = BaseBranchCompany.objects.get(code = active)

    base_po_type = BasePOType.objects.all()

    template_name = 'purchaseOrderItem/editPOItem.html'
    heading_message = 'Model Formset Demo'

    #ถ้า user login แก้ไขรหัสใบสั่งซื้อได้
    isEditPO = is_edit_po_id(request.user)

    #ถ้า user login แก้ไขผู้อนุมัติบสั่งซื้อได้
    isEditApproverUserPO = is_edit_approver_user_po(request.user)

    #ดึง item ที่ทำใบ po แล้ว
    #itemList = RequisitionItem.objects.filter(requisit__purchase_requisition_id__isnull = False, is_receive = False, product__isnull = False)
    #เปลี่ยนให้ดึงสินค้าเฉพาะที่ตัดมาจากใบขอซื้อแล้วเท่านั้น 14-09-2022
    itemList = PurchaseOrderItem.objects.values('item__id','item__product__id','item__product_name', 'item__requisit__pr_ref_no', 'item__quantity_pr', 'item__product__unit__id', 'unit__id', 'item__quantity_used', 'quantity', 'po__cp').filter(po__id = po_id).annotate(
                    q_remain = ExpressionWrapper(F('item__quantity_pr') -  (F('item__quantity_used') - F('quantity')) , output_field = models.DecimalField()),
                )
    company = BaseBranchCompany.objects.get(code = request.session['company_code'])
    #distributorList = Distributor.objects.filter(affiliated = company.affiliated)
    #distributorList = Distributor.objects.all().values('id','name','credit__id','vat_type__id')

    po_data = PurchaseOrder.objects.get(id = po_id)
    po_items = PurchaseOrderItem.objects.filter(po = po_id)
    try:
        rate_dist = RateDistributor.objects.get(po = po_id)
    except RateDistributor.DoesNotExist:
        rate_dist = None
    
    form = PurchaseOrderForm(instance=po_data)
    if request.method == "POST":
        formset = PurchaseOrderItemInlineFormset(request.POST, request.FILES, instance=po_data)
        price_form = PurchaseOrderPriceForm(request.POST, instance=po_data)
        form = PurchaseOrderForm(request.POST, request.FILES, instance=po_data)
        form_rate = RateDistributorForm(request.POST, request.FILES, instance=rate_dist)
        if formset.is_valid() and price_form.is_valid() and form.is_valid() and form_rate.is_valid():
            # save ราคาใบ po
            price = price_form.save(commit=False)
            if not price.discount:
                price.discount = 0.00
            if not price.freight:
                price.freight = 0.00

            're approve po'
            if price.is_re_approve == False:
                price.is_re_approve = get_bool(isReApprove)
            if get_bool(isReApprove) == True:
                #เปิดให้ผู้อนุมัติ อนุมัติรายการใหม่
                price.approver_status_id = 1
                price.approver_update = None
                #เปลี่ยนสถานะยกเลิก เป็นรายการปกติ
                price.is_cancel = False
                price.cancel_reason = None
            price.save()

            #ประเมินร้านค้า
            rate_contact = form_rate.save(commit=False)
            rate_contact.po = po_data
            rate_contact.organizer_user = po_data.stockman_user
            if rate_contact.price_rate and rate_contact.quantity_rate and rate_contact.service_rate and rate_contact.safety_rate:
                rate_contact.save()

            po_form = form.save()
            #แก้เลขที่ผูก cp po_ref_no
            try:
                cp = ComparisonPrice.objects.get(id = po_form.cp_id)
                cp.po_ref_no = po_form.ref_no
                cp.save()
            except ComparisonPrice.DoesNotExist:
                pass

            # save po item
            instances = formset.save(commit=False)
            for instance in instances:
                instance.save()
                try:
                    r_item = RequisitionItem.objects.get(id = instance.item.id)
                    #คำนวนหาจำนวนสินค้าในใบเปรียบเทียบและใบสั่งซื้อที่ทำไปแล้ว 24-09-2022
                    sum_po_cp = calculateSumQuantityCPAndPOItem(instance.item)
                    remain = r_item.quantity_pr - sum_po_cp

                    r_item.quantity_used = sum_po_cp

                    if remain <= 0:
                        r_item.is_used = True
                        r_item.save()
                        #เช็คว่าดึงรายการในใบขอเบิกไปใช้หมดหรือยัง
                        check_item = RequisitionItem.objects.filter(requisit = r_item.requisit, is_used = False).distinct()
                        if not check_item:
                            try:
                                pr = PurchaseRequisition.objects.get(requisition = r_item.requisit)
                                pr.is_complete = True
                                pr.save()
                            except:
                                pass
                    else:
                        r_item.is_used = False
                        r_item.save()
                        pr = PurchaseRequisition.objects.get(requisition = r_item.requisit)
                        pr.is_complete = False
                        pr.save()
                except RequisitionItem.DoesNotExist:
                    pass
            for obj in formset.deleted_objects:
                try:
                    d_item = RequisitionItem.objects.get(id = obj.item.id)
                    d_item.quantity_used = calculateSumQuantityCPAndPOItem(obj.item) - obj.quantity
                    d_item.is_used = False
                    d_item.save()
                    #เช็คว่าดึงรายการในใบขอเบิกไปใช้หมดหรือยัง
                    check_item = RequisitionItem.objects.filter(requisit = d_item.requisit, is_used = False).distinct()
                    if check_item:
                        try:
                            pr = PurchaseRequisition.objects.get(requisition = d_item.requisit)
                            pr.is_complete = False
                            pr.save()
                        except:
                            pass
                except RequisitionItem.DoesNotExist:
                    pass
                obj.delete()
            formset.save_m2m()
            return redirect('viewPO')
    else:
        formset = PurchaseOrderItemInlineFormset(instance=po_data)
        price_form = PurchaseOrderPriceForm(instance=po_data)
        form = PurchaseOrderForm(instance=po_data)
        form_rate = RateDistributorForm(instance=rate_dist)

    po = PurchaseOrder.objects.get(id = po_id)
    items = PurchaseOrderItem.objects.filter(po = po_id)

    new_pr_id = dict()
    if items:
        for obj in items:
            if obj.item.requisit.purchase_requisition_id not in new_pr_id:
                new_pr_id[obj.item.requisit.purchase_requisition_id] = obj

    new_pr = dict()
    for id in new_pr_id:
        try:
            pr = PurchaseRequisition.objects.get(id = id)
            new_pr[pr] = pr
        except PurchaseRequisition.DoesNotExist:
            pass

    #'distributorList': distributorList,
    context = {
        'po':po,
        'po_items':po_items,
        'po_page': "tab-active",
        'po_show': "show",
        'formset': formset,
        'price_form':price_form,
        'form': form,
        'form_rate': form_rate,
        'isFromPR':isFromPR,
        'itemList':itemList,
        'new_pr':new_pr,
        'isEditPO':isEditPO,
        'isEditApproverUserPO':isEditApproverUserPO,
        'bc':bc,
        'base_po_type':base_po_type,
        'rate_dist': rate_dist,
        'heading': heading_message,
        active :"active show",
		"disableTab":"disableTab",
		"colorNav":"disableNav"
    }
    return render(request, template_name, context)

@login_required(login_url='signIn')
def viewPOApprove(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)
    data = PurchaseOrder.objects.filter(approver_status_id = 1, amount__isnull = False, amount__gt = 0, branch_company__code__in = company_in)

    #กรองข้อมูล
    myFilter = PurchaseOrderFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)

    context = {
                'pos':dataPage,
                'filter':myFilter,
                'ap_po_page': "tab-active",
                'ap_po_show': "show",
                active :"active show",
                "colorNav":"enableNav"
              }

    return render(request,'purchaseOrderApprove/viewPOApprove.html',context)

def editPOApprove(request, po_id, isFromHome):
    active = request.session['company_code']

    try:
        company_in = findCompanyIn(request)
    except:
        company_in = BaseBranchCompany.objects.filter(code = active).values('code')

    base_po_type = BasePOType.objects.all()
    po = PurchaseOrder.objects.get(id = po_id)
    bc = BaseBranchCompany.objects.get(code = po.branch_company)

    items = PurchaseOrderItem.objects.filter(po = po)
    new_pr_id = dict()
    if items:
        for obj in items:
            if obj.item.requisit.purchase_requisition_id not in new_pr_id:
                new_pr_id[obj.item.requisit.purchase_requisition_id] = obj

    new_pr = dict()
    for id in new_pr_id:
        try:
            pr = PurchaseRequisition.objects.get(id = id)
            new_pr[pr] = pr
        except PurchaseRequisition.DoesNotExist:
            pass

    #get permission with position login
    try:
        user_profile = UserProfile.objects.get(user_id = request.user.id)

        permiss = BasePermission.objects.filter(codename ='CAAPO')
        permiss_po = PositionBasePermission.objects.filter(position_id = user_profile.position_id, base_permission__codename='CAAPO', branch_company__code__in = company_in).values('branch_company__code')
    except:
        permiss_po  = None


    #ใบสั่งซื้อที่ fix ตามสิทธิ
    isPermiss = False
    isPermiss_po1 = False
    isPermiss_po2 = False

    #เคสที่ดึงมาจากใบเปรียบเทียบมีชื่อคนอนุมัติอยู่แล้ว
    if request.user.is_authenticated:
        try:
            #ดึงข้อมูล PurchaseOrder
            isPermiss_po1 = PurchaseOrder.objects.filter(id = po_id, approver_user = request.user, branch_company__code__in = company_in).exists() #หาสถานะรอดำเนินการของผู้อนุมัติ
        except PurchaseOrder.DoesNotExist:
            pass

    #ถ้าเป็นผู้อนุมัติที่มีสิทธิ
    if permiss_po:
        try:
            #ดึงข้อมูล PurchaseOrder
            isPermiss_po2 = PurchaseOrder.objects.filter(id = po_id, approver_status = 1, amount__isnull = False, amount__gt = 0, approver_user__isnull = True, cp__isnull = True, branch_company__code__in = permiss_po).exists() #หาสถานะรอดำเนินการของผู้อนุมัติ
        except PurchaseOrder.DoesNotExist:
            pass

    if isPermiss_po1 or isPermiss_po2:
        isPermiss = True

    '''
    if po_item:
        if permiss_po and not po_item.approver_user:
            isPermiss  = True
        elif po_item.approver_user == request.user:
            isPermiss  = True    
    '''

    if request.method == 'POST':
        post_status = request.POST['status'] or None
        status = BaseApproveStatus.objects.get(name = post_status)

        obj = PurchaseOrder.objects.get(id = po_id)
        #ผู้อนุมัติ
        if(isPermiss):
            obj.approver_status = status
            obj.approver_user_id = request.user.id
            obj.approver_update = datetime.datetime.now()
        obj.save()
        return redirect('editPOApprove', po_id = po_id, isFromHome = isFromHome)

    context = {
                'po':po,
                'items':items,
                'isPermiss':isPermiss,
                'isFromHome':isFromHome,
                'new_pr':new_pr,
                'bc':bc,
                'base_po_type':base_po_type,
                'ap_po_page': "tab-active",
                'ap_po_show': "show",
                active :"active show",
			    "disableTab":"disableTab",
			    "colorNav":"disableNav"
    }
    return render(request, 'purchaseOrderApprove/editPOApprove.html',context)

@login_required(login_url='signIn')
def viewComparePricePO(request):
    active = request.session['company_code']

    #data = ComparisonPrice.objects.filter(Q(examiner_status_id = 1) | Q(approver_status_id = 1))
    data = ComparisonPrice.objects.filter(Q(po_ref_no = "") & ~Q(examiner_status_id = 3) & ~Q(approver_status_id = 3) & ~Q(special_approver_status_id = 3) | (Q(examiner_status_id = 1) & Q(is_re_approve = True)), is_cancel = False, branch_company__code = active)

    prs = ComparisonPriceItem.objects.values('item__requisit__pr_ref_no','item__requisit__purchase_requisition_id','cp').annotate(Count('item')).order_by().filter(item__count__gt=0, cp__in=data)

    #กรองข้อมูล
    myFilter = ComparisonPriceFilter(request.GET, queryset = data)
    data = myFilter.qs

    bidder = ComparisonPriceDistributor.objects.values('cp_id','distributor__name','distributor__id','amount').filter(cp__in=data)
    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)
    context = {
        'cps':dataPage,
        'prs': prs,
        'filter':myFilter,
        'bidder':bidder,
        'cp_page': "tab-active",
        'cp_show': "show",
        active :"active show",
        "colorNav":"enableNav",
    }

    return render(request, 'comparePricePO/viewComparePricePO.html',context)

def createComparePricePO(request):
    form = ComparisonPriceForm(request.POST or None, initial={'organizer': request.user.id, 'approver_status':1, 'examiner_status':1})
    if form.is_valid():
        new_contact = form.save()
        return HttpResponseRedirect(reverse('createComparePricePOItem', args=(new_contact.pk, 'False')))

    context = {
        'form':form,
        'cp_page': "tab-active",
        'cp_show': "show",
    }
    return render(request, 'comparePricePO/createComparePricePO.html', context)

def editComparePricePO(request, cp_id):
    cp = ComparisonPrice.objects.get(id=cp_id)
    form = ComparisonPriceForm(instance=cp)
    if request.method == 'POST':
        form = ComparisonPriceForm(request.POST, instance=cp)
        if form.is_valid():
            form.save()
            return redirect('viewComparePricePO')

    context = {
        'form':form,
        'cp_page': "tab-active",
        'cp_show': "show",
    }

    return render(request, 'comparePricePO/createComparePricePO.html', context)

def removeComparePricePO(request, cp_id):
    cp = ComparisonPrice.objects.get(id = cp_id)

    bidder = ComparisonPriceDistributor.objects.filter(cp = cp_id)

    #ลบ item ใน ComparisonPriceItem ด้วย
    items = ComparisonPriceItem.objects.filter(cp = cp_id)
    for obj in items:
        try:
            d_item = RequisitionItem.objects.get(id = obj.item.id)
            #คำนวนหาจำนวนสินค้าในใบเปรียบเทียบและใบสั่งซื้อที่ทำไปแล้ว 24-09-2022
            cp_sum_quantity = calculateSumQuantityCPItem(obj.item, cp_id)
            sum_po_cp = calculateSumQuantityCPAndPOItem(obj.item)
            remain = d_item.quantity_pr - (sum_po_cp - cp_sum_quantity)

            d_item.quantity_used = sum_po_cp - cp_sum_quantity

            if remain <= 0:
                d_item.is_used = True
                d_item.save()
            else:
                d_item.is_used = False
                d_item.save()

            #เช็คว่าดึงรายการในใบขอเบิกไปใช้หมดหรือยัง
            check_item = RequisitionItem.objects.filter(requisit = d_item.requisit, is_used = False).distinct()
            if check_item:
                try:
                    pr = PurchaseRequisition.objects.get(requisition = d_item.requisit)
                    pr.is_complete = False
                    pr.save()
                except:
                    pass
        except RequisitionItem.DoesNotExist:
            pass
    items.delete()

    #ลบผู้จำหน่าย
    bidder.delete()

    #ลบใบเปรียบเทียบทีหลัง
    cp.delete()
    return redirect('viewComparePricePO')

def prepareComparePricePO(request):
    cp = ComparisonPrice.objects.create(
        organizer = request.user,
        approver_status_id = 1,
        examiner_status_id = 1,
    )
    cp.save()
    return HttpResponseRedirect(reverse('createComparePricePOItem', args=(cp.id, 'False')))

def createComparePricePOItem(request, cp_id, isReApprove):
    active = request.session['company_code']
    cp = ComparisonPrice.objects.get(id=cp_id)

    #ดึง item ที่ทำใบ po แล้ว
    #itemList = RequisitionItem.objects.filter(requisit__purchase_requisition_id__isnull = False, is_receive = False, product__isnull = False)
    #เปลี่ยนให้ดึงสินค้าเฉพาะที่ตัดมาจากใบขอซื้อแล้วเท่านั้น 14-09-2022
    itemList = ComparisonPriceItem.objects.values('item__id','item__product__id','item__product_name', 'item__requisit__pr_ref_no', 'item__quantity_pr', 'item__product__unit__id').filter(cp = cp_id)
    #
    company = BaseBranchCompany.objects.get(code = request.session['company_code'])
    #distributorList = Distributor.objects.filter(affiliated = company.affiliated)
    #distributorList = Distributor.objects.all().values('id','name','credit__id','vat_type__id')

    #ร้านแรก
    try:
        bidder = ComparisonPriceDistributor.objects.filter(cp = cp_id).order_by('id').first()
        bidder_oldest = bidder.id
    except (ComparisonPriceDistributor.DoesNotExist, AttributeError):
        bidder_oldest = None

    #ร้านทั้งหมดในใบเปรียบเทียบ
    try:
        bidder_have = ComparisonPriceDistributor.objects.filter(cp = cp_id)
    except (ComparisonPriceDistributor.DoesNotExist, AttributeError):
        bidder_have = None

    if request.method == 'GET':
        bookform = CPDModelForm(request.GET or None, initial={'cp': cp_id})
        formset = CPitemFormset(queryset=ComparisonPriceItem.objects.none())
        're approve cp'
        cp.is_re_approve = get_bool(isReApprove)
        if get_bool(isReApprove) == True:
            cp.approver_status_id = 1
            cp.approver_update = None

            cp.examiner_status_id = 1
            cp.examiner_update= None
            #ถ้าเป็นแบบพิเศษยอดเกิน 200,000
            if cp.is_special_approve_cm:
                cp.special_approver_user = None
                cp.special_approver_status_id = 1
                cp.special_approver_update = None
            else:
                cp.special_approver_user = None
                cp.special_approver_status_id = 4
                cp.special_approver_update = None

            cp.save()

    elif request.method == 'POST':
        bookform = CPDModelForm(request.POST or None, request.FILES, initial={'cp': cp_id})
        formset = CPitemFormset(request.POST)
        if bookform.is_valid() and formset.is_valid():
            # first save this book, as its reference will be used in `Author`
            book = bookform.save(commit=False)
            if not book.discount:
                book.discount = 0.00
            if not book.freight:
                book.freight = 0.00
            book.save()

            for form in formset:
                # so that `book` instance can be attached.
                if form.cleaned_data.get('item'):
                    cpi = form.save(commit=False)
                    cpi.bidder = book 
                    cpi.cp = cp_id
                    cpi.save()
                    #ตัดสินค้าหน้า PR
                    try:
                        r_item = RequisitionItem.objects.get(id = cpi.item.id)
                        #คำนวนหาจำนวนสินค้าในใบเปรียบเทียบและใบสั่งซื้อที่ทำไปแล้ว 24-09-2022
                        sum_po_cp = calculateSumQuantityCPAndPOItem(cpi.item)
                        remain = r_item.quantity_pr - sum_po_cp

                        r_item.quantity_used = sum_po_cp

                        if remain <= 0:
                            r_item.is_used = True
                            r_item.save()
                        else:
                            r_item.is_used = False
                            r_item.save()

                        #เช็คว่าดึงรายการในใบขอเบิกไปใช้หมดหรือยัง
                        checkRequisitionItemAllUsed(r_item.requisit)

                    except RequisitionItem.DoesNotExist:
                        pass
            return redirect('createComparePricePOItem', cp_id = cp_id, isReApprove = 'False')

    cpd = ComparisonPriceDistributor.objects.filter(cp = cp_id)
    cp_item = ComparisonPriceItem.objects.filter(cp = cp_id)
    
    #'distributorList': distributorList,
    context = {
        'link_cp_id':cp_id,
        'cpd':cpd,
        'cp_item':cp_item,
        'bidder_oldest':bidder_oldest,
        'bidder_have': bidder_have,
        'itemList': itemList,
        'bookform': bookform,
        'formset': formset,
        'cp_page': "tab-active",
        'cp_show': "show",
        active :"active show",
		"disableTab":"disableTab",
		"colorNav":"disableNav"
    }
    return render(request, 'comparePricePOItem/createComparePricePOItem.html',context)

def autocompalteDistributor(request):
    if 'term' in request.GET:
        term = request.GET.get('term')
        qs = Distributor.objects.filter(Q(id__icontains = term) | Q(name__icontains = term))[:15]
        titles = list()
        for obj in qs:
            titles.append(obj.id +"-"+ obj.name)
    return JsonResponse(titles, safe=False)

def searchDataDistributor(request):
    if 'id_distributor' in request.GET:
        id = request.GET.get('id_distributor')
        qs = Distributor.objects.get(id = id)

    data = {
        'credit': qs.credit.id,
        'vat_type': qs.vat_type.id,
    }
    return JsonResponse(data)

def getRateDistributor(request):
    if 'id_distributor' in request.GET:
        id = request.GET.get('id_distributor')
        company_code = request.GET.get('company_code')

        rate_counsel = RateDistributor.objects.filter(~Q(counsel = ""), distributor = id, counsel__isnull = False, po__branch_company__code = company_code).values('counsel','organizer_user__first_name','organizer_user__last_name','organizer_update').order_by('-id')[:5]
        num_grade_all = RateDistributor.objects.filter(distributor = id, po__branch_company__code = company_code).count()
        ratings = RateDistributor.objects.filter(distributor = id, po__branch_company__code = company_code).aggregate(avg_total_rate = Avg('total_rate'),  avg_price_rate = Avg('price_rate'), avg_quantity_rate = Avg('quantity_rate'), avg_duration_rate = Avg('duration_rate'), avg_service_rate = Avg('service_rate'), avg_safety_rate = Avg('safety_rate'))

    data = {
        'rate_counsel':list(rate_counsel),
        'ratings':ratings,
        'num_grade_all':num_grade_all,
    }
    return JsonResponse(data)

def setDataDistributor(request):
    if 'id_distributor' in request.GET:
        id = request.GET.get('id_distributor')
        qs = Distributor.objects.get(id = id)
        val = qs.id + "-" + qs.name
    data = {
        'val': val,
    }
    return JsonResponse(data)

def editComparePricePOItemFromPR(request, cp_id , cpd_id):
    active = request.session['company_code']

    #ดึง item ที่ทำใบ po แล้ว
    #itemList = RequisitionItem.objects.filter(requisit__purchase_requisition_id__isnull = False, is_receive = False, product__isnull = False)
    #เปลี่ยนให้ดึงสินค้าเฉพาะที่ตัดมาจากใบขอซื้อแล้วเท่านั้น 14-09-2022
    itemList = ComparisonPriceItem.objects.values('item__id','item__product__id','item__product_name', 'item__requisit__pr_ref_no', 'quantity', 'item__product__unit__id', 'unit__id').filter(cp = cp_id)

    #company = BaseBranchCompany.objects.get(code = request.session['company_code'])
    #distributorList = Distributor.objects.filter(affiliated = company.affiliated)
    #distributorList = Distributor.objects.all().values('id','name','credit__id','vat_type__id')

    data = ComparisonPriceDistributor.objects.get(id = cpd_id)
    if request.method == "POST":
        formset = CPitemInlineFormset(request.POST, request.FILES, instance=data)
        bookform = CPDModelForm(request.POST, request.FILES, instance=data)

        if formset.is_valid() and bookform.is_valid():
            # save ราคาใบ po
            book = bookform.save(commit=False)
            if not book.discount:
                book.discount = 0.00
            if not book.freight:
                book.freight = 0.00
            book.save()

            # save po item
            instances = formset.save(commit=False)
            for instance in instances:
                instance.bidder = book
                instance.cp = cp_id
                instance.save()
                #ตัดสินค้าหน้า PR
                try:
                    r_item = RequisitionItem.objects.get(id = instance.item.id)
                    #คำนวนหาจำนวนสินค้าในใบเปรียบเทียบและใบสั่งซื้อที่ทำไปแล้ว 24-09-2022
                    sum_po_cp = calculateSumQuantityCPAndPOItem(instance.item)
                    remain = r_item.quantity_pr - sum_po_cp

                    r_item.quantity_used = sum_po_cp

                    if remain <= 0:
                        r_item.is_used = True
                        r_item.save()
                    else:
                        r_item.is_used = False
                        r_item.save()
                    
                    #เช็คว่าดึงรายการในใบขอเบิกไปใช้หมดหรือยัง
                    checkRequisitionItemAllUsed(r_item.requisit)

                except RequisitionItem.DoesNotExist:
                    pass
            for obj in formset.deleted_objects:
                try:
                    d_item = RequisitionItem.objects.get(id = obj.item.id)
                    d_item.quantity_used = calculateSumQuantityCPAndPOItem(obj.item) - obj.quantity
                    d_item.is_used = False
                    d_item.save()

                    #เช็คว่าดึงรายการในใบขอเบิกไปใช้หมดหรือยัง
                    checkRequisitionItemAllUsed(d_item.requisit)
                except RequisitionItem.DoesNotExist:
                    pass
                obj.delete()
            formset.save_m2m()
            return redirect('createComparePricePOItem', cp_id = cp_id, isReApprove = 'False')
    else:
        bookform = CPDModelForm(instance=data)
        formset = CPitemInlineFormset(instance=data)

    cpd = ComparisonPriceDistributor.objects.filter(cp = cp_id)
    cp_item = ComparisonPriceItem.objects.filter(cp = cp_id)

    #'distributorList': distributorList,
    context = {
        'link_cp_id' : cp_id,
        'link_cpd_id' : cpd_id,
        'cpd':cpd,
        'cp_item':cp_item,
        'itemList':itemList,
        'bookform': bookform,
        'formset': formset,
        'cp_page': "tab-active",
        'cp_show': "show",
        active :"active show",
		"disableTab":"disableTab",
		"colorNav":"disableNav"
    }
    return render(request, 'comparePricePOItem/editComparePricePOItemFromPR.html',context)

def editComparePricePOItem(request, cp_id , cpd_id):
    active = request.session['company_code']
    cp = ComparisonPrice.objects.get(id=cp_id)

    #ดึง item ที่ทำใบ po แล้ว
    #itemList = RequisitionItem.objects.filter(requisit__purchase_requisition_id__isnull = False, is_receive = False, product__isnull = False)
    #เปลี่ยนให้ดึงสินค้าเฉพาะที่ตัดมาจากใบขอซื้อแล้วเท่านั้น 14-09-2022
    itemList = ComparisonPriceItem.objects.values('item__id','item__product__id','item__product_name', 'item__requisit__pr_ref_no', 'item__quantity_pr', 'item__product__unit__id').filter(cp = cp_id)
    #
    company = BaseBranchCompany.objects.get(code = request.session['company_code'])
    #distributorList = Distributor.objects.filter(affiliated = company.affiliated)
    #distributorList = Distributor.objects.all().values('id','name','credit__id','vat_type__id')

    data = ComparisonPriceDistributor.objects.get(id = cpd_id)
    numDistributor = ComparisonPriceDistributor.objects.filter(cp = cp_id).count()
    if request.method == "POST":
        formset = CPitemInlineFormset(request.POST, request.FILES, instance=data)
        bookform = CPDModelForm(request.POST, request.FILES, instance=data)

        if formset.is_valid() and bookform.is_valid():
            # save ราคาใบ po
            book = bookform.save(commit=False)
            if not book.discount:
                book.discount = 0.00
            book.save()

            # save po item
            instances = formset.save(commit=False)
            for instance in instances:
                instance.bidder = book 
                instance.cp = cp_id
                instance.save()
                #ตัดสินค้าหน้า PR
                try:
                    r_item = RequisitionItem.objects.get(id = instance.item.id)
                    #คำนวนหาจำนวนสินค้าในใบเปรียบเทียบและใบสั่งซื้อที่ทำไปแล้ว 24-09-2022
                    sum_po_cp = calculateSumQuantityCPAndPOItem(instance.item)
                    remain = r_item.quantity_pr - sum_po_cp
                    r_item.quantity_used = sum_po_cp

                    if remain <= 0:
                        r_item.is_used = True
                        r_item.save()
                        #เช็คว่าดึงรายการในใบขอเบิกไปใช้หมดหรือยัง
                        check_item = RequisitionItem.objects.filter(requisit = r_item.requisit, is_used = False).distinct()
                        if not check_item:
                            try:
                                pr = PurchaseRequisition.objects.get(requisition = r_item.requisit)
                                pr.is_complete = True
                                pr.save()
                            except:
                                pass
                    else:
                        r_item.is_used = False
                        r_item.save()
                        pr = PurchaseRequisition.objects.get(requisition = r_item.requisit)
                        pr.is_complete = False
                        pr.save()
                        
                except RequisitionItem.DoesNotExist:
                    pass

            for obj in formset.deleted_objects:
                try:
                    d_item = RequisitionItem.objects.get(id = obj.item.id)
                    d_item.quantity_used = calculateSumQuantityCPAndPOItem(obj.item) - obj.quantity
                    d_item.is_used = False
                    d_item.save()
                    #เช็คว่าดึงรายการในใบขอเบิกไปใช้หมดหรือยัง
                    check_item = RequisitionItem.objects.filter(requisit = d_item.requisit, is_used = False).distinct()
                    if check_item:
                        try:
                            pr = PurchaseRequisition.objects.get(requisition = d_item.requisit)
                            pr.is_complete = False
                            pr.save()
                        except:
                            pass
                except RequisitionItem.DoesNotExist:
                    pass   
                obj.delete()
            formset.save_m2m()
            return redirect('createComparePricePOItem',cp_id = cp_id, isReApprove = 'False')
    else:
        bookform = CPDModelForm(instance=data)
        formset = CPitemInlineFormset(instance=data)

    cpd = ComparisonPriceDistributor.objects.filter(cp = cp_id)
    cp_item = ComparisonPriceItem.objects.filter(cp = cp_id)

    #'distributorList': distributorList,
    context = {
        'link_cp_id' : cp_id,
        'link_cpd_id' : cpd_id,
        'cpd':cpd,
        'cp_item':cp_item,
        'itemList':itemList,
        'bookform': bookform,
        'formset': formset,
        'cp_page': "tab-active",
        'cp_show': "show",
        active :"active show",
		"disableTab":"disableTab",
		"colorNav":"disableNav"
    }
    return render(request, 'comparePricePOItem/editComparePricePOItem.html',context)

def removeComparePriceDistributor(request, cp_id, cpd_id):
    numDistributor = ComparisonPriceDistributor.objects.filter(cp = cp_id).count()
    cpd = ComparisonPriceDistributor.objects.get(id = cpd_id)
    #ลบ item ใน PurchaseOrder ด้วย
    items = ComparisonPriceItem.objects.filter(bidder = cpd)

    cp = ComparisonPrice.objects.get(id  = cpd.cp.id)
    if numDistributor == 1:
        if cp.is_re_approve:
            messages.warning(request, "ไม่สามารถลบรายการได้ เนื่องจากเป็นใบเปรียบเทียบราคาฉบับแก้ไข\nกรุณายกเลิกรายการพร้อมใส่เหตุผลการยกเลิก")
            redirect('createComparePricePOItem', cp_id = cp_id, isReApprove = 'False')
        else:
            for obj in items:
                try:
                    d_item = RequisitionItem.objects.get(id = obj.item.id)
                    #คำนวนหาจำนวนสินค้าในใบเปรียบเทียบและใบสั่งซื้อที่ทำไปแล้ว 24-09-2022
                    cp_sum_quantity = calculateSumQuantityCPItem(obj.item, cp_id)
                    sum_po_cp = calculateSumQuantityCPAndPOItem(obj.item)
                    remain = d_item.quantity_pr - (sum_po_cp - cp_sum_quantity)

                    if remain <= 0:
                        d_item.is_used = True
                        d_item.save()
                    else:
                        d_item.is_used = False
                        d_item.save()

                    #เช็คว่าดึงรายการในใบขอเบิกไปใช้หมดหรือยัง
                    check_item = RequisitionItem.objects.filter(requisit = d_item.requisit, is_used = False).distinct()
                    if check_item:
                        try:
                            pr = PurchaseRequisition.objects.get(requisition = d_item.requisit)
                            pr.is_complete = False
                            pr.save()
                        except:
                            pass
                except RequisitionItem.DoesNotExist:
                    pass

            items.delete()
            #ลบ PurchaseOrder ทีหลัง
            cpd.delete()
    else:
        items.delete()
        #ลบ PurchaseOrder ทีหลัง
        cpd.delete()

    return redirect('createComparePricePOItem', cp_id = cp_id, isReApprove = 'False')

def printComparePricePO(request, cp_id):
    active = request.session['company_code']
    bc = BaseBranchCompany.objects.get(code = active)

    try:
        bidder_oldest = ComparisonPriceDistributor.objects.filter(cp = cp_id).order_by('id').first()
        items_oldest = ComparisonPriceItem.objects.filter(bidder = bidder_oldest.id)
    except (ComparisonPriceDistributor.DoesNotExist, ComparisonPriceItem.DoesNotExist, AttributeError):
        items_oldest = None

    pr_ref_no = ""
    new_pr_id = dict()
    if items_oldest:
        for obj in items_oldest:
            if obj.item.requisit.purchase_requisition_id not in new_pr_id:
                new_pr_id[obj.item.requisit.purchase_requisition_id] = obj

    new_pr = dict()
    for id in new_pr_id:
        try:
            pr = PurchaseRequisition.objects.get(id = id)
            pr_ref_no += pr.ref_no + ", "
            new_pr[pr] = pr
        except PurchaseRequisition.DoesNotExist:
            pass 

    cp = ComparisonPrice.objects.get(id=cp_id)
    form = CPSelectBidderForm(instance=cp)
    if request.method == 'POST':
        form = CPSelectBidderForm(request.POST, instance=cp)
        if form.is_valid():
            f_cp = form.save()
            
            #เปลี่ยนทุกร้านค้าให้ไม่เลือกก่อน
            all_bidder = ComparisonPriceDistributor.objects.filter(cp = f_cp)
            for bd in all_bidder:
                bd.is_select = False
                bd.save()

            #set ร้านค้าที่เลือก
            select_bidder = ComparisonPriceDistributor.objects.get(cp = f_cp, distributor = f_cp.select_bidder)
            select_bidder.is_select = True
            #save
            select_bidder.save()

            #เก็บวันที่เลือกร้านครั้งแรก
            if f_cp.select_bidder and f_cp.select_bidder_update is None:
                f_cp.select_bidder_update = datetime.datetime.now()

            #หาว่าเป็นใบเปรียบเทียบ(ยอดเกิน 200,000) แบบอนุมัติ 2 คนหรือไม่ ถ้าเป็นให้ set special_approver_status เป็นรอดำเนินการ
            isSpecialCP = is_special_approver_cp(cp.id)
            if isSpecialCP:
                f_cp.special_approver_status_id = 1
                f_cp.is_special_approve_cm = True
            else:
                f_cp.special_approver_status_id = 4
                f_cp.is_special_approve_cm = False
            #save
            f_cp.save()
            return redirect('viewComparePricePO')

    bidder = ComparisonPriceDistributor.objects.filter(cp = cp_id).order_by('amount', 'cp')
    itemName = ComparisonPriceItem.objects.filter(cp = cp_id).order_by('bidder__amount', 'id')

    baseSparesType = BaseSparesType.objects.all()
    context = {
        'cp':cp,
        'bidder' : bidder,
        'items_oldest' : items_oldest,
        'itemName':itemName,
        'pr_ref_no': pr_ref_no,
        'new_pr': new_pr,
        'form':form,
        'baseSparesType' : baseSparesType,
        'bc':bc,
        'cp_page': "tab-active",
        'cp_show': "show",
        active :"active show",
		"disableTab":"disableTab",
		"colorNav":"disableNav"
    }
    return render(request, 'comparePricePO/printComparePricePO.html',context)

def showComparePricePO(request, cp_id, mode):
    active = request.session['company_code']
    try:
        bidder_oldest = ComparisonPriceDistributor.objects.filter(cp = cp_id).order_by('id').first()
        items_oldest = ComparisonPriceItem.objects.filter(bidder = bidder_oldest.id)
    except (ComparisonPriceDistributor.DoesNotExist, ComparisonPriceItem.DoesNotExist, AttributeError):
        items_oldest = None

    cp = ComparisonPrice.objects.get(id=cp_id)
    bc = BaseBranchCompany.objects.get(code = cp.branch_company)

    baseSparesType = BaseSparesType.objects.all()

    bidder = ComparisonPriceDistributor.objects.filter(cp = cp_id).order_by('amount', 'cp')
    itemName = ComparisonPriceItem.objects.filter(cp = cp_id).order_by('bidder__amount', 'id')

    #หาว่าเป็นใบเปรียบเทียบ(ยอดเกิน 200,000) แบบอนุมัติ 2 คนหรือไม่
    isSpecialCP = is_special_approver_cp(cp_id)

    pr_ref_no = ""
    new_pr_id = dict()
    if items_oldest:
        for obj in items_oldest:
            if obj.item.requisit.purchase_requisition_id not in new_pr_id:
                new_pr_id[obj.item.requisit.purchase_requisition_id] = obj
    
    new_pr = dict()
    for id in new_pr_id:
        try:
            pr = PurchaseRequisition.objects.get(id = id)
            pr_ref_no += pr.ref_no + ", "
            new_pr[pr] = pr
        except PurchaseRequisition.DoesNotExist:
            pass

    #ถ้า user login เป็นจัดซื้อ
    isPurchasing = is_purchasing(request.user)

    form = ComparisonPriceAddressCompanyForm(instance=cp)
    if request.method == 'POST' and 'btnformCPa' in request.POST:
        form = ComparisonPriceAddressCompanyForm(request.POST, request.FILES, instance=cp)
        if form.is_valid():
            form.save()
            return redirect('showComparePricePO', cp_id = cp_id , mode = mode)
        
    cancel_form = CPCancelForm(instance=cp)
    if request.method == 'POST' and 'btnformCancelCP' in request.POST:
        cancel_form = CPCancelForm(request.POST, instance=cp)
        if cancel_form.is_valid():
            cc = cancel_form.save(commit=False)
            if cc.cancel_reason is not None:
                cc.is_cancel = True
                cc.save()
                #rePOItemIsUse(po_id)#หากยกเลิกรายการ เช็ค item และนำใบขอซื้อกลับมาทำใหม่
            return redirect('showComparePricePO', cp_id = cp_id , mode = mode)

    page = CPPageMode(mode)
    show = CPShowMode(mode)
    context = {
            'cp':cp,
            'bidder' : bidder,
            'items_oldest' : items_oldest,
            'itemName':itemName,
            'baseSparesType':baseSparesType,
            'pr_ref_no': pr_ref_no,
            'new_pr': new_pr,
            'isPurchasing':isPurchasing,
            'isSpecialCP':isSpecialCP,
            'form':form,
            'cancel_form':cancel_form,
            'bc':bc,
            page: "tab-active",
            show: "show",
            active :"active show",
			"disableTab":"disableTab",
			"colorNav":"disableNav"
    }
    return render(request, 'comparePricePO/showComparePricePO.html',context)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def createPOFromComparisonPrice(request, cp_id):
    active = request.session['company_code']
    company = BaseBranchCompany.objects.get(code = active)


    #ถ้า user login แก้ไขรหัสใบสั่งซื้อได้
    isEditPO = is_edit_po_id(request.user)

    #ถ้า user login แก้ไขผู้อนุมัติบสั่งซื้อได้
    isEditApproverUserPO = is_edit_approver_user_po(request.user)

    bc = ComparisonPrice.objects.get(id = cp_id)

    #ถ้าเป็นยอดเกิน 200,000 เปลี่ยนคนอนุมัติ
    if bc.is_special_approve_cm:
        approver_user = bc.special_approver_user
    else:
        approver_user = bc.approver_user

    #ประเมินร้านค้า
    rate_counsel = RateDistributor.objects.filter(~Q(counsel = ""), distributor = bc.select_bidder, counsel__isnull = False, po__branch_company__code = company).values('counsel','organizer_user__first_name','organizer_user__last_name','organizer_update').order_by('-id')[:5]
    num_grade_all = RateDistributor.objects.filter(distributor = bc.select_bidder, po__branch_company__code = company).count()
    ratings = RateDistributor.objects.filter(distributor = bc.select_bidder, po__branch_company__code = company).aggregate(avg_total_rate = Avg('total_rate'),  avg_price_rate = Avg('price_rate'), avg_quantity_rate = Avg('quantity_rate'), avg_duration_rate = Avg('duration_rate'), avg_service_rate = Avg('service_rate'), avg_safety_rate = Avg('safety_rate'))

    #set ค่าเริ่มต้นของเจ้าหน้าที่พัสดุ และสเตตัสการอนุมัติเป็น รอดำเนินการ
    form = PurchaseOrderFromComparisonPriceForm(request, request.POST or None, initial={'cp': cp_id , 'address_company': bc.address_company , 'approver_user' : approver_user})
    form_rate = RateDistributorForm(request.POST or None, initial={'distributor': bc.select_bidder})
    if form.is_valid() and form_rate.is_valid():
        try:
            new_contact = form.save(commit=False)
            cp = ComparisonPrice.objects.get(id = new_contact.cp.id)
            cpd = ComparisonPriceDistributor.objects.get(cp = new_contact.cp.id, distributor = cp.select_bidder)
            new_contact.distributor = cp.select_bidder
            new_contact.credit = cpd.credit
            new_contact.stockman_user = cp.organizer
            new_contact.vat_type = cpd.vat_type
            new_contact.quotation_pdf = cpd.quotation_pdf
            new_contact.approver_status_id = 1
            #new_contact.approver_user = cp.approver_user
            new_contact.branch_company = company

            rate_contact = form_rate.save(commit=False)
            rate_contact.organizer_user = cp.organizer
            rate_contact.distributor = cp.select_bidder
            rate_contact.po = new_contact
        except:
            return HttpResponseRedirect(reverse('createPOFromComparisonPrice', args=(cp_id,)))
        else:
            new_contact.save()
            rate_contact.save()

            cp.po_ref_no = new_contact.ref_no
            cp.save()
            return HttpResponseRedirect(reverse('createPOItemFromComparisonPrice', args=(new_contact.pk,)))

    context = {
        'form':form,
        'form_rate':form_rate,
        'isEditPO':isEditPO,
        'isEditApproverUserPO':isEditApproverUserPO,
        'cp':bc,
        'rate_counsel':rate_counsel,
        'num_grade_all':num_grade_all,
        'ratings':ratings,
        'po_page': "tab-active",
        'po_show': "show",
        active :"active show",
		"disableTab":"disableTab",
		"colorNav":"disableNav"
    }

    return render(request, "purchaseOrder/createPO.html", context)

def createPOItemFromComparisonPrice(request, po_id):
    active = request.session['company_code']
    bc = BaseBranchCompany.objects.get(code = active)

    base_po_type = BasePOType.objects.all()

    template_name = 'purchaseOrderItem/createPOItemFromComparisonPrice.html'
    heading_message = 'Model Formset Demo'
    po_data = PurchaseOrder.objects.get(id=po_id)
    po_items = PurchaseOrderItem.objects.filter(po = po_id)
    cp = ComparisonPrice.objects.get(id = po_data.cp.id)

    #ดึง item ที่ทำใบ po แล้ว
    #itemList = RequisitionItem.objects.filter(requisit__purchase_requisition_id__isnull = False, is_receive = False, product__isnull = False)
    #เปลี่ยนให้ดึงสินค้าเฉพาะที่ตัดมาจากใบขอซื้อแล้วเท่านั้น 14-09-2022
    itemList = ComparisonPriceItem.objects.values('item__id','item__product__id','item__product_name', 'item__requisit__pr_ref_no', 'item__quantity_pr', 'item__product__unit__id', 'unit__id').filter(cp = cp.id)

    cp_item = ComparisonPriceItem.objects.filter(cp = po_data.cp.id, bidder__distributor = po_data.cp.select_bidder)
    cpd_price = ComparisonPriceDistributor.objects.get(cp = po_data.cp.id, distributor = po_data.cp.select_bidder)
    if request.method == 'GET':
        formset = PurchaseOrderItemModelFormset(queryset=PurchaseOrderItem.objects.none())
        price_form = PurchaseOrderPriceForm()
    elif request.method == 'POST':
        formset = PurchaseOrderItemModelFormset(request.POST)
        price_form = PurchaseOrderPriceForm(request.POST, instance=po_data)
        if formset.is_valid() and price_form.is_valid():

            # save ราคาใบ po
            price_form.save()
            # save po item
            for form in formset:
                # only save if name is present
                if form.cleaned_data.get('item'):
                    obj = form.save(commit=False)
                    obj.po = po_data
                    obj.save()
            #redirect นอก loop
            return redirect('viewPO')

    po = PurchaseOrder.objects.get(id = po_id)
    items = PurchaseOrderItem.objects.filter(po = po_id)
    new_pr_id = dict()
    if items:
        for obj in items:
            if obj.item.requisit.purchase_requisition_id not in new_pr_id:
                new_pr_id[obj.item.requisit.purchase_requisition_id] = obj

    new_pr = dict()
    for id in new_pr_id:
        try:
            pr = PurchaseRequisition.objects.get(id = id)
            new_pr[pr] = pr
        except PurchaseRequisition.DoesNotExist:
            pass

    context = {
        'po':po,
        'po_items':po_items,
        'cp_item':cp_item,
        'cpd_price':cpd_price,
        'po_page': "tab-active",
        'po_show': "show",
        'formset': formset,
        'price_form': price_form,
        'itemList':itemList,
        'new_pr':new_pr,
        'heading': heading_message,
        'bc':bc,
        'base_po_type':base_po_type,
        active :"active show",
		"disableTab":"disableTab",
		"colorNav":"disableNav"
    }
    return render(request, template_name, context)

@login_required(login_url='signIn')
def viewCPApprove(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)
    data = ComparisonPrice.objects.filter(Q(examiner_status_id = 1) | Q(approver_status_id = 1) & ~Q(examiner_status_id = 3), select_bidder_id__isnull = False, branch_company__code__in = company_in)

    #กรองข้อมูล
    myFilter = ComparisonPriceFilter(request.GET, queryset = data)
    data = myFilter.qs

    bidder = ComparisonPriceDistributor.objects.filter(cp__in=data)
    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)
    context = {
        'cps':dataPage,
        'filter':myFilter,
        'bidder':bidder,
        'ap_cp_page': "tab-active",
        'ap_cp_show': "show",
        active :"active show",
        "colorNav":"enableNav"
    }
    return render(request, "comparePricePOApprove/viewCPApprove.html",context)

def printCPApprove(request, cp_id, isFromHome):
    active = request.session['company_code']

    try:
        company_in = findCompanyIn(request)
    except:
        company_in = BaseBranchCompany.objects.filter(code = active).values('code')

    try:
        bidder_oldest = ComparisonPriceDistributor.objects.filter(cp = cp_id).order_by('id').first()
        items_oldest = ComparisonPriceItem.objects.filter(bidder = bidder_oldest.id)
    except (ComparisonPriceDistributor.DoesNotExist, ComparisonPriceItem.DoesNotExist, AttributeError):
        items_oldest = None

    cp = ComparisonPrice.objects.get(id=cp_id)
    bc = BaseBranchCompany.objects.get(code = cp.branch_company)

    baseSparesType = BaseSparesType.objects.all()

    bidder = ComparisonPriceDistributor.objects.filter(cp = cp_id).order_by('amount', 'cp')
    itemName = ComparisonPriceItem.objects.filter(cp = cp_id).order_by('bidder__amount', 'id')

    isApprover = False
    isExaminer = False
    isApproverSpecial = False

    #หาว่าเป็นใบเปรียบเทียบ(ยอดเกิน 200,000) แบบอนุมัติ 3 คนหรือไม่เพื่อทำ form อนุมัติแบบอนุมัติ 3
    isSpecialCP = is_special_approver_cp(cp_id)

    try:
        cpd_select = ComparisonPriceDistributor.objects.get(cp = cp, distributor = cp.select_bidder)
    except ComparisonPriceDistributor.DoesNotExist:
        cpd_select = None


    try:
        ex_item = ComparisonPrice.objects.get(id = cp_id, select_bidder = cpd_select.distributor, examiner_status = 1, examiner_user = request.user)
        if(ex_item):
            isExaminer = True
    except ComparisonPrice.DoesNotExist:
        ex_item = None

    try:
        ap_item = ComparisonPrice.objects.get(id = cp_id, select_bidder = cpd_select.distributor, approver_status = 1, approver_user = request.user)
        if(ap_item):
            isApprover = True
    except ComparisonPrice.DoesNotExist:
        ap_item = None

    #หาว่าเป็นคนที่มีสิทธิอนุมัติคนที่ 2 ในใบเปรียบเทียบนี้หรือไม่
    #get permission with position login
    try:
        user_profile = UserProfile.objects.get(user_id = request.user.id)

        permiss = BasePermission.objects.filter(codename ='CASCP')
        permiss_cm = PositionBasePermission.objects.filter(position_id = user_profile.position_id, base_permission__codename='CASCP', branch_company__code__in = company_in).values('branch_company__code')
    except:
        permiss_cm  = None

    try:
        in_company = BaseBranchCompany.objects.filter(userprofile = user_profile, code__in = permiss_cm).exists()
    except:
        pass

    #ใบสั่งซื้อที่ fix ตามสิทธิ

    #เคสที่ดึงมาจากใบเปรียบเทียบมีชื่อคนอนุมัติอยู่แล้ว
    if permiss_cm and in_company:
        try:
            #ดึงข้อมูล PurchaseOrder
            isApproverSpecial = ComparisonPrice.objects.filter(id = cp_id, examiner_status = 2, approver_status = 2, special_approver_status = 1, is_special_approve_cm = True, branch_company__code__in = permiss_cm).exists() #หาสถานะรอดำเนินการของผู้อนุมัติ
        except ComparisonPrice.DoesNotExist:
            pass


    ''' ผู้ตรวจสอบและผู้อนุมัติใบเปรียบเทียบราคาแบบเก่าเปลี่ยนเป็นแบบ fix ชื่อ
    #get permission with position login
    #ผู้ตรวจสอบ
    try:
        user_profile = UserProfile.objects.get(user_id = request.user.id)
        permiss = BasePermission.objects.filter(codename__in= ['CAECP1','CAECP2','CAECP3','CAECP4','CAECPD','CAECPA'])
        isPermissAE = PositionBasePermission.objects.filter(position_id = user_profile.position_id, base_permission__codename__in= ['CAECP1','CAECP2','CAECP3','CAECP4','CAECPD','CAECPA']).prefetch_related(Prefetch('base_permission', queryset=permiss)).values('base_permission')
    except:
        isPermissAE = None

    pmAE = []
    if(isPermissAE):
        for i in isPermissAE:
            obj = BasePermission.objects.get(id = i['base_permission'])
            pmAE.append(obj)

    #ผู้อนุมัติ
    try:
        permiss = BasePermission.objects.filter(codename__in= ['CAACP1','CAACP2','CAACP3','CAACP4','CAACPD','CAACPA'])
        isPermissAA = PositionBasePermission.objects.filter(position_id = user_profile.position_id, base_permission__codename__in= ['CAACP1','CAACP2','CAACP3','CAACP4','CAACPD','CAACPA']).prefetch_related(Prefetch('base_permission', queryset=permiss)).values('base_permission')
    except:
        isPermissAA = None

    pmAA = []
    if(isPermissAA):
        for i in isPermissAA:
            obj = BasePermission.objects.get(id = i['base_permission'])
            pmAA.append(obj)


    if isPermissAA and cpd_select:
        for aa in pmAA:
            if aa.codename == 'CAACPD' or aa.codename == 'CAACPA':
                try:
                    if aa.codename == 'CAACPD' and cp.cm_type.id == '1':
                        isApprover = True
                    if aa.codename == 'CAACPA' and cp.cm_type.id == '2':
                        isApprover = True
                except:
                    pass
            else:
                if aa.ap_amount_min <= cpd_select.amount <= aa.ap_amount_max and cpd_select.cp.cm_type is None:
                    isApprover = True


    if isPermissAE and cpd_select:
        for ae in pmAE:
            if ae.codename == 'CAECPD' or ae.codename == 'CAECPA':
                try:
                    if ae.codename == 'CAECPD' and cp.cm_type.id == '1':
                        isExaminer = True
                    if ae.codename == 'CAECPA' and cp.cm_type.id == '2':
                        isExaminer = True
                except:
                    pass
            else:
                if ae.ap_amount_min <= cpd_select.amount <= ae.ap_amount_max and cpd_select.cp.cm_type is None:
                    isExaminer = True

    '''

    pr_ref_no = ""
    new_pr_id = dict()
    if items_oldest:
        for obj in items_oldest:
            if obj.item.requisit.purchase_requisition_id not in new_pr_id:
                new_pr_id[obj.item.requisit.purchase_requisition_id] = obj

    new_pr = dict()
    for id in new_pr_id:
        try:
            pr = PurchaseRequisition.objects.get(id = id)
            pr_ref_no += pr.ref_no + ", "
            new_pr[pr] = pr
        except PurchaseRequisition.DoesNotExist:
            pass

    if request.method == 'POST':
        post_status = request.POST['status'] or None
        status = BaseApproveStatus.objects.get(name = post_status)

        obj = ComparisonPrice.objects.get(id = cp_id)
        #ผู้อนุมัติ
        if(isApprover and isExaminer):
            obj.examiner_status = status
            obj.examiner_user_id = request.user.id
            obj.examiner_update = datetime.datetime.now()

            obj.approver_status = status
            obj.approver_user_id = request.user.id
            obj.approver_update = datetime.datetime.now()
        elif(isApprover):
            obj.approver_status = status
            obj.approver_user_id = request.user.id
            obj.approver_update = datetime.datetime.now()
        elif(isExaminer):
            obj.examiner_status = status
            obj.examiner_user_id = request.user.id
            obj.examiner_update = datetime.datetime.now()
        elif(isApproverSpecial):
            obj.special_approver_status = status
            obj.special_approver_user_id = request.user.id
            obj.special_approver_update = datetime.datetime.now()
        obj.save()
        return redirect('printCPApprove', cp_id = cp_id, isFromHome = isFromHome)

    context = {
        'cp':cp,
        'bidder' : bidder,
        'items_oldest' : items_oldest,
        'itemName':itemName,
        'baseSparesType':baseSparesType,
        'new_pr': new_pr,
        'pr_ref_no': pr_ref_no,
        'isApprover': isApprover,
        'isExaminer': isExaminer,
        'isFromHome': isFromHome,
        'isApproverSpecial': isApproverSpecial,
        'isSpecialCP':isSpecialCP,
        'bc':bc,
        'ap_cp_page': "tab-active",
        'ap_cp_show': "show",
        active :"active show",
		"disableTab":"disableTab",
		"colorNav":"disableNav"
    }
    return render(request, 'comparePricePOApprove/printCPApprove.html',context)

def searchItemExpress(request):
    strName = "<ul class='list-group list-group-flush'>"
    name = request.GET.get('title', None)
    product = Product.objects.filter(name__icontains=name)
    index = 1
    for i in product:
        strName += "<li class='list-group-item'>"+ str(index) + ")  " + i.id + "  " + i.name + "</li>"
        index += 1
    strName += "</ul>"

    data = {
        'instance': strName,
    }
    return JsonResponse(data)

def viewReceive(request):
    active = request.session['company_code']
    data = PurchaseOrder.objects.filter(approver_status_id = 2, is_receive = False, is_cancel = False, branch_company__code = active)

    #กรองข้อมูล
    myFilter = PurchaseOrderFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)

    prs = PurchaseOrderItem.objects.filter(po__in=data).values('item__requisit__pr_ref_no','item__requisit__purchase_requisition_id','po')
    context = {
        'pos':dataPage,
        'filter':myFilter,
        'prs':prs,
        'rc_page': "tab-active",
        'rc_show': "show",
        active :"active show",
        "colorNav":"enableNav",
    }
    return render(request, 'receive/viewReceive.html',context)

def removeReceive(request, rc_id):
    rc = Receive.objects.get(id = rc_id)
    #ลบ item ใน PurchaseOrder ด้วย
    items = ReceiveItem.objects.filter(rc = rc)
    items.delete()
    #ลบ PurchaseOrder ทีหลัง
    rc.delete()
    return redirect('viewReceive')

def createReceive(request):
    #set ค่าเริ่มต้นของผู้รับสินค้า
    form = ReceiveForm(request.POST or None, initial={'receive_user': request.user})
    if form.is_valid():
        new_contact = form.save(commit=False)
        po = PurchaseOrder.objects.get(id = new_contact.po_id)
        new_contact.total_price = po.total_price
        new_contact.discount = po.discount
        new_contact.total_after_discount = po.total_after_discount
        new_contact.vat = po.vat
        new_contact.amount = po.amount
        new_contact.freight = po.freight
        new_contact.save()

        #create receive item
        items = PurchaseOrderItem.objects.filter(po = new_contact.po)
        if items:
            for item in items:
                rc_item = ReceiveItem.objects.create(
                    item = item.item,
                    quantity = item.quantity,
                    unit = item.unit,
                    unit_price = item.unit_price,
                    price = item.price,
                    rc = new_contact,
                )
                rc_item.save()

        return HttpResponseRedirect(reverse('editReceiveItem', args=(new_contact.pk,)))
    context = {
        'form' : form,
        'rc_page': "tab-active",
        'rc_show': "show",
    }
    return render(request, 'receive/createReceive.html',context)

def editReceiveItem(request, rc_id):
    rc = Receive.objects.get(id = rc_id)
    items = PurchaseOrderItem.objects.filter(po = rc.po)

    if request.method == "POST":
        formset = ReceiveItemInlineFormset(request.POST, request.FILES, instance=rc)
        receive_form = ReceivePriceForm(request.POST, instance=rc)
        if formset.is_valid() and receive_form.is_valid():
            # save ราคาใบ po
            receive_form.save()
            # save po item
            instances = formset.save(commit=False)
            for instance in instances:
                instance.save()
            formset.save_m2m()
            return redirect('viewReceive')
    else:
        formset = ReceiveItemInlineFormset(instance=rc)
        receive_form = ReceivePriceForm(instance=rc)

    context = {
        'rc': rc,
        'items': items,
        'receive_form': receive_form,
        'formset': formset,
        'rc_page': "tab-active",
        'rc_show': "show",
    }
    return render(request, 'receiveItem/editReceiveItem.html',context)

#ดึงรายการใบขอซื้อกลับมาทำใหม่
def reBuyPR(request, pr_id):
    active = request.session['company_code']
    pr = PurchaseRequisition.objects.get(id = pr_id)

    try:
        pr_item = RequisitionItem.objects.filter(requisit = pr.requisition)
        for item in pr_item:
            item.is_used = False
            item.save()

        pr.is_complete = False
        pr.note = pr.note[2:]
        pr.save()

        return HttpResponseRedirect(reverse('createCMorPO', args=(pr_id,)))
    except RequisitionItem.DoesNotExist:
        pass
    
    context = {
        'pr_page': "tab-active",
        'pr_show': "show",
        active :"active show",
        "disableTab":"disableTab",
        "colorNav":"disableNav"
    }

    return render(request, "purchaseRequisition/showPR.html", context)

def reApprovePR(request, pr_id):
        active = request.session['company_code']
        bc = BaseBranchCompany.objects.get(code = active)

        pr = PurchaseRequisition.objects.get(id = pr_id)

        pr.approver_status_id = 1 #กำหนดค่าเริ่มต้น
        pr.approver_update = None
        
        #พัสดุ
        #pr.stockman_user_id = request.user.id
        #pr.stockman_update = datetime.datetime.now()

        #ถ้า 2 ตัวแรก เกิดจากการ CLOSE ให้ตัดออก
        if pr.note:
            first_two = pr.note[:2]
            if first_two == 'C#':
                pr.note = pr.note[2:]

        pr.is_complete = False

        pr.is_re_approve = True #เพิ่มสถานะ re_approve
        pr.save()

        items = RequisitionItem.objects.filter(requisit = pr.requisition)
        for item in items:
            item.is_used = False
            item.save()

        requisition = Requisition.objects.get(id = pr.requisition.id)
            
        baseUrgency = BaseUrgency.objects.all()
        baseUnit = BaseUnit.objects.all()
        baseProduct = Product.objects.all().values('id', 'name')
        #ถ้ามีประเภทใบขอเบิก ดึงตามประเภทใบขอเบิก
        if requisition.rq_type:
            baseCar = BaseCar.objects.filter(Q(rq_type = requisition.rq_type)| Q(rq_type = 4)).values('id', 'code', 'name')
            baseRepairType = BaseRepairType.objects.filter(Q(rq_type = requisition.rq_type)| Q(rq_type = 4)).values('id', 'name')
        else:
            baseCar = BaseCar.objects.all().values('id', 'code', 'name')
            baseRepairType = BaseRepairType.objects.all().values('id', 'name')

        #form
        form_pr = PurchaseRequisitionForm(request, instance=pr)
        if request.method == 'POST':
            form_pr = PurchaseRequisitionForm(request, request.POST, instance=pr)
            if form_pr.is_valid():
                form_pr.save()
                return HttpResponseRedirect(reverse('viewPRHistory'))

        #ถ้า user มีสิทธิ edit pr
        isEditPR = is_edit_pr(request.user)

        context = {
            'users' : items,
            'requisition' : requisition,
            'pr' : pr,
            'baseUrgency' : baseUrgency,
            'baseUnit' : baseUnit,
            'baseProduct' : baseProduct,
            'baseCar': baseCar,
            'baseRepairType': baseRepairType,
            'bc':bc,
            'form_pr':form_pr,
            'isEditPR' : isEditPR,
            'pr_page': "tab-active",
            'pr_show': "show",
            active :"active show",
            "disableTab":"disableTab",
            "colorNav":"disableNav"
        }

        return render(request, "purchaseRequisition/reApprovePR.html", context)

def closePR(request, pr_id):
    pr = PurchaseRequisition.objects.get(id = pr_id)
    requisition = Requisition.objects.get(purchase_requisition_id = pr_id)
    r_items = RequisitionItem.objects.filter(requisit = requisition.id)
    #set pr ทำรายการสำเร็จ
    pr.is_complete = True
    #ใส่ stamp ว่าเกิดจากการ CLOSE
    if pr.note:
        pr.note = "C#" + pr.note
    else:
        pr.note = "C#"

    pr.save()
    #set รายการสินค้า นำไปใช้แล้ว
    for i in r_items:
        i.is_used = True
        i.save()
    return redirect('viewPR')

def export(request):
    person_resource = DistributorResource()
    dataset = person_resource.export()
    response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="persons.xls"'
    return response

def correct_date_format(date_str):
    formats = ['%Y-%m-%d %H:%M:%S', '%d/%m/%Y']
    for date_format in formats:
        try:
            date_obj = datetime.datetime.strptime(date_str, date_format)
            return date_obj.strftime('%Y-%m-%d')
        except ValueError:
            continue
    return "Invalid date format"

def uploadReceive(request):
    active = request.session['company_code']
    if request.method == 'POST':
        uploaded_file = request.FILES.get('myfile', False)
        df = pd.read_excel(uploaded_file)
        df = df.dropna(how='all')
        filtered_data = df.values.tolist()

        if filtered_data:
            preview_data = []

            for data in filtered_data:

                if not data or len(data) <= 10 or data[1] is None or data[10] is None or pd.isna(data[1]) or pd.isna(data[10]):
                    continue
                else:
                    strRefNo = data[10]
                    refNo = strRefNo.split(' ')[0] if ' ' in strRefNo else strRefNo

                    #change data[1] to corract format
                    data_col_1 = correct_date_format(str(data[1]))

                    preview_data.append({
                        "po_ref_no": refNo,
                        "po_receive_update": data_col_1,
                    })

        # Store preview data in session temporarily
        request.session['preview_data'] = preview_data

        # Return preview data to the template for rendering
        context = {
            'preview_data': preview_data,
            'rc_page': "tab-active",
            'rc_show': "show",
            active: "active show",
            "disableTab": "disableTab",
            "colorNav": "disableNav",
        }
        return render(request, 'receive/uploadReceive.html', context)

    # If no POST data, render the upload form
    context = {
        'rc_page': "tab-active",
        'rc_show': "show",
        active: "active show",
        "disableTab": "disableTab",
        "colorNav": "disableNav",
    }
    return render(request, 'receive/uploadReceive.html', context)

def confirmUpload(request):
    if request.method == 'POST':
        preview_data = request.session.get('preview_data', [])

        for data in preview_data:
            try:
                po = PurchaseOrder.objects.get(ref_no=data['po_ref_no'])
                po.is_receive = True
                po.receive_update = data['po_receive_update']
                po.save()
                
                try:
                    rd = RateDistributor.objects.filter(po=po).order_by('-id').first()
                    #คำนวน rating (duration_rate) ระยะเวลาที่รับจริง - วันที่กำหนดรับของ
                    #ถ้าปีห่างกันมากกว่าแสดงว่าปีเป็น คศ และ พศ ให้ทำการแปลงเป็นคศ ก่อน

                    if rd:
                        if years_between(po.due_receive_update, data['po_receive_update']) > 500:
                            durationRate = calculateDurationRate(convertDateBEtoBC(str(data['po_receive_update'])), po.due_receive_update)
                        else:
                            durationRate = calculateDurationRate(data['po_receive_update'], po.due_receive_update)
                                
                        #ต้องมีคะแนนประเมินทุกรายการถึงจะให้คำนวน totalRate ได้
                        if rd.price_rate and rd.quantity_rate and durationRate and rd.service_rate and rd.safety_rate:
                            #คำนวน rating (total_rate)
                            totalRate = calculateTotalRate(rd.price_rate, rd.quantity_rate, durationRate, rd.service_rate, rd.safety_rate)
                            #คำนวน grade rate
                            grade_rate = calculateGradeRate(totalRate)

                            rd.duration_rate = durationRate
                            rd.total_rate = totalRate
                            rd.grade_id = grade_rate
                            rd.save()
                except RateDistributor.DoesNotExist:
                    pass

                try:
                    po_item = PurchaseOrderItem.objects.filter(po__ref_no = data['po_ref_no'])
                    for i in po_item: 
                        i.is_receive = True
                        i.save()
                        try:
                            item = RequisitionItem.objects.get(id = i.item_id)
                            item.is_receive = True
                            item.save()
                        except RequisitionItem.DoesNotExist:
                            pass
                except PurchaseOrderItem.DoesNotExist or PurchaseOrderItem.MultipleObjectsReturned:
                    pass

            except PurchaseOrder.DoesNotExist:
                pass

        # Clear preview data from session after saving
        request.session.pop('preview_data', None)
        return redirect('viewReceive')

    return redirect('uploadReceive')

def showReceive(request, rc_id):
    rc = Receive.objects.get(id = rc_id)
    items = ReceiveItem.objects.filter(rc = rc)

    context = {
        'rc':rc,
        'items':items,
        'rc_page': "tab-active",
        'rc_show': "show",
    }
    return render(request, 'receive/showReceive.html',context)

def viewRequisitionHistory(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    month = datetime.datetime.now().month
    year = datetime.datetime.now().year
    #data = Requisition.objects.filter(created__year__lte = year,created__month__lt = month, purchase_requisition_id__isnull = False)
    
    data = Requisition.objects.filter(is_edit = False, branch_company__code__in = company_in)
    #data = Requisition.objects.filter(Q(invoice_id__isnull = False) | Q(purchase_requisition_id__isnull = False) , branch_company__code__in = company_in)

    #กรองข้อมูล
    myFilter = RequisitionFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)

    context = {
            'requisitions':dataPage,
            'filter':myFilter,
            'h_requisitions_page': "tab-active",
            'h_requisitions_show': "show",
            active :"active show",
            "colorNav":"enableNav"
    }

    return render(request,'history/viewRequisition.html', context)

@login_required(login_url='signIn')
def viewPRHistory(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)
    
    data = PurchaseRequisition.objects.filter(Q(purchase_status_id = 2, approver_status_id = 2), branch_company__code__in = company_in)
    #requisit = PurchaseRequisition.objects.filter(Q(purchase_status_id = 2, approver_status_id = 2)).values("requisition")

    #กรองข้อมูล
    myFilter = PurchaseRequisitionFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)

    '''
    ri = RequisitionItem.objects.filter(quantity_pr__gt=0, requisit__in = requisit)

    ri_dict = {}
    for item in ri:
        requisition_id = item.requisition_id
        if requisition_id not in ri_dict:
            ri_dict[requisition_id] = []
        ri_dict[requisition_id].append(item)   
    '''

    context = {
                'prs':dataPage,
                'filter':myFilter,
                'h_pr_page': "tab-active",
                'h_pr_show': "show",
                active :"active show",
                "colorNav":"enableNav"
              }

    return render(request,'history/viewPR.html',context)


@login_required(login_url='signIn')
def viewMAHistory(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    data = Maintenance.objects.filter(Q(is_cancel = True) | Q(is_complete__isnull=False), branch_company__code__in = company_in)

    #กรองข้อมูล
    myFilter = MaintenanceFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)

    context = {
                'mas':dataPage,
                'filter':myFilter,
                'h_ma_page': "tab-active",
                'h_ma_show': "show",
                active :"active show",
                "colorNav":"enableNav"
              }

    return render(request,'history/viewMA.html',context)

def viewPOHistory(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    data = PurchaseOrder.objects.filter(Q(~Q(approver_status_id = 1), is_receive = True), is_cancel = False, branch_company__code__in = company_in)

    #กรองข้อมูล
    myFilter = PurchaseOrderFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)

    prs = PurchaseOrderItem.objects.filter(po__in=data).values('item__requisit__pr_ref_no','item__requisit__purchase_requisition_id','item__product_name','po')
    context = {
        'pos':dataPage,
        'filter':myFilter,
        'prs': prs,
        'h_po_page': "tab-active",
        'h_po_show': "show",
        active :"active show",
        "colorNav":"enableNav"
    }
    return render(request, "history/viewPO.html", context)

def viewComparePricePOHistory(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    data = ComparisonPrice.objects.filter(Q(examiner_status_id = 2, approver_status_id = 2) & ~Q(special_approver_status_id = 3) & ~Q(special_approver_status_id = 1), is_cancel = False, branch_company__code__in = company_in)

    #กรองข้อมูล
    myFilter = ComparisonPriceFilter(request.GET, queryset = data)
    data = myFilter.qs

    bidder = ComparisonPriceDistributor.objects.values('cp_id','distributor__name','distributor__id','amount').filter(cp__in=data)
    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)

    prs = ComparisonPriceItem.objects.values('item__requisit__pr_ref_no','item__requisit__purchase_requisition_id','cp').annotate(Count('item')).order_by().filter(item__count__gt=0, cp__in=data)
    context = {
        'cps':dataPage,
        'filter':myFilter,
        'bidder':bidder,
        'prs': prs,
        'h_cp_page': "tab-active",
        'h_cp_show': "show",
        active :"active show",
        "colorNav":"enableNav"
    }

    return render(request, 'history/viewComparePricePO.html',context)

#Incomplate อันนี้ไม่ได้ใช้อยู่ 19-08-2024
def viewRequisitionHistoryIncomplete(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    month = datetime.datetime.now().month
    year = datetime.datetime.now().year
    #data = Requisition.objects.filter(created__year__lte = year,created__month__lt = month, purchase_requisition_id__isnull = False)
    data = Requisition.objects.filter(purchase_requisition_id__isnull = False, branch_company__code__in = company_in)

    #กรองข้อมูล
    myFilter = RequisitionFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)

    context = {
            'requisitions':dataPage,
            'filter':myFilter,
            'h_i_requisitions_page': "tab-active",
            'h_i_requisitions_show': "show",
             active :"active show",
            "colorNav":"enableNav"
    }

    return render(request,'historyIncomplete/viewRequisition.html', context)

def viewPRHistoryIncomplete(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    data = PurchaseRequisition.objects.filter(Q(purchase_status_id = 3) | Q(approver_status_id = 3) | Q(approver_status_id = 4), branch_company__code__in = company_in)

    #กรองข้อมูล
    myFilter = PurchaseRequisitionFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)

    context = {
                'prs':dataPage,
                'filter':myFilter,
                'h_i_pr_page': "tab-active",
                'h_i_pr_show': "show",
                active :"active show",
                "colorNav":"enableNav"
              }

    return render(request,'historyIncomplete/viewPR.html',context)

def viewPOHistoryIncomplete(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    data = PurchaseOrder.objects.filter(Q(approver_status_id = 3) | Q(is_cancel = True), branch_company__code__in = company_in)

    #กรองข้อมูล
    myFilter = PurchaseOrderFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)

    context = {
        'pos':dataPage,
        'filter':myFilter,
        'h_i_po_page': "tab-active",
        'h_i_po_show': "show",
         active :"active show",
         "colorNav":"enableNav"
    }
    return render(request, "historyIncomplete/viewPO.html", context)

def viewComparePricePOHistoryIncomplete(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    data = ComparisonPrice.objects.filter(Q(examiner_status_id = 3) | Q(approver_status_id = 3) | Q(special_approver_status_id = 3) | Q(is_cancel = True), branch_company__code__in = company_in)

    #กรองข้อมูล
    myFilter = ComparisonPriceFilter(request.GET, queryset = data)
    data = myFilter.qs

    bidder = ComparisonPriceDistributor.objects.values('cp_id','distributor__name','distributor__id','amount').filter(cp__in=data)
    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)
    context = {
        'cps':dataPage,
        'filter':myFilter,
        'bidder':bidder,
        'h_i_cp_page': "tab-active",
        'h_i_cp_show': "show",
         active :"active show",
         "colorNav":"enableNav"
    }

    return render(request, 'historyIncomplete/viewComparePricePO.html',context)

def viewPOReport(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)
    data = PurchaseOrder.objects.filter(approver_status = 2, branch_company__code__in = company_in, is_cancel = False).distinct() #ใส่ distinct เพราะ filter purchaseorderitem__item__machine ทำให้เบิ้ลรายการตาม items

    #กรองข้อมูล
    myFilter = PurchaseOrderFilter(request.GET, queryset = data)
    data = myFilter.qs

    prs = PurchaseOrderItem.objects.values('item__requisit__pr_ref_no','item__requisit__purchase_requisition_id','po')

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)

    context = {
        'pos':dataPage,
        'filter':myFilter,
        'prs':prs,
        'rp_po_page': "tab-active",
        'rp_po_show': "show",
        active :"active show",
        "colorNav":"enableNav"
    }
    return render(request, "report/viewPO.html", context)

def viewPOItemReport(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)
    data = PurchaseOrderItem.objects.filter(po__approver_status = 2, po__branch_company__code__in = company_in, po__is_cancel = False).order_by('-po__created')

    #กรองข้อมูล
    myFilter = PurchaseOrderItemFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)

    context = {
        'po_item':dataPage,
        'filter':myFilter,
        'rp_poi_page': "tab-active",
        'rp_poi_show': "show",
        active :"active show",
        "colorNav":"enableNav"
    }
    return render(request, "report/viewPOItem.html", context)

def viewRateDistributorReport(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)
    #data = RateDistributor.objects.filter(po__branch_company__code__in = company_in, grade__isnull = False)
    rd = RateDistributor.objects.filter(po__branch_company__code__in = company_in).values('distributor__id')
    data = Distributor.objects.filter(id__in = rd).values('id', 'name', 'address')

    #กรองข้อมูล
    #myFilter = RateDistributorFilter(request.GET, queryset = data)
    myFilter = DistributorFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)

    context = {
        'ds':dataPage,
        'filter':myFilter,
        'rp_rd_page': "tab-active",
        'rp_rd_show': "show",
        active :"active show",
        "colorNav":"enableNav"
    }
    return render(request, "report/viewRD.html", context)

def showRateDistributor(request, pk):
    active = request.session['company_code']
    company_in = findCompanyIn(request)
    distributor = Distributor.objects.get(id = pk)
    
    #ประเมินร้านค้า
    rate_counsel = RateDistributor.objects.filter(~Q(counsel = ""), distributor__id = pk, counsel__isnull = False, po__branch_company__code__in = company_in).values('counsel','organizer_user__first_name','organizer_user__last_name','organizer_update').order_by('-id')[:5]
    num_grade_all = RateDistributor.objects.filter(distributor__id = pk, po__branch_company__code__in = company_in).count()
    ratings = RateDistributor.objects.filter(distributor__id = pk, po__branch_company__code__in = company_in).aggregate(avg_total_rate = Avg('total_rate'),  avg_price_rate = Avg('price_rate'), avg_quantity_rate = Avg('quantity_rate'), avg_duration_rate = Avg('duration_rate'), avg_service_rate = Avg('service_rate'), avg_safety_rate = Avg('safety_rate'))

    rd_data = RateDistributor.objects.filter(distributor__id = pk, po__branch_company__code__in = company_in).order_by('-po__created')
    
    paginator = Paginator(rd_data, 10)
    page_number = request.GET.get('page')
    po_rd = paginator.get_page(page_number)

    context = {
        'rp_rd_page': "tab-active",
        'rp_rd_show': "show",
        'rate_counsel':rate_counsel,
        'num_grade_all':num_grade_all,
        'ratings':ratings,
        'distributor':distributor,
        'po_rd':po_rd,
        active :"active show",
		"disableTab":"disableTab",
		"colorNav":"disableNav"
    }
    return render(request, "rateDistributor/showRateDistributor.html", context)

''' old รายงานใบสั่งซื้อเก่า
def exportExcelPO(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="PO_Report_"'+active+'".xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('รายงานใบสั่งสินค้า', cell_overwrite_ok=True) # this will make a sheet named Users Data

    # Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['เลขที่', 'วันที่', 'บริษัท', 'รหัสผู้จำหน่าย', 'ผู้จำหน่าย','วันที่กำหนดรับของ', 'รับของวันที่', 'เครดิต', 'V', 'ส่วนลด', 'มูลค่าสินค้า','VAT.', 'รวมทั้งสิ้น', 'ผู้สั่งสินค้า', 'ราคาส่วนต่างใบเปรียบเทียบ', 'ราคาที่ประหยัดได้', 'เลขที่ใบขอซื้อ', 'วันที่อนุมัติใบขอซื้อ','รายการสินค้า','ใช้ในระบบงาน','วันที่ต้องการ', 'ระดับความเร่งด่วน',  'ระยะเวลาในการซื้อ','ความล่าช้าในการรับของ']

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style) # at 0 row 0 column 

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()
    date_style = xlwt.easyxf(num_format_str='DD/MM/YYYY')
    decimal_style = xlwt.easyxf(num_format_str='#,##0.00')

    stockman_user = request.GET.get('stockman_user') or None
    ref_no = request.GET.get('ref_no') or None
    distributor = request.GET.get('distributor') or None
    amount_min = request.GET.get('amount_min') or None
    amount_max = request.GET.get('amount_max') or None
    start_created = request.GET.get('start_created') or None
    end_created = request.GET.get('end_created') or None

    my_q = Q()
    if stockman_user is not None:
        my_q = Q(stockman_user = stockman_user)
    if ref_no is not None:
        my_q &= Q(ref_no__icontains = ref_no)
    if distributor is not None:
        my_q &= Q(distributor__name__startswith = distributor)
    if start_created is not None:
        my_q &= Q(created__gte = start_created)
    if end_created is not None:
        my_q &=Q(created__lte = end_created)
    if amount_min is not None:
        my_q &= Q(amount__gte = amount_min)
    if amount_max is not None :
        my_q &=Q(amount__lte = amount_max)

    my_q &=Q(approver_status = 2, is_cancel = False)

    #ถ้ามีสิทธิดูรายงานของบริษัททั้งหมด ในแท็ป ALL จะดึงรายงานของทุกๆบริษัทมา
    if  is_view_report_all(request.user) and active == 'ALL':
        pass
    else:
        my_q &=Q(branch_company__code__in = company_in)

    rows = PurchaseOrder.objects.filter(
        my_q
    ).values_list('ref_no', 'created', 'branch_company__name', 'distributor', 'distributor__name','due_receive_update', 'receive_update', 'credit__name', 'vat_type__id','discount', 'total_after_discount','vat' ,'amount','stockman_user__first_name','cp__amount_diff').annotate(
        save_price=Case(
			When(vat_type_id = 1, then= F('total_price')- (F('total_after_discount') + F('vat'))),
			When(vat_type_id = 0, then= F('total_price')-F('total_after_discount')),
			When(vat_type_id = 2, then= F('total_price')-F('total_after_discount')),
			)).order_by('amount')

    total_price = PurchaseOrder.objects.filter(my_q).values_list('total_after_discount', flat=True)
    sum_total_price = sum(total_price)

    vat = PurchaseOrder.objects.filter(my_q).values_list('vat', flat=True)
    sum_vat = sum(vat)

    amount = PurchaseOrder.objects.filter(my_q).values_list('amount', flat=True)
    sum_amount = sum(amount)

    count = PurchaseOrder.objects.filter(my_q).count()
    
    #หารายละเอียดของสินค้าที่ออกใบสั่งซื้อ
    po = PurchaseOrder.objects.filter(my_q).order_by('amount')
    po_items = PurchaseOrderItem.objects.filter(po__in = po).values('po__ref_no', 'item__requisit__pr_ref_no', 'item__product__name', 'item__machine', 'item__desired_date', 'item__urgency', 'item__requisit').order_by('po__amount')
    
    po_items_temp = PurchaseOrderItem.objects.filter(po__in = po).values('item__requisit').order_by('po__amount')
    pr = PurchaseRequisition.objects.filter(requisition__in = po_items_temp).values('ref_no', 'approver_update')

    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            if isinstance(row[col_num], datetime.date):
                ws.write(row_num, col_num, row[col_num], date_style)
            elif col_num == 10 or col_num == 11 or col_num == 12 or col_num == 14 or col_num == 15:
                ws.write(row_num, col_num, row[col_num], decimal_style)
            else:
                ws.write(row_num, col_num, row[col_num], font_style)
            
            strPr = ""
            strPrItem = ""
            strPrMachine = ""
            strPrDesired = ""
            strPrUrgency = ""

            # loop po item
            for item in po_items:
                if row[0] == item['po__ref_no']:
                    strPr = str(item['item__requisit__pr_ref_no'])
                    strPrItem = " ".join([strPrItem, str(item['item__product__name']), ", "])
                    strPrMachine = " ".join([strPrMachine, str(item['item__machine']), ", "])
                    strPrDesired = item['item__desired_date']
                    strPrUrgency = str(item['item__urgency'])
            #remove , in last item
            strPrItem = strPrItem[:-2]
            strPrMachine = strPrMachine[:-2]

            strApproverUpdate = None
            strDateDiff = None
            for p in pr:
                if strPr == p['ref_no']:
                    strApproverUpdate = p['approver_update']

            strTempReceiveUpdate = None
            strDateDelay = None
            if row[6] and strPrDesired:
                strTempReceiveUpdate = str(row[6])
                strReceiveUpdate = convertDateBEtoBC(strTempReceiveUpdate)
                
                #DateDelay
                if row[5]:
                    if years_between(row[5], strTempReceiveUpdate) > 500:
                        strDateDelay = str(days_between_nagative(strReceiveUpdate, row[5])) + " วัน"
                    else:
                        strDateDelay = str(days_between_nagative(strTempReceiveUpdate, row[5])) + " วัน"

                #DateDiff
                if row[6]:
                    if years_between(row[6], strPrDesired) > 500:
                        strDateDiff = str(days_between_nagative(strReceiveUpdate, strPrDesired)) + " วัน"
                    else:
                        strDateDiff = str(days_between_nagative(strTempReceiveUpdate, strPrDesired)) + " วัน"

            
            ws.write(row_num, 16, strPr, font_style)
            ws.write(row_num, 17, strApproverUpdate, date_style)
            ws.write(row_num, 18, strPrItem, font_style)
            ws.write(row_num, 19, strPrMachine, font_style)
            ws.write(row_num, 20, strPrDesired, date_style)
            ws.write(row_num, 21, strPrUrgency, font_style)
            ws.write(row_num, 22, strDateDiff, font_style)
            ws.write(row_num, 23, strDateDelay, font_style)

    ws.write(row_num+1, 0, "รวมทั้งสิ้น", font_style)
    ws.write(row_num+1, 1, count, font_style)
    ws.write(row_num+1, 2, "ใบ", font_style)
    ws.write(row_num+1, 9, sum_total_price, decimal_style)
    ws.write(row_num+1, 10, sum_vat, decimal_style)
    ws.write(row_num+1, 11, sum_amount, decimal_style)

    ws.write(row_num+3, 0, "ระดับความเร่งด่วน", font_style)
    ws.write(row_num+3, 1, "1 = A", font_style)
    ws.write(row_num+3, 2, "ภายใน 48 ชม.", font_style)

    ws.write(row_num+4, 1, "2 = B", font_style)
    ws.write(row_num+4, 2, "3-5 วัน", font_style)

    ws.write(row_num+5, 1, "3 = C", font_style)
    ws.write(row_num+5, 2, "7 วัน", font_style)

    ws.write(row_num+6, 1, "4 = D", font_style)
    ws.write(row_num+6, 2, "15 วัน", font_style)

    ws.write(row_num+8, 0, "ระยะเวลาในการซื้อ คือ วันที่ต้องการเทียบกับรับของวันที่", font_style)
    ws.write(row_num+8, 1, "น้อยกว่า 0", font_style)
    ws.write(row_num+8, 2, "รับของช้ากว่าวันที่ต้องการ", font_style)

    ws.write(row_num+9, 1, "เท่ากับ 0", font_style)
    ws.write(row_num+9, 2, "รับของตรงกับวันที่ต้องการ", font_style)

    ws.write(row_num+10, 1, "มากกว่า 0", font_style)
    ws.write(row_num+10, 2, "รับของเร็วกว่าวันที่ต้องการ", font_style)

    ws.write(row_num+12, 0, "ความล่าช้าในการรับของ คือ วันที่กำหนดรับของเทียบกับรับของวันที่", font_style)
    ws.write(row_num+12, 1, "น้อยกว่า 0", font_style)
    ws.write(row_num+12, 2, "รับของช้ากว่าวันที่กำหนดรับของ", font_style)

    ws.write(row_num+13, 1, "เท่ากับ 0", font_style)
    ws.write(row_num+13, 2, "รับของตรงกับวันที่กำหนดรับของ", font_style)

    ws.write(row_num+14, 1, "มากกว่า 0", font_style)
    ws.write(row_num+14, 2, "รับของเร็วกว่าวันที่กำหนดรับของ", font_style)
    wb.save(response)

    return response
'''

def cal_days_between_nagative(day1, day2):
    strDate = None
    
    if day1 and day2:
        if years_between(day2, day1) > 500:
            strDate = str(days_between_nagative(convertDateBEtoBC(str(day1)), str(day2))) + " วัน"
        else:
            strDate = str(days_between_nagative(str(day1), str(day2))) + " วัน"
    return strDate

def normalize_datetime(value):
    if value is None:
        return None
    if is_aware(value):
        value = make_naive(value)  #Convert to naive datetime
    return value.strftime('%Y-%m-%d')  #Format as string

def exportExcelPOToExpress(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    stockman_user = request.GET.get('stockman_user') or None
    ref_no = request.GET.get('ref_no') or None
    distributor = request.GET.get('distributor') or None
    amount_min = request.GET.get('amount_min') or None
    amount_max = request.GET.get('amount_max') or None
    start_created = request.GET.get('start_created') or None
    end_created = request.GET.get('end_created') or None
    item_machine = request.GET.get('item_machine') or None
    item_rq_note = request.GET.get('item_rq_note') or None

    my_q = Q()
    if stockman_user is not None:
        my_q = Q(po__stockman_user = stockman_user)
    if ref_no is not None:
        my_q &= Q(po__ref_no__icontains = ref_no)
    if distributor is not None:
        my_q &= Q(po__distributor__name__startswith = distributor)
    if start_created is not None:
        my_q &= Q(po__created__gte = start_created)
    if end_created is not None:
        my_q &=Q(po__created__lte = end_created)
    if amount_min is not None:
        my_q &= Q(po__amount__gte = amount_min)
    if amount_max is not None :
        my_q &=Q(po__amount__lte = amount_max)
    if item_machine is not None:
        my_q &= Q(item__machine__icontains=item_machine)
    if item_rq_note is not None:
        my_q &= Q(item__requisit__note__icontains=item_rq_note)

    my_q &=Q(po__approver_status = 2, po__is_cancel = False)

    #ถ้ามีสิทธิดูรายงานของบริษัททั้งหมด ในแท็ป ALL จะดึงรายงานของทุกๆบริษัทมา
    if  is_view_report_all(request.user) and active == 'ALL':
        pass
    else:
        my_q &=Q(po__branch_company__code__in = company_in)

    queryset = PurchaseOrderItem.objects.filter(my_q).order_by('po__id')
    if not queryset.exists():
        return HttpResponse("No data to export.")

    data1 = {'เลขที่': queryset.values_list('po__ref_no', flat=True),
            'วันที่': queryset.values_list('po__created', flat=True),
            'รหัสบริษัท': queryset.values_list('po__branch_company__code', flat=True),
            'บริษัท': queryset.values_list('po__branch_company__name', flat=True),
            'รหัสผู้จำหน่าย': queryset.values_list('po__distributor', flat=True),
            'ผู้จำหน่าย': queryset.values_list('po__distributor__name', flat=True),
            'แผนกคชจ.': queryset.values_list('item__requisit__expense_dept__name', flat=True),
            'วันที่กำหนดรับของ': queryset.values_list('po__due_receive_update', flat=True),
            'รับของวันที่': queryset.values_list('po__receive_update', flat=True),
            'เครดิต': queryset.values_list('po__credit__name', flat=True),
            'V': queryset.values_list('po__vat_type__id', flat=True),
            'ส่วนลดท้ายบิล': queryset.values_list('po__discount', flat=True),
            'มูลค่าสินค้า': queryset.values_list('po__total_after_discount', flat=True),
            'VAT.': queryset.values_list('po__vat', flat=True),
            'รวมทั้งสิ้น': queryset.values_list('po__amount', flat=True),
            'ผู้สั่งสินค้า': queryset.values_list('po__stockman_user__first_name', flat=True),
            'ราคาส่วนต่างใบเปรียบเทียบ': queryset.values_list('po__cp__amount_diff', flat=True),
            'ราคาที่ประหยัดได้': queryset.annotate(
                    save_price=Case(
                        When(po__vat_type_id = 1, then= F('po__total_price')- (F('po__total_after_discount') + F('po__vat'))),
                        When(po__vat_type_id = 0, then= F('po__total_price')-F('po__total_after_discount')),
                        When(po__vat_type_id = 2, then= F('po__total_price')-F('po__total_after_discount')),
                        )).values_list('save_price', flat=True),        
            'เลขที่ใบขอซื้อ': queryset.values_list('item__requisit__pr_ref_no', flat=True),
            'วันที่อนุมัติใบขอซื้อ' :[
                normalize_datetime(
                    PurchaseRequisition.objects.filter(requisition = rq).values_list('approver_update', flat=True).first()
                )
                for rq in queryset.values_list('item__requisit', flat=True)
            ],
            'รหัสสินค้า' : queryset.values_list('item__product__id', flat=True),
            'รายการสินค้า' : queryset.values_list('item__product__name', flat=True),
            'กลุ่มสินค้า' : queryset.values_list('item__product__category__name', flat=True),
            'จำนวน' : queryset.values_list('quantity', flat=True),
            'ราคาต่อหน่วย' : queryset.values_list('unit_price', flat=True),
            'ส่วนลด' : queryset.values_list('discount', flat=True),
            'ราคารวม' : queryset.values_list('price', flat=True),
            'ใช้ในระบบงาน' : queryset.annotate(
                car_name_code=Concat('item__requisit__car__name', 'item__requisit__car__code'),
                machine=Case(
                    When(~Q(car_name_code =''), then='car_name_code'),
                    default='item__machine'
                )).values_list('machine', flat=True), #ใช้ในระบบงาน version ใหม่ดึงจาก car.name + car.code อันเก่าดึงจาก machine 10/10/2024
            'วันที่ต้องการ' : queryset.values_list('item__desired_date', flat=True),
            'ระดับความเร่งด่วน' : queryset.values_list('item__urgency', flat=True),
            'ระยะเวลาในการซื้อ' :  [ cal_days_between_nagative(qs[1], qs[0]) for qs in queryset.values_list('item__desired_date', 'po__receive_update')],
            'ความล่าช้าในการรับของ' :  [ cal_days_between_nagative(qs['po__receive_update'], qs['po__due_receive_update']) for qs in queryset.values('po__due_receive_update', 'po__receive_update')],
            }

    df1 = pd.DataFrame(data1)

    data2 = {'เลขที่': 'รวมทั้งสิ้น',
            'วันที่': [queryset.values('po__id').distinct().count()],
            'บริษัท': 'ใบ',
            'รหัสผู้จำหน่าย': '',
            'ผู้จำหน่าย': '',
            'วันที่กำหนดรับของ': '',
            'รับของวันที่': '',
            'เครดิต': '',
            'V': '',
            'ส่วนลดท้ายบิล': '',
            'มูลค่าสินค้า': [queryset.values('po__id').distinct().aggregate(s_t_f_d = Sum('po__total_after_discount'))['s_t_f_d']],
            'VAT.': [queryset.values('po__id').distinct().aggregate(s_vat = Sum('po__vat'))['s_vat']],
            'รวมทั้งสิ้น': [queryset.values('po__id').distinct().aggregate(s_amount = Sum('po__amount'))['s_amount']],
            }

    df2 = pd.DataFrame(data2, index=[0])

    data3 = {"เลขที่": ['ระดับความเร่งด่วน', '', '', ''], "วันที่": ['1 = A', '2 = B', '3 = C', '4 = D'], "บริษัท": ['ภายใน 48 ชม.', '3-5 วัน', '7 วัน', '15 วัน']}
    df3 = pd.DataFrame(data3, index=[0, 1, 2, 3])

    data4 = {"เลขที่": ['ระยะเวลาในการซื้อ คือ วันที่ต้องการเทียบกับรับของวันที่', '', ''], "วันที่": ['น้อยกว่า 0', 'เท่ากับ 0', 'มากกว่า 0'], "บริษัท": ['รับของช้ากว่าวันที่ต้องการ', 'รับของตรงกับวันที่ต้องการ', 'รับของเร็วกว่าวันที่ต้องการ']}
    df4 = pd.DataFrame(data4, index=[0, 1, 2])

    data5 = {"เลขที่": ['ความล่าช้าในการรับของ คือ วันที่กำหนดรับของเทียบกับรับของวันที่', '', ''], "วันที่": ['น้อยกว่า 0', 'เท่ากับ 0', 'มากกว่า 0'], "บริษัท": ['รับของช้ากว่าวันที่กำหนดรับของ', 'รับของตรงกับวันที่กำหนดรับของ', 'รับของเร็วกว่าวันที่ต้องการ']}
    df5 = pd.DataFrame(data5, index=[0, 1, 2])

    frames = [df1, df2, df3, df4, df5]
    result = pd.concat(frames)

    output = BytesIO()
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=PO_to_Express_Report_({active}).xlsx'

    with pd.ExcelWriter(response, engine='xlsxwriter', options={'strings_to_numbers': True}) as writer:
        result.to_excel(writer, index=False)

    return response

def exportExcelPO(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    stockman_user = request.GET.get('stockman_user') or None
    ref_no = request.GET.get('ref_no') or None
    distributor = request.GET.get('distributor') or None
    amount_min = request.GET.get('amount_min') or None
    amount_max = request.GET.get('amount_max') or None
    start_created = request.GET.get('start_created') or None
    end_created = request.GET.get('end_created') or None
    item_machine = request.GET.get('item_machine') or None
    item_rq_note = request.GET.get('item_rq_note') or None

    my_q = Q()
    if stockman_user is not None:
        my_q = Q(stockman_user = stockman_user)
    if ref_no is not None:
        my_q &= Q(ref_no__icontains = ref_no)
    if distributor is not None:
        my_q &= Q(distributor__name__startswith = distributor)
    if start_created is not None:
        my_q &= Q(created__gte = start_created)
    if end_created is not None:
        my_q &=Q(created__lte = end_created)
    if amount_min is not None:
        my_q &= Q(amount__gte = amount_min)
    if amount_max is not None :
        my_q &=Q(amount__lte = amount_max)
    if item_machine is not None:
        my_q &= Q(purchaseorderitem__item__machine__icontains=item_machine)
    if item_rq_note is not None:
        my_q &= Q(purchaseorderitem__item__requisit__note__icontains=item_rq_note)
    
    my_q &=Q(approver_status = 2, is_cancel = False)

    #ถ้ามีสิทธิดูรายงานของบริษัททั้งหมด ในแท็ป ALL จะดึงรายงานของทุกๆบริษัทมา
    if  is_view_report_all(request.user) and active == 'ALL':
        pass
    else:
        my_q &=Q(branch_company__code__in = company_in)

    queryset = PurchaseOrder.objects.filter(my_q).distinct().order_by('amount', 'id')#ใส่ distinct เพราะ filter purchaseorderitem__item__machine ทำให้เบิ้ลรายการตาม items
    if not queryset.exists():
        return HttpResponse("No data to export.")

    data1 = {'เลขที่': queryset.values_list('ref_no', flat=True),
            'วันที่': queryset.values_list('created', flat=True),
            'รหัสบริษัท': queryset.values_list('branch_company__code', flat=True),
            'บริษัท': queryset.values_list('branch_company__name', flat=True),
            'รหัสผู้จำหน่าย': queryset.values_list('distributor', flat=True),
            'ผู้จำหน่าย': queryset.values_list('distributor__name', flat=True),
            'วันที่กำหนดรับของ': queryset.values_list('due_receive_update', flat=True),
            'รับของวันที่': queryset.values_list('receive_update', flat=True),
            'เครดิต': queryset.values_list('credit__name', flat=True),
            'V': queryset.values_list('vat_type__id', flat=True),
            'ส่วนลดท้ายบิล': queryset.values_list('discount', flat=True),
            'มูลค่าสินค้า': queryset.values_list('total_after_discount', flat=True),
            'VAT.': queryset.values_list('vat', flat=True),
            'รวมทั้งสิ้น': queryset.values_list('amount', flat=True),
            'ผู้สั่งสินค้า': queryset.values_list('stockman_user__first_name', flat=True),
            'ราคาส่วนต่างใบเปรียบเทียบ': queryset.values_list('cp__amount_diff', flat=True),
            'ราคาที่ประหยัดได้': queryset.annotate(
                    save_price=Case(
                        When(vat_type_id = 1, then= F('total_price')- (F('total_after_discount') + F('vat'))),
                        When(vat_type_id = 0, then= F('total_price')-F('total_after_discount')),
                        When(vat_type_id = 2, then= F('total_price')-F('total_after_discount')),
                        )).values_list('save_price', flat=True),
            'เลขที่ใบขอซื้อ': [ PurchaseOrderItem.objects.filter(po = id).values_list('item__requisit__pr_ref_no', flat=True).first() for id in queryset.values_list('id', flat=True)],
            'วันที่อนุมัติใบขอซื้อ' : [
                normalize_datetime(
                    PurchaseRequisition.objects.filter(
                        requisition=PurchaseOrderItem.objects.filter(po=id)
                        .values_list('item__requisit', flat=True)
                        .first()
                    ).values_list('approver_update', flat=True).first()
                )
                for id in queryset.values_list('id', flat=True)
            ],
            'รหัสสินค้า' : [', '.join(map(str, list(PurchaseOrderItem.objects.filter(po=id).values_list('item__product__id', flat=True).distinct()))) for id in queryset.values_list('id', flat=True)],
            'รายการสินค้า' : [', '.join(map(str, list(PurchaseOrderItem.objects.filter(po=id).values_list('item__product__name', flat=True).distinct()))) for id in queryset.values_list('id', flat=True)],
            'ใช้ในระบบงาน' : [', '.join(map(str, list(PurchaseOrderItem.objects.filter(po=id)
                                                .annotate(car_name_code=Concat('item__requisit__car__name', 'item__requisit__car__code'), machine=Case(When(~Q(car_name_code =''), then='car_name_code'),default='item__machine'))
                                                .values_list('machine', flat=True).distinct()))) 
                                                for id in queryset.values_list('id', flat=True)], #ใช้ในระบบงาน version ใหม่ดึงจาก car.name + car.code อันเก่าดึงจาก machine 10/10/2024
            'วันที่ต้องการ' :   [', '.join(map(str, list(PurchaseOrderItem.objects.filter(po=id).values_list('item__desired_date', flat=True).distinct()))) for id in queryset.values_list('id', flat=True)],
            'ระดับความเร่งด่วน' :  [', '.join(map(str, list(PurchaseOrderItem.objects.filter(po=id).values_list('item__urgency', flat=True).distinct()))) for id in queryset.values_list('id', flat=True)],
            'ระยะเวลาในการซื้อ' : [ cal_days_between_nagative(qs[1], PurchaseOrderItem.objects.filter(po = qs[0]).values_list('item__desired_date', flat=True).first()) for qs in queryset.values_list('id', 'receive_update')],
            'ความล่าช้าในการรับของ' : [ cal_days_between_nagative(qs['receive_update'], qs['due_receive_update']) for qs in queryset.values('due_receive_update', 'receive_update')],
            }

    df1 = pd.DataFrame(data1)

    data2 = {'เลขที่': 'รวมทั้งสิ้น',
            'วันที่': [queryset.aggregate(c_id = Count('id'))['c_id']],
            'บริษัท': 'ใบ',
            'รหัสผู้จำหน่าย': '',
            'ผู้จำหน่าย': '',
            'วันที่กำหนดรับของ': '',
            'รับของวันที่': '',
            'เครดิต': '',
            'V': '',
            'ส่วนลดท้ายบิล': '',
            'มูลค่าสินค้า': [queryset.aggregate(s_t_f_d = Sum('total_after_discount'))['s_t_f_d']],
            'VAT.': [queryset.aggregate(s_vat = Sum('vat'))['s_vat']],
            'รวมทั้งสิ้น': [queryset.aggregate(s_amount = Sum('amount'))['s_amount']],
            }

    df2 = pd.DataFrame(data2, index=[0])

    data3 = {"เลขที่": ['ระดับความเร่งด่วน', '', '', ''], "วันที่": ['1 = A', '2 = B', '3 = C', '4 = D'], "บริษัท": ['ภายใน 48 ชม.', '3-5 วัน', '7 วัน', '15 วัน']}
    df3 = pd.DataFrame(data3, index=[0, 1, 2, 3])

    data4 = {"เลขที่": ['ระยะเวลาในการซื้อ คือ วันที่ต้องการเทียบกับรับของวันที่', '', ''], "วันที่": ['น้อยกว่า 0', 'เท่ากับ 0', 'มากกว่า 0'], "บริษัท": ['รับของช้ากว่าวันที่ต้องการ', 'รับของตรงกับวันที่ต้องการ', 'รับของเร็วกว่าวันที่ต้องการ']}
    df4 = pd.DataFrame(data4, index=[0, 1, 2])

    data5 = {"เลขที่": ['ความล่าช้าในการรับของ คือ วันที่กำหนดรับของเทียบกับรับของวันที่', '', ''], "วันที่": ['น้อยกว่า 0', 'เท่ากับ 0', 'มากกว่า 0'], "บริษัท": ['รับของช้ากว่าวันที่กำหนดรับของ', 'รับของตรงกับวันที่กำหนดรับของ', 'รับของเร็วกว่าวันที่ต้องการ']}
    df5 = pd.DataFrame(data5, index=[0, 1, 2])

    frames = [df1, df2, df3, df4, df5]
    result = pd.concat(frames)

    output = BytesIO()
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=PO_Report_({active}).xlsx'

    with pd.ExcelWriter(response, engine='xlsxwriter', options={'strings_to_numbers': True}) as writer:
        result.to_excel(writer, index=False)

    return response

def exportExcelSummaryByProductValue(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    output = BytesIO()
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/ms-excel'
    )
    response['Content-Disposition'] = f'attachment; filename=Report_Product_By_Value({active}).xls'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('รายงานสรุปตามมูลค่าสินค้า', cell_overwrite_ok=True) # this will make a sheet named Users Data

    # Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['รหัสสินค้า', 'รายละเอียด', 'หน่วยนับ', 'ราคาต่อหน่วย','ราคาเฉลี่ยต่อหน่วย', 'จำนวนครั่งที่สั่งซื้อ', 'ปริมาณซื้อสุทธิ', 'รวมมูลค่าซื้อ', 'ความเร่งด่วนเฉลี่ย', 'ใช้ในระบบงาน']

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style) # at 0 row 0 column 

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()
    decimal_style = xlwt.easyxf(num_format_str='#,##0.00')

    item_product_id_from = request.GET.get('item_product_id_from') or None
    item_product_id_to = request.GET.get('item_product_id_to') or None
    item_product_name = request.GET.get('item_product_name') or None
    item_machine = request.GET.get('item_machine') or None
    item_rq_note = request.GET.get('item_rq_note') or None
    distri_id = request.GET.get('distri_id') or None
    distributor = request.GET.get('distributor') or None
    stockman_user = request.GET.get('stockman_user') or None
    start_created = request.GET.get('start_created') or None
    end_created = request.GET.get('end_created') or None
    unit_price_min = request.GET.get('unit_price_min') or None
    unit_price_max = request.GET.get('unit_price_max') or None
    category = request.GET.get('category') or None

    my_q = Q()
    if item_product_id_from is not None:
        my_q = Q(item__product_id__gte = item_product_id_from)
    if item_product_id_to is not None:
        my_q &= Q(item__product_id__lte = item_product_id_to)
    if item_product_name is not None:
        my_q &= Q(item__product_name__icontains = item_product_name)
    if stockman_user is not None:
        my_q &= Q(po__stockman_user = stockman_user)
    if category is not None:
        my_q &= Q(item__product__category = category)
    if distri_id is not None:
        my_q &= Q(po__distributor__id__startswith = distri_id)
    if distributor is not None:
        my_q &= Q(po__distributor__name__startswith = distributor)
    if start_created is not None:
        my_q &= Q(po__created__gte = start_created)
    if end_created is not None:
        my_q &=Q(po__created__lte = end_created)
    if unit_price_min is not None:
        my_q &= Q(unit_price__gte = unit_price_min)
    if unit_price_max is not None :
        my_q &=Q(unit_price__lte = unit_price_max)
    if item_machine is not None :
        my_q &=Q(item__machine__icontains = item_machine)
    if item_rq_note is not None :
        my_q &=Q(item__requisit__note__icontains = item_rq_note)

    my_q &=Q(po__approver_status = 2, po__is_cancel = False)

    #ถ้ามีสิทธิดูรายงานทั้งหมด ในแท็ป ALL จะดึงรายงานของทุกๆบริษัทมา
    if  is_view_report_all(request.user) and active == 'ALL':
        pass
    else:
        my_q &=Q(po__branch_company__code__in = company_in)

    rows = PurchaseOrderItem.objects.filter(
        my_q
    ).values_list('item__product_id', 'item__product__name', 'item__product__unit__name','item__product_id').annotate(avg = Avg('unit_price'), count = Count('item__product_id'), quantity = Sum('quantity'), total_price = Sum('price'), ugency = Round(Avg('item__urgency'))).order_by('-total_price')

    po_items = PurchaseOrderItem.objects.filter(my_q).annotate(
                car_name_code=Concat('item__requisit__car__name', 'item__requisit__car__code'),
                mc=Case(
                    When(~Q(car_name_code =''), then='car_name_code'),
                    default='item__machine'
                )).values('item__product_id','unit_price','mc') #ใช้ในระบบงาน version ใหม่ดึงจาก car.name + car.code อันเก่าดึงจาก machine 10/10/2024

    total_price = PurchaseOrderItem.objects.filter(my_q).aggregate(Sum('price'))
    sum_total_price = total_price['price__sum']
    
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            if col_num > 3:
                ws.write(row_num, col_num, row[col_num], decimal_style)
            else:
                ws.write(row_num, col_num, row[col_num], font_style)


        strUnitPrice = ""
        strMachine = ""

        # loop po item
        for item in po_items:
            if row[0] == item['item__product_id']:
                strUnitPrice = " ".join([strUnitPrice, str(item['unit_price']), ", "])
                strMachine = " ".join([strMachine, str(item['mc']), ", "])
        #remove , in last item
        strUnitPrice = strUnitPrice[:-2]
        strMachine = strMachine[:-2]
        ws.write(row_num, 3, strUnitPrice, decimal_style)
        ws.write(row_num, 9, strMachine, font_style)

    ws.write(row_num+1, 0, "รวมทั้งสิ้น", font_style)
    ws.write(row_num+1, 7, sum_total_price, decimal_style)

    ws.write(row_num+3, 0, "ระดับความเร่งด่วน", font_style)
    ws.write(row_num+3, 1, "1.00 = A", font_style)
    ws.write(row_num+3, 2, "ภายใน 48 ชม.", font_style)

    ws.write(row_num+4, 1, "2.00 = B", font_style)
    ws.write(row_num+4, 2, "3-5 วัน", font_style)

    ws.write(row_num+5, 1, "3.00 = C", font_style)
    ws.write(row_num+5, 2, "7 วัน", font_style)

    ws.write(row_num+6, 1, "4.00 = D", font_style)
    ws.write(row_num+6, 2, "15 วัน", font_style)

    wb.save(response)

    return response

def exportExcelSummaryByProductFrequently(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    output = BytesIO()
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/ms-excel'
    )
    response['Content-Disposition'] = f'attachment; filename=Report_Product_By_Frequently({active}).xls'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('รายงานสรุปตามจำนวนครั้งที่ซื้อ', cell_overwrite_ok=True) # this will make a sheet named Users Data

    # Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['รหัสสินค้า', 'รายละเอียด', 'หน่วยนับ', 'ราคาต่อหน่วย','ราคาเฉลี่ยต่อหน่วย', 'จำนวนครั่งที่สั่งซื้อ', 'ปริมาณซื้อสุทธิ', 'รวมมูลค่าซื้อ', 'ความเร่งด่วนเฉลี่ย', 'ใช้ในระบบงาน']

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style) # at 0 row 0 column 

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()
    decimal_style = xlwt.easyxf(num_format_str='#,##0.00')

    item_product_id_from = request.GET.get('item_product_id_from') or None
    item_product_id_to = request.GET.get('item_product_id_to') or None
    item_product_name = request.GET.get('item_product_name') or None
    item_machine = request.GET.get('item_machine') or None
    item_rq_note = request.GET.get('item_rq_note') or None
    distri_id = request.GET.get('distri_id') or None
    distributor = request.GET.get('distributor') or None
    stockman_user = request.GET.get('stockman_user') or None
    start_created = request.GET.get('start_created') or None
    end_created = request.GET.get('end_created') or None
    unit_price_min = request.GET.get('unit_price_min') or None
    unit_price_max = request.GET.get('unit_price_max') or None
    category = request.GET.get('category') or None

    my_q = Q()
    if item_product_id_from is not None:
        my_q = Q(item__product_id__gte = item_product_id_from)
    if item_product_id_to is not None:
        my_q &= Q(item__product_id__lte = item_product_id_to)
    if item_product_name is not None:
        my_q &= Q(item__product_name__icontains = item_product_name)
    if stockman_user is not None:
        my_q &= Q(po__stockman_user = stockman_user)
    if category is not None:
        my_q &= Q(item__product__category = category)
    if distri_id is not None:
        my_q &= Q(po__distributor__id__startswith = distri_id)
    if distributor is not None:
        my_q &= Q(po__distributor__name__startswith = distributor)
    if start_created is not None:
        my_q &= Q(po__created__gte = start_created)
    if end_created is not None:
        my_q &=Q(po__created__lte = end_created)
    if unit_price_min is not None:
        my_q &= Q(unit_price__gte = unit_price_min)
    if unit_price_max is not None :
        my_q &=Q(unit_price__lte = unit_price_max)
    if item_machine is not None :
        my_q &=Q(item__machine__icontains = item_machine)
    if item_rq_note is not None:
        my_q &= Q(item__requisit__note__icontains=item_rq_note)

    my_q &=Q(po__approver_status = 2, po__is_cancel = False)

    #ถ้ามีสิทธิดูรายงานของบริษัททั้งหมด ในแท็ป ALL จะดึงรายงานของทุกๆบริษัทมา
    if  is_view_report_all(request.user) and active == 'ALL':
        pass
    else:
        my_q &=Q(po__branch_company__code__in = company_in)

    rows = PurchaseOrderItem.objects.filter(
        my_q
    ).values_list('item__product_id', 'item__product__name', 'item__product__unit__name','item__product_id').annotate(avg = Avg('unit_price'), count = Count('item__product_id'), quantity = Sum('quantity'), total_price = Sum('price'), ugency = Round(Avg('item__urgency'))).order_by('-count')

    po_items = PurchaseOrderItem.objects.filter(my_q).annotate(
                car_name_code=Concat('item__requisit__car__name', 'item__requisit__car__code'),
                mc=Case(
                    When(~Q(car_name_code =''), then='car_name_code'),
                    default='item__machine'
                )).values('item__product_id','unit_price','mc') #ใช้ในระบบงาน version ใหม่ดึงจาก car.name + car.code อันเก่าดึงจาก machine 10/10/2024

    total_price = PurchaseOrderItem.objects.filter(my_q).aggregate(Sum('price'))
    sum_total_price = total_price['price__sum']
    
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            if col_num > 3:
                ws.write(row_num, col_num, row[col_num], decimal_style)
            else:
                ws.write(row_num, col_num, row[col_num], font_style)

        strUnitPrice = ""
        strMachine = ""

        # loop po item
        for item in po_items:
            if row[0] == item['item__product_id']:
                strUnitPrice = " ".join([strUnitPrice, str(item['unit_price']), ", "])
                strMachine = " ".join([strMachine, str(item['mc']), ", "])
        #remove , in last item
        strUnitPrice = strUnitPrice[:-2]
        strMachine = strMachine[:-2]
        
        ws.write(row_num, 3, strUnitPrice, decimal_style)
        ws.write(row_num, 9, strMachine, font_style)

    ws.write(row_num+1, 0, "รวมทั้งสิ้น", font_style)
    ws.write(row_num+1, 7, sum_total_price, decimal_style)

    ws.write(row_num+3, 0, "ระดับความเร่งด่วน", font_style)
    ws.write(row_num+3, 1, "1.00 = A", font_style)
    ws.write(row_num+3, 2, "ภายใน 48 ชม.", font_style)

    ws.write(row_num+4, 1, "2.00 = B", font_style)
    ws.write(row_num+4, 2, "3-5 วัน", font_style)

    ws.write(row_num+5, 1, "3.00 = C", font_style)
    ws.write(row_num+5, 2, "7 วัน", font_style)

    ws.write(row_num+6, 1, "4.00 = D", font_style)
    ws.write(row_num+6, 2, "15 วัน", font_style)

    wb.save(response)

    return response

#ค้นหารายการสินค้าที่สั่งซื้อ 5 รายการล่าสุด
def searchLastPoItem(request):
    strName = "<table class='table table-striped'><thead class = 'thead-dark'><tr><th>เลขที่</th><th>วันที่</th><th>ราคาต่อหน่วย</th><th>หน่วย</th><th>เกรด/ยี่ห้อ</th><th>จากร้าน</th><th>ซื้อโดย</th><th>เจ้าหน้าที่จัดซื้อ</th></thead></tr>"
    item = request.GET.getlist('itemList[]', None)
    
    product = RequisitionItem.objects.filter(id__in = item).values('product__id')
    count = 0

    for pd in product:
        try:
            po_item = PurchaseOrderItem.objects.filter(item__product__id = pd['product__id'], unit_price__isnull = False , po__approver_status = 2).order_by('-po__created')[:5]

            strName = ''.join([strName, "<tr class='bg-info text-white'><td colspan='8'><b>" + po_item[0].item.product.id + " : "+ po_item[0].item.product.name +"</b></td></tr>"])
            index = 1
            if po_item:
                for i in po_item:

                    if i.po.cp:
                        cp_item = ComparisonPriceItem.objects.filter(item = i.item, cp = i.po.cp.id , bidder__is_select = True)
                        brand = cp_item[0].brand
                    else:
                        brand = ''   

                    ''' กดเข้าไปดูแต่ละ po ได้
                    show_po_url = reverse('showPO', args=[i.po.id, 4])

                    href_string = f'<a href="{show_po_url}">{i.po.created.strftime("%d/%m/%Y")}</a>'

                    strName = ''.join([strName, "<tr>"])
                    strName = ''.join([
                        strName,
                        "<td>" + str(index) + ")</td>",
                        "<td>" + href_string + "</td>",
                        "<td>" + str(i.item.product_name) + "</td>",
                        "<td>" + f'{i.unit_price:,}' + "</td>",
                        "<td>" + str(i.unit.name) + "</td>",
                        "<td>" + str(brand) + "</td>",
                        "<td>" + str(i.po.distributor.name) + "</td>",
                        "<td>" + str(i.po.branch_company.name) + "</b></td>",
                        "<td>" + str(i.po.stockman_user) + "</b></td>"
                    ])
                    '''

                    strName = ''.join([strName, "<tr>"])
                    strName = ''.join([strName, "<td>" + str(index) + ")</td><td><b>"+ i.po.created.strftime("%d/%m/%Y") + "</td><td>"+ f'{i.unit_price:,}' + "</td><td>" + str(i.unit.name) + "</td><td>" + str(brand) + "</td><td>" + str(i.po.distributor.name) + "</td><td>" + str(i.po.branch_company.name) + "</b></td>"+ "</td><td>" + str(i.po.stockman_user) + "</b></td>"])
                    strName = ''.join([strName, "</tr>"])
                    index += 1
            count += 1
        except IndexError or PurchaseOrderItem.DoesNotExist:
            strName = ''.join([strName, "<tr><td class='alert alert-warning' colspan='8'><b>ไม่มีรายการ "+pd['product__id']+" ที่ซื้อล่าสุด</b></td></tr>"])

    strName = ''.join([strName, "</table>"])

    data = {
        'instance': strName,
    }
    return JsonResponse(data)

def get_grade(average_total_rate):
    if 90 <= average_total_rate <= 100:
        return 'A'
    elif 80 <= average_total_rate < 90:
        return 'B'
    elif 70 <= average_total_rate < 80:
        return 'C'
    elif 60 <= average_total_rate < 70:
        return 'D'
    elif 60 < average_total_rate :
        return 'F'
    else:
        return '-'

def exportToExcelRateDistributor(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    company_name = BranchCompanyBaseAdress.objects.filter(branch_company__code = active).values_list('address__name_th', flat=True)[:1] or None

    output = BytesIO()
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/ms-excel'
    )
    response['Content-Disposition'] = f'attachment; filename=rate_distributor_report_({active}).xlsx'

    wb = Workbook()
    ws = wb.active

    distributor_id_from = request.GET.get('distributor_id_from') or None
    distributor_id_to = request.GET.get('distributor_id_to') or None
    distributor = request.GET.get('distributor') or None

    my_q = Q()
    if distributor_id_from is not None:
        my_q &= Q(distributor__id__gte = distributor_id_from)
    if distributor_id_to is not None:
        my_q &= Q(distributor__id__lte = distributor_id_to)
    if distributor is not None :
        my_q &=Q(distributor__name__icontains = distributor)

    rd = RateDistributor.objects.filter(my_q, po__branch_company__code__in = company_in, grade__isnull = False).values('distributor__id')
    distributors = Distributor.objects.filter(id__in = rd)

    title = [company_name[0] if company_name else 'บริษัททั้งหมด', '', '', 'บัญชีรายชื่อผู้ขายที่ผ่านการอนุมัติแล้ว Approved Provider List : APL']
    ws.append(title)
    ws.row_dimensions[1].height = 40
    
    header = ['ลำดับ', 'รหัส Provide','รายชื่อผู้ขาย/ผู้รับจ้าง (Provide)', 'ที่อยู่','เบอร์โทร']
    six_month_periods = get_six_month_periods()

    for period_start, period_end in six_month_periods:
        header.append(f'ครั้งที่ { "1" if period_start.strftime("%m") == "01" else "2" } ปี {period_start.strftime("%Y")}')
    
    ws.append(header)
    ws.row_dimensions[2].height = 40

    column_num = len(six_month_periods)

    for index, distributor_data in enumerate(distributors):
        distributor_id = distributor_data.id
        distributor = Distributor.objects.get(pk=distributor_id)
        row = [index+1]
        row.append(distributor.id)
        row.append(distributor.name)
        row.append(distributor.address)
        row.append(distributor.tel)

        for period_start, period_end in six_month_periods:
            avg_total_rate = RateDistributor.objects.filter(
                distributor=distributor,
                organizer_update__range=[period_start, period_end],
                po__branch_company__code__in = company_in,
            ).aggregate(Avg('total_rate'))['total_rate__avg'] or 0
            grade = get_grade(avg_total_rate)
            row.append(grade)

        ws.append(row)
        ws.row_dimensions[index+3].height = 30

    try:
        # Set the column widths
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[2].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) 
            ws.column_dimensions[column].width = adjusted_width
    
        # Center-align cell content and get cell index
        for row_index, row in enumerate(ws.iter_rows(), 1):
            for cell_index, cell in enumerate(row, 1):
                if row_index == 1:
                    cell.alignment = Alignment(horizontal='center')

                if cell_index > 5:
                    cell.alignment = Alignment(horizontal='center')

        side = Side(border_style='thin', color='000000')
        set_border(ws, side)

        ws.merge_cells(start_row=1, start_column = 1, end_row=1, end_column=3)
        ws.merge_cells(start_row=1, start_column = 4, end_row=1, end_column= column_num + 5)
    except:
        pass 

    wb.save(response)
    return response

def get_six_month_periods():
    start_date = date(2023, 1, 1)
    end_date = datetime.date.today()

    periods = []
    while start_date < end_date:
        period_end = getLastDayInSixMonth(start_date)
        periods.append((start_date, period_end))
        start_date = period_end + timedelta(days=1)

    return periods

def getLastDayInSixMonth(start_date):
    try:
        period_end = start_date.replace(month=start_date.month + 6) - timedelta(days=1)
    except ValueError:
        period_end = datetime.date(start_date.year, 12, 31)
    
    return  period_end

def setSessionCompany(request):
    name = request.GET.get('title', None)
    request.session['company_code'] = name
    company = BaseBranchCompany.objects.get(code = request.session['company_code'])
    request.session['company'] = company.affiliated.name_sh
    data = {
        'instance': request.session['company_code'],
    }
    return JsonResponse(data)

def viewInvoice(request):
    active = request.session['company_code']
    data = Invoice.objects.filter(branch_company__code = active, is_express = False)

    #กรองข้อมูล
    myFilter = InvoidFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)
    
    context = {
        'ivs': dataPage,
        'isDelIv' : is_del_iv(request.user),
        'isSupplies' : is_supplies(request.user),
        'filter':myFilter,
        'iv_page': "tab-active",
        'iv_show': "show",
        active :"active show",
        "colorNav":"enableNav"
    }
    return render(request, "invoice/viewInvoice.html", context)

def showInvoice(request, iv_id, mode):
    active = request.session['company_code']

    items = InvoiceItem.objects.filter(iv = iv_id)
    iv = Invoice.objects.get(id = iv_id)

    page = IVPageMode(mode)
    show = IVShowMode(mode)

    context = {
        'items': items,
        'iv': iv,
        page:"tab-active",
        show:"show",
        active :"active show",
		"disableTab":"disableTab",
		"colorNav":"disableNav"
    }

    return render(request, "invoice/showInvoice.html", context)

def removeInvoice(request, iv_id):
    iv = Invoice.objects.get(id = iv_id)

    #set invoice_id เป็น None ด้วย
    rq = Requisition.objects.get(invoice_id = iv_id)
    rq.invoice_id = None
    rq.iv_ref_no = None
    rq.is_edit = True
    rq.save()

    #ลบ item ใน PurchaseOrder ด้วย
    items = InvoiceItem.objects.filter(iv = iv)
    items.delete()
    #ลบ PurchaseOrder ทีหลัง
    iv.delete()
    return redirect('viewInvoice')

def exportExcelRQ(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    ref_no = request.GET.get('ref_no') or None
    name = request.GET.get('name') or None
    section = request.GET.get('section') or None
    organizer = request.GET.get('organizer') or None
    start_created = request.GET.get('start_created') or None
    end_created = request.GET.get('end_created') or None

    my_q = Q()
    if ref_no is not None:
        my_q &= Q(requisit__ref_no__icontains = ref_no)
    if name is not None:
        my_q &= Q(requisit__name__first_name__icontains = name)
    if section is not None:
        my_q &= Q(requisit__section = section)
    if organizer is not None :
        my_q &=Q(requisit__organizer = organizer)
    if start_created is not None:
        my_q &= Q(requisit__created__gte = start_created)
    if end_created is not None:
        my_q &=Q(requisit__created__lte = end_created)

    my_q &=Q(requisit__purchase_requisition_id__isnull = False)

    #ถ้ามีสิทธิดูรายงานของบริษัททั้งหมด ในแท็ป ALL จะดึงรายงานของทุกๆบริษัทมา
    if  is_view_report_all(request.user) and active == 'ALL':
        pass
    else:
        my_q &=Q( requisit__branch_company__code__in = company_in )

    queryset = RequisitionItem.objects.filter(my_q).order_by('requisit__id')
    if not queryset.exists():
        return HttpResponse("No data to export.")

    data1 = {
            'วันที่': queryset.values_list('requisit__created', flat=True),
            'บริษัท': queryset.values_list('requisit__branch_company__name', flat=True),            
            'ทะเบียนรถ/เครื่องจักร/หน่วยงาน': queryset.annotate(
                                            car_name_code=Concat('requisit__car__name', 'requisit__car__code'),
                                            mc =Case(
                                                When(~Q(car_name_code =''), then='car_name_code'),
                                                default='machine'
                                            )).values_list('mc', flat=True), #ใช้ในระบบงาน version ใหม่ดึงจาก car.name + car.code อันเก่าดึงจาก machine 10/10/2024
            'ประเภทการซ่อม': queryset.values_list('requisit__repair_type__name', flat=True),
            'เลขที่ใบขอเบิก': queryset.values_list('requisit__ref_no', flat=True),
            'รหัสสินค้า': queryset.values_list('product__id', flat=True),
            'สินค้า': queryset.values_list('product_name', flat=True),
            'จำนวน': queryset.values_list('quantity', flat=True),
            'หน่วย': queryset.values_list('unit', flat=True),
            'หน่วยงาน': queryset.values_list('requisit__agency__name', flat=True),
            'แผนกคชจ.': queryset.values_list('requisit__expense_dept__name', flat=True)
            }

    df1 = pd.DataFrame(data1)

    data2 = {
            'วันที่': 'รวมใบขอเบิก',
            'บริษัท': [queryset.values('requisit__id').distinct().count()],
            'ทะเบียนรถ/เครื่องจักร/หน่วยงาน': 'ใบ',
            'ประเภทการซ่อม': '',
            'เลขที่ใบขอเบิก': '',
            'รหัสสินค้า': '',
            'สินค้า': '',
            'จำนวน': '',
            'หน่วย': '',
            'หน่วยงาน': '',
            'แผนกคชจ.': '',
            }

    df2 = pd.DataFrame(data2, index=[0])

    frames = [df1, df2]
    result = pd.concat(frames)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Requisition_Report_({active}).xlsx'

    with pd.ExcelWriter(response, engine='xlsxwriter', options={'strings_to_numbers': True}) as writer:
        result.to_excel(writer, index=False)

    return response

def exportExcelIVToExpress(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    ref_no = request.GET.get('ref_no') or None
    bring_name = request.GET.get('bring_name') or None
    payer_name = request.GET.get('payer_name') or None
    expense_dept = request.GET.get('expense_dept') or None
    start_created = request.GET.get('start_created') or None
    end_created = request.GET.get('end_created') or None

    my_q = Q()
    if ref_no is not None:
        my_q &= Q(iv__ref_no__icontains = ref_no)
    if bring_name is not None:
        my_q &= Q(iv__bring_name__startswith = bring_name)
    if start_created is not None:
        my_q &= Q(iv__created__gte = start_created)
    if end_created is not None:
        my_q &=Q(iv__created__lte = end_created)
    if payer_name is not None:
        my_q &= Q(iv__payer_name = payer_name)
    if expense_dept is not None :
        my_q &=Q(iv__expense_dept = expense_dept)

    #ถ้ามีสิทธิดูรายงานของบริษัททั้งหมด ในแท็ป ALL จะดึงรายงานของทุกๆบริษัทมา
    if  is_view_report_all(request.user) and active == 'ALL':
        pass
    else:
        my_q &=Q(iv__branch_company__code__in = company_in)

    queryset = InvoiceItem.objects.filter(my_q).order_by('iv__id')
    if not queryset.exists():
        return HttpResponse("No data to export.")

    data1 = {'เลขที่ใบจ่ายสินค้าภายใน': queryset.values_list('iv__ref_no', flat=True),
            'วันที่': queryset.values_list('iv__created', flat=True),
            'รหัสบริษัท': queryset.values_list('iv__branch_company__code', flat=True),
            'บริษัท': queryset.values_list('iv__branch_company__name', flat=True),
            'แผนกคชจ.': queryset.values_list('iv__expense_dept__name', flat=True),
            'รหัสสินค้า' : queryset.values_list('item__product__id', flat=True),
            'รายการสินค้า' : queryset.values_list('item__product__name', flat=True),
            'จำนวน' : queryset.values_list('quantity', flat=True),
            'หน่วย' : queryset.values_list('unit', flat=True),
            'ผู้เบิก' : queryset.values_list(
                        Concat(F('iv__bring_name__first_name'),  Value(" "), F('iv__bring_name__last_name')), 
                        flat=True
                    ),
            'ผู้จ่าย' : queryset.values_list(
                        Concat(F('iv__payer_name__first_name'),  Value(" "), F('iv__payer_name__last_name')),
                        flat=True
                    ),
            'หน่วย' : queryset.values_list('unit', flat=True),
            'หมายเหตุ' :queryset.values_list(
                            Concat(F('iv__car__name'), F('iv__car__code')), 
                            flat=True
                        ),
            'หมายเหตุเพิ่มเติม' : queryset.values_list('iv__note', flat=True),
            }

    df1 = pd.DataFrame(data1)
    frames = [df1]
    result = pd.concat(frames)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=IB_to_Express_Report_({active}).xlsx'

    with pd.ExcelWriter(response, engine='xlsxwriter', options={'strings_to_numbers': True}) as writer:
        result.to_excel(writer, index=False)

    return response

#################################
############# API ###############
#################################

############# Login API ###############
class LoginApiView(APIView):
    permission_classes = []

    def post(self, request: Request):
        username = request.data.get("username")
        password = request.data.get("password")

        user = authenticate(username=username, password=password)
        if user is not None:
            tokens = create_jwt_pair_for_user(user)

            response = {"message": "Login Successfull", "tokens": tokens}
            return Response(data=response, status=status.HTTP_200_OK)

        else:
            return Response(data={"message": "Invalid email or password"})

    def get(self, request: Request):
        content = {"user": str(request.user), "auth": str(request.auth)}

        return Response(data=content, status=status.HTTP_200_OK)

############# SignUp API ###############   
class SignUpApiView(generics.GenericAPIView):
    serializer_class = SignUpSerializer
    permission_classes = []

    def post(self, request: Request):
        data = request.data

        serializer = self.serializer_class(data=data)

        if serializer.is_valid():
            serializer.save()

            response = {"message": "User Created Successfully", "data": serializer.data}

            return Response(data=response, status=status.HTTP_201_CREATED)

        return Response(data=serializer.errors, status=status.HTTP_400_BAD_REQUEST)

############# PO API ###############
class SmallResultsSetPagination(PageNumberPagination):
    page_size = 100  # or any number suitable for your frontend

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def apiOverviewPO(request):
    api_urls = {
        'Api Overview Po': '/po/api/',
        'Detail PO View':'/po/api/detail/<str:ref_no>/',
        'Detail PO Item View':'/po/items/api/detail/<str:ref_no>/',
        'Detail PO Item by Product Id View':'/po/product/api/detail/<str:ref_no>/<str:prod_id>',
        'All PO View Between Date':'/po/api/all/between/<str:start_date>/<str:end_date>/',
        'All PO Item View Between Date':'/po/items/api/all/between/<str:start_date>/<str:end_date>/',
    }
    return Response(api_urls)

############# PO  Items API ###############
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def detailPO(request, ref_no):
    queryset = PurchaseOrder.objects.get(ref_no = ref_no)
    serializer = PurchaseOrderSerializer(queryset, many = False)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def detailPOItems(request, ref_no):
    queryset = PurchaseOrderItem.objects.filter(po__ref_no = ref_no)
    serializer = PurchaseOrderItemSerializer(queryset, many = True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def detailPOProductItems(request, ref_no, prod_id):
    queryset = PurchaseOrderItem.objects.filter(po__ref_no = ref_no, item__product_id = prod_id)
    serializer = PurchaseOrderItemSerializer(queryset, many = True)
    items_as_dict = {str(i): item for i, item in enumerate(serializer.data)}
    return Response({"items": items_as_dict})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def allPO(request, start_date, end_date):
    queryset = PurchaseOrder.objects.filter(
       created__range=[start_date, end_date]
    ).select_related(
        'address_company',
        'po_type',
        'distributor',
        'vat_type',
        'credit',
        'delivery',
    ).prefetch_related(
        'purchaseorderitem_set__item__requisit__car'
    )
    paginator = SmallResultsSetPagination()
    result_page = paginator.paginate_queryset(queryset, request)
    serializer = PurchaseOrderSerializer(result_page, many=True)
    return paginator.get_paginated_response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def allPOItems(request, start_date, end_date):
    queryset = PurchaseOrderItem.objects.filter(
       po__created__range=[start_date, end_date]
    ).select_related(
        'po',
        'item__product',
        'unit',
    )
    paginator = SmallResultsSetPagination()
    result_page = paginator.paginate_queryset(queryset, request)
    serializer = PurchaseOrderItemSerializer(result_page, many=True)
    return paginator.get_paginated_response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def apiOverviewCar(request):
    api_urls = {
        'Api Overview Car':'/car/api/',
        'All Car View':'/car/api/all/',
    }
    return Response(api_urls)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def allCar(request):
    queryset = BaseCar.objects.all()
    paginator = SmallResultsSetPagination()
    result_page = paginator.paginate_queryset(queryset, request)
    serializer = BaseCarSerializer(result_page, many=True)
    return paginator.get_paginated_response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def apiOverviewDriverAndCar(request):
    api_urls = {
        'Api Overview Driver and Car':'/driver/car/api/',
        'All Car View':'/driver/car/api/all/',
    }
    return Response(api_urls)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def allDriverAndCar(request):
    queryset = UserCarDepartment.objects.filter(
       car__isnull = False
    )
    paginator = SmallResultsSetPagination()
    result_page = paginator.paginate_queryset(queryset, request)
    serializer = UserCarDepartmentSerializer(result_page, many=True)
    return paginator.get_paginated_response(serializer.data)

def exportExcelByExpense(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    stockman_user = request.GET.get('stockman_user') or None
    ref_no = request.GET.get('ref_no') or None
    distributor = request.GET.get('distributor') or None
    amount_min = request.GET.get('amount_min') or None
    amount_max = request.GET.get('amount_max') or None
    start_created = request.GET.get('start_created') or None
    end_created = request.GET.get('end_created') or None

    start_date, end_date = get_month_start_end()
    if not start_created:
        start_created = start_date
    if not end_date:
        end_created = end_date

    my_q = Q()
    if stockman_user is not None:
        my_q = Q(po__stockman_user = stockman_user)
    if ref_no is not None:
        my_q &= Q(po__ref_no__icontains = ref_no)
    if distributor is not None:
        my_q &= Q(po__distributor__name__startswith = distributor)
    if start_created is not None:
        my_q &= Q(po__created__gte = start_created)
    if end_created is not None:
        my_q &=Q(po__created__lte = end_created)
    if amount_min is not None:
        my_q &= Q(po__amount__gte = amount_min)
    if amount_max is not None :
        my_q &=Q(po__amount__lte = amount_max)

    my_q &=Q(po__approver_status = 2, po__is_cancel = False)
    my_q &=Q(item__requisit__agency__name__isnull = False, item__requisit__expense_dept__name__isnull = False)
    
    #ถ้ามีสิทธิดูรายงานของบริษัททั้งหมด ในแท็ป ALL จะดึงรายงานของทุกๆบริษัทมา
    if  is_view_report_all(request.user) and active == 'ALL':
        pass
    else:
        my_q &=Q(po__branch_company__code__in = company_in)

    # Your existing queryset
    queryset = PurchaseOrderItem.objects.filter(my_q).values(
        'item__requisit__agency__name',
        'item__requisit__expense_dept__name'
    ).annotate(
        sum_amount=Sum('price'),
        oil_lir=Sum(Case(When(item__product__category__slug='OL', then='quantity'), output_field=models.DecimalField())),
        oil_amount=Sum(Case(When(item__product__category__slug='OL', then='price'), output_field=models.DecimalField())),
        sum_other=Sum(Case(When(~Q(item__product__category__slug='OL'), then='price'), output_field=models.DecimalField()))
    ).order_by(
        'item__requisit__agency__name', 'item__requisit__expense_dept__name'
    )

    if not queryset.exists():
        return HttpResponse("No data to export.")

    # Prepare data
    internal_data = []
    external_data = []

    for item in queryset:
        row = {
            'แผนก': item['item__requisit__expense_dept__name'],
            'ลิตร': item['oil_lir'] or '',
            'บาท': item['oil_amount'] or '',
            'น้ำมันม.หล่อลื่น+': item['sum_other'] or '',
            'รวม': item['sum_amount'] or 0
        }
        
        if "ภายนอก" in item['item__requisit__agency__name']:
            external_data.append(row)
        else:
            internal_data.append(row)

    internal_data = prepare_data(internal_data) or []
    external_data = prepare_data(external_data) or []

    # สร้าง Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "รายงานค่าใช้จ่ายพัสดุ"

    # สไตล์
    header_font = Font(bold=True, size=12)
    section_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'),
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    start_created = safe_str_to_date(start_created) or start_date
    end_created = safe_str_to_date(end_created)  or end_date

    # ส่วนหัวหลัก
    ws.merge_cells('A1:E1')
    ws['A1'] = "PO รายงานค่าใช้จ่ายพัสดุ " + start_created.strftime('%d/%m/%Y') + " ถึง " + end_created.strftime('%d/%m/%Y')
    ws['A1'].font = header_font
    ws['A1'].alignment = center_alignment

    # หัวคอลัมน์
    ws.append(['แผนกคชจ.', 'น้ำมันโซล่า', '', 'น้ำมันม.หล่อลื่น+', 'รวมจำนวนเงิน'])
    ws.append(['', 'ลิตร', 'บาท', 'ค่าอะไหล่/สิ่งของ', 'บาท'])

    # Merge และจัดรูปแบบเซลล์ A2 และ A3
    ws.merge_cells('A2:A3')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A2'].font = Font(bold=True)
    ws['A2'] = "แผนก"  # ข้อความที่คุณต้องการแสดง
        
    # รวมหัวคอลัมน์และจัดสไตล์
    ws.merge_cells('B2:C2')
    for row in ws.iter_rows(min_row=2, max_row=3):
        for cell in row:
            cell.font = Font(bold=True)
            cell.alignment = center_alignment
            cell.border = border

    current_row = 4
    if internal_data:
        # ข้อมูลหน่วยงานภายใน
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws[f'A{current_row}'] = "หน่วยงานภายใน"
        ws[f'A{current_row}'].font = section_font
        ws[f'A{current_row}'].alignment = center_alignment
        current_row += 1

        for item in internal_data:
            ws.append([
                item['แผนก'],
                item['ลิตร'],
                item['บาท'],
                item['น้ำมันม.หล่อลื่น+'],
                item['รวม']
            ])
            current_row += 1

        # ผลรวมหน่วยงานภายใน
        ws.append([
            "รวมหน่วยงานภายใน",
            safe_sum(internal_data, 'ลิตร'),
            safe_sum(internal_data, 'บาท'),
            safe_sum(internal_data, 'น้ำมันม.หล่อลื่น+'),
            safe_sum(internal_data, 'รวม')
        ])
        for col in range(1, 6):
            cell = ws.cell(row=current_row, column=col)
            cell.font = Font(bold=True)
            if col in [2, 3, 4, 5]:  # Numeric columns
                cell.number_format = '#,##0.00'
        current_row += 1

    if external_data:
        # ข้อมูลหน่วยงานภายนอก
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws[f'A{current_row}'] = "หน่วยงานภายนอก"
        ws[f'A{current_row}'].font = section_font
        ws[f'A{current_row}'].alignment = center_alignment
        current_row += 1

        for item in external_data:
            ws.append([
                item['แผนก'],
                item['ลิตร'],
                item['บาท'],
                item['น้ำมันม.หล่อลื่น+'],
                item['รวม']
            ])
            current_row += 1
            
        # ผลรวมหน่วยงานภายนอก
        ws.append([
            "รวมหน่วยงานภายนอก",
            safe_sum(external_data, 'ลิตร'),
            safe_sum(external_data, 'บาท'),
            safe_sum(external_data, 'น้ำมันม.หล่อลื่น+'),
            safe_sum(external_data, 'รวม')
        ])
        for col in range(1, 6):
            cell = ws.cell(row=current_row, column=col)
            cell.font = Font(bold=True)
            if col in [2, 3, 4, 5]:  # Numeric columns
                cell.number_format = '#,##0.00'
        current_row += 1

    # Add one empty row as separator
    ws.append([''] * 5)  # Empty row separator

    # Add GRAND TOTAL row (sum of both internal and external)
    grand_total_row = current_row + 1

    # Calculate grand totals
    grand_total_liter = safe_sum(internal_data, 'ลิตร') + safe_sum(external_data, 'ลิตร')
    grand_total_baht = safe_sum(internal_data, 'บาท') + safe_sum(external_data, 'บาท')
    grand_total_other = safe_sum(internal_data, 'น้ำมันม.หล่อลื่น+') + safe_sum(external_data, 'น้ำมันม.หล่อลื่น+')
    grand_total_amount = safe_sum(internal_data, 'รวม') + safe_sum(external_data, 'รวม')

    # Create grand total row (showing all sums)
    ws[f'A{grand_total_row}'] = "รวมทั้งหมด"
    ws[f'B{grand_total_row}'] = grand_total_liter
    ws[f'C{grand_total_row}'] = grand_total_baht
    ws[f'D{grand_total_row}'] = grand_total_other
    ws[f'E{grand_total_row}'] = grand_total_amount

    # Apply formatting to grand total row
    for col in range(1, 6):
        cell = ws.cell(row=grand_total_row, column=col)
        cell.font = Font(bold=True)
        if col in [2, 3, 4, 5]:  # Numeric columns
            cell.number_format = '#,##0.00'
        if col == 1:  # Label column
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Apply special border style
    for col in range(1, 6):
        ws.cell(row=grand_total_row, column=col).border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='double')  # Double line for emphasis
        )

    # ใส่กรอบให้ทุกเซลล์
    for row in ws.iter_rows(min_row=1, max_row=current_row, max_col=5):
        for cell in row:
            cell.border = border

    # Adjust column widths if needed
    column_widths = {'A': 40, 'B': 15, 'C': 15, 'D': 20, 'E': 15}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # สร้าง HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Expense by Department({active}).xlsx'
    
    wb.save(response)
    return response

# แก้ไขการคำนวณผลรวมให้รองรับค่าว่าง
def safe_sum(items, key):
    return sum(float(item[key]) for item in items if item[key] and str(item[key]).replace('.','',1).isdigit())

# แปลงข้อมูลให้ปลอดภัย
def prepare_data(data):
    for item in data:
        for key in item:
            if item[key] is None:
                item[key] = ''
            elif isinstance(item[key], (int, float)):
                continue
            elif str(item[key]).replace('.','',1).isdigit():
                item[key] = float(item[key])
        return data
    
def get_month_start_end(date=None):
    if date is None:
        date = datetime.date.today()
    
    start_date = date.replace(day=1)  # First day of the month
    next_month = date.month % 12 + 1
    next_month_year = date.year + (1 if next_month == 1 else 0)
    end_date = datetime.date(next_month_year, next_month, 1) - datetime.timedelta(days=1)  # Last day of the month
    
    return start_date, end_date

def safe_str_to_date(date_str):
    if date_str and isinstance(date_str, str):  # Check if it's a valid string
        try:
            return datetime.datetime.strptime(date_str, '%Y-%m-%d')  # Convert only if valid
        except ValueError:
            pass  # In case of an invalid format, ignore and return None
    return None  # Return None if input is invalid or empty

# Flow เก่า 29/04/2568
def viewExInvoice_old(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)
    '''
    data = ExOESTNH.objects.using('pg_db').filter(comcod = 'SLC').order_by('-docdat', '-docnum')

    #กรองข้อมูล
    myFilter = ExOESTNHFilter(request.GET, queryset = data)
    data = myFilter.qs

    '''
    data = Invoice.objects.filter(is_express = True, branch_company__code__in = company_in)

    #กรองข้อมูล
    myFilter = InvoidFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)
    
    context = {
        'ivs': dataPage,
        'filter':myFilter,
        'ex_old_iv_page': "tab-active",
        'ex_old_iv_show': "show",
        active :"active show",
        "colorNav":"enableNav"
    }
    return render(request, "express/viewExInvoice_old.html", context)

#เฉพาะศิลาชัยก่อน 28/04/2025
def updateInvoiceFromExpress(request):
    if request.method == "GET":
        active = request.session.get('company_code')
        company_in = findCompanyIn(request)
        b_com = BaseBranchCompany.objects.filter(code = active)

        update_invoice(b_com)

        return JsonResponse({"success": True, "message": "Invoices updated successfully"})

    return JsonResponse({"success": False, "message": "Invalid request"}, status=400)

#เฉพาะศิลาชัยก่อน 28/04/2025
def update_invoice(all_company):
    for b_com in all_company:
        find_iv_id = list(
            InvoiceItem.objects.filter(
                is_express=False,
                iv__ref_no__startswith = b_com.invoice_code,
                iv__branch_company__affiliated__name = b_com.affiliated.name
            )
            .values_list('iv__ref_no', flat=True)
            .order_by('iv__ref_no').distinct()
        )
            
        iv_ex = ExOESTND.objects.using('pg_db').filter(
            docnum__in = find_iv_id , comcod = b_com.affiliated.name
        ).order_by('docnum', 'seqnum').values('seqnum', 'docnum', 'stkcod', 'ordqty', 'unitpr', 'trnval')

        for ex in iv_ex:
            invoice_item = InvoiceItem.objects.filter(
                iv__ref_no=ex['docnum'], 
                item__product__id=ex['stkcod'], 
                iv__branch_company__affiliated__name = b_com.affiliated.name,
            ).first()

            if invoice_item:
                invoice_item.is_express = True
                invoice_item.quantity = ex['ordqty']
                invoice_item.unit_price = ex['unitpr']
                invoice_item.price = ex['trnval']
                invoice_item.save()

        try:
            af_iv = InvoiceItem.objects.filter(
                is_express = True,
                iv__branch_company__affiliated__name = b_com.affiliated.name,
                iv__ref_no__startswith = b_com.invoice_code,
            ).values_list('iv', flat=True).distinct()

            iv = Invoice.objects.filter(id__in=af_iv, is_express=False)
            for i in iv:
                i.total_price = InvoiceItem.objects.filter(iv=i.id).aggregate(total_sum=Sum('price'))['total_sum']
                i.is_express = True
                i.save()
        except Invoice.DoesNotExist:
            pass

#เฉพาะศิลาชัยก่อน 28/04/2025
def send_4am_summary():
    try:
        #all_company = BaseBranchCompany.objects.filter(~Q(code = 'ALL'))
        all_company = BaseBranchCompany.objects.filter(code = 'SLC')
        update_invoice(all_company)
    except :
        pass

''' ปิด BackgroundScheduler 08-05-2025
# Schedule the tasks
#เฉพาะศิลาชัยก่อน 28/04/2025
scheduler = BackgroundScheduler()
scheduler.add_job(
    send_4am_summary,
    trigger=CronTrigger(hour=4, minute=0),
    id="4am_summary",
    replace_existing=True,
)
scheduler.start()
'''

def exportExcelByIVExpense(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)
    b_com = BaseBranchCompany.objects.get(code = active)

    start_created = request.GET.get('start_created') or None
    end_created = request.GET.get('end_created') or None

    start_date, end_date = get_month_start_end()
    if not start_created:
        start_created = start_date
    if not end_date:
        end_created = end_date

    my_q = Q()
    if start_created is not None:
        my_q &= Q(iv__created__gte = start_created)
    if end_created is not None:
        my_q &=Q(iv__created__lte = end_created)

    my_q &=Q(is_express = True)
    my_q &=Q(iv__requisit__agency__name__isnull = False, iv__expense_dept__isnull = False)
    
    #ถ้ามีสิทธิดูรายงานของบริษัททั้งหมด ในแท็ป ALL จะดึงรายงานของทุกๆบริษัทมา
    if  is_view_report_all(request.user) and active == 'ALL':
        pass
    else:
        my_q &=Q(iv__branch_company__code__in = company_in)

    iv_filters = Q()
    if b_com.invoice_code:
        iv_filters |= Q(docnum__startswith=b_com.invoice_code)
    if b_com.oi_invoice_code:
        iv_filters |= Q(docnum__startswith=b_com.oi_invoice_code)

    #หน่วยงานภายใน - อะไหล่
    l_iv = list(
        ExOESTNH.objects.using('pg_db')
        .filter(
            iv_filters,
            comcod=b_com.affiliated.name,
            docdat__range=(start_created, end_created)
        )
        .values_list('docnum', flat=True)
    )

    in_queryset = ExOESTND.objects.using('pg_db').filter(
        comcod=b_com.affiliated.name,
        docnum__in=l_iv
    ).extra(
        select={
            'depcod': """(
                SELECT depcod 
                FROM "OESTNH" 
                WHERE "OESTNH"."docnum" = "OESTND"."docnum" 
                AND "OESTNH"."comcod" = "OESTND"."comcod" 
                LIMIT 1
            )"""
        }
    ).values(
        'depcod'
    ).annotate(
        sum_amount=Sum('trnval'),
        oil_lir=Sum(Case(
            When(stktyp ='น้ำมันดีเซล', then='ordqty'),
            output_field=models.DecimalField()
        )),
        eg_o_amount=Sum(Case(
            When(stktyp ='น้ำมันหล่อลื่น', then='trnval'),
            output_field=models.DecimalField()
        )),
        eg_o_oil_lir=Sum(Case(
            When(stktyp ='น้ำมันหล่อลื่น', then='ordqty'),
            output_field=models.DecimalField()
        )),
        oil_amount=Sum(Case(
            When(stktyp ='น้ำมันดีเซล', then='trnval'),
            output_field=models.DecimalField()
        )),
        sum_other=Sum(Case(
            When(stktyp ='อะไหล่', then='trnval'),
            output_field=models.DecimalField()
        ))
    ).order_by('depcod')

    #หน่วยงานภายนอก - อะไหล่
    soc_filters = Q()
    if b_com.soc_code:
        soc_filters |= Q(docnum__startswith=b_com.soc_code)
    if b_com.oi_soc_code:
        soc_filters |= Q(docnum__startswith=b_com.oi_soc_code)

    l_soc = list(
                    ExOEINVH.objects.using('pg_db').filter(soc_filters, comcod = b_com.affiliated.name, docdate__range=(start_created, end_created)).values_list('docnum', flat=True)
                )

    ex_queryset = ExOEINVD.objects.using('pg_db').filter(
        comcod = b_com.affiliated.name,
        docnum__in=l_soc
    ).extra(
        select={
            'cusnam': """(
                SELECT cusnam 
                FROM "OEINVH" 
                WHERE "OEINVH"."docnum" = "OEINVD"."docnum" 
                AND "OEINVH"."comcod" = "OEINVD"."comcod" 
                LIMIT 1
            )"""
        }
    ).values(
        'cusnam'
    ).annotate(
        sum_amount=Sum('trnval'),
        oil_lir=Sum(Case(
            When(stktyp ='น้ำมันดีเซล', then='ordqty'),
            output_field=models.DecimalField()
        )),
        eg_o_amount=Sum(Case(
            When(stktyp ='น้ำมันหล่อลื่น', then='trnval'),
            output_field=models.DecimalField()
        )),
        eg_o_oil_lir=Sum(Case(
            When(stktyp ='น้ำมันหล่อลื่น', then='ordqty'),
            output_field=models.DecimalField()
        )),
        oil_amount=Sum(Case(
            When(stktyp ='น้ำมันดีเซล', then='trnval'),
            output_field=models.DecimalField()
        )),
        sum_other=Sum(Case(
            When(stktyp ='อะไหล่', then='trnval'),
            output_field=models.DecimalField()
        ))
    ).order_by('cusnam')

    if not in_queryset.exists():
        return HttpResponse("No data to export.")

    # Prepare data
    internal_data = []
    external_data = []

    # 1. สร้าง Dict depcod -> name
    dept_map = {
        str(d['id']): d['name']
        for d in BaseExpenseDepartment.objects.filter(
            id__in=[item['depcod'].strip() for item in in_queryset if item['depcod'] is not None]
        ).values('id', 'name')
    }

    # 2. วนลูปสร้าง internal_data
    internal_data = []
    for item in in_queryset:
        dep_name = dept_map.get(item['depcod'].strip(), item['depcod'].strip())  # ถ้าไม่มี depcod ก็ใส่ '-'
        row = {
            'แผนก': dep_name,
            'ลิตร': item['oil_lir'] or '-',
            'บาท': item['oil_amount'] or '-',
            'ลิตร.': item['eg_o_oil_lir'] or '-',
            'บาท.': item['eg_o_amount'] or '-',
            'ค่าอะไหล่': item['sum_other'] or '-',
            'รวม': item['sum_amount'] or 0
        }
        internal_data.append(row)

    for item in ex_queryset:
        row = {
            'แผนก': item['cusnam'] or '-',
            'ลิตร': item['oil_lir'] or '-',
            'บาท': item['oil_amount'] or '-',
            'ลิตร.': item['eg_o_oil_lir'] or '-',
            'บาท.': item['eg_o_amount'] or '-',
            'ค่าอะไหล่': item['sum_other'] or '-',
            'รวม': item['sum_amount'] or 0
        }
        external_data.append(row)

    internal_data = prepare_data(internal_data) or []
    external_data = prepare_data(external_data) or []

    # สร้าง Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "รายงานค่าใช้จ่ายพัสดุ"

    # สไตล์
    header_font = Font(bold=True, size=12)
    section_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'),
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    start_created = safe_str_to_date(start_created) or start_date
    end_created = safe_str_to_date(end_created)  or end_date

    # ส่วนหัวหลัก
    ws.merge_cells('A1:G1')
    ws['A1'] = "รายงานค่าใช้จ่ายพัสดุ " + start_created.strftime('%d/%m/%Y') + " ถึง " + end_created.strftime('%d/%m/%Y')
    ws['A1'].font = header_font
    ws['A1'].alignment = center_alignment

    # หัวคอลัมน์
    ws.append(['แผนกคชจ.', 'น้ำมันโซล่า', '', 'น้ำมันหล่อลื่น', '', 'ค่าอะไหล่/สิ่งของ+', 'รวมจำนวนเงิน'])
    ws.append(['', 'ลิตร', 'บาท', 'ลิตร.', 'บาท.', 'ค่าแรง', 'บาท'])

    # Merge และจัดรูปแบบเซลล์ A2 และ A3
    ws.merge_cells('A2:A3')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A2'].font = Font(bold=True)
    ws['A2'] = "แผนก"  # ข้อความที่คุณต้องการแสดง
        
    # รวมหัวคอลัมน์และจัดสไตล์
    ws.merge_cells('B2:C2')
    ws.merge_cells('D2:E2')
    for row in ws.iter_rows(min_row=2, max_row=3):
        for cell in row:
            cell.font = Font(bold=True)
            cell.alignment = center_alignment
            cell.border = border

    current_row = 4
    if internal_data:
        # ข้อมูลหน่วยงานภายใน
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws[f'A{current_row}'] = "หน่วยงานภายใน"
        ws[f'A{current_row}'].font = section_font
        ws[f'A{current_row}'].alignment = center_alignment
        current_row += 1

        for item in internal_data:
            ws.append([
                item['แผนก'],
                item['ลิตร'],
                item['บาท'],
                item['ลิตร.'],
                item['บาท.'],
                item['ค่าอะไหล่'],
                item['รวม']
            ])
            for col in range(1, 8):
                cell = ws.cell(row=current_row, column=col)
                if col in [2, 3, 4, 5, 6, 7]:  # Numeric columns
                    cell.number_format = '#,##0.00'
            current_row += 1

        # ผลรวมหน่วยงานภายใน
        ws.append([
            "รวมหน่วยงานภายใน",
            safe_sum(internal_data, 'ลิตร'),
            safe_sum(internal_data, 'บาท'),
            safe_sum(internal_data, 'ลิตร.'),
            safe_sum(internal_data, 'บาท.'),
            safe_sum(internal_data, 'ค่าอะไหล่'),
            safe_sum(internal_data, 'รวม')
        ])
        for col in range(1, 8):
            cell = ws.cell(row=current_row, column=col)
            cell.font = Font(bold=True)
            if col in [2, 3, 4, 5, 6, 7]:  # Numeric columns
                cell.number_format = '#,##0.00'
        current_row += 1

    if external_data:
        # ข้อมูลหน่วยงานภายนอก
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws[f'A{current_row}'] = "หน่วยงานภายนอก"
        ws[f'A{current_row}'].font = section_font
        ws[f'A{current_row}'].alignment = center_alignment
        current_row += 1

        for item in external_data:
            ws.append([
                item['แผนก'],
                item['ลิตร'],
                item['บาท'],
                item['ลิตร.'],
                item['บาท.'],
                item['ค่าอะไหล่'],
                item['รวม']
            ])
            for col in range(1, 8):
                cell = ws.cell(row=current_row, column=col)
                if col in [2, 3, 4, 5, 6, 7]:  # Numeric columns
                    cell.number_format = '#,##0.00'
            current_row += 1
            
        # ผลรวมหน่วยงานภายนอก
        ws.append([
            "รวมหน่วยงานภายนอก",
            safe_sum(external_data, 'ลิตร'),
            safe_sum(external_data, 'บาท'),
            safe_sum(external_data, 'ลิตร.'),
            safe_sum(external_data, 'บาท.'),
            safe_sum(external_data, 'ค่าอะไหล่'),
            safe_sum(external_data, 'รวม')
        ])
        for col in range(1, 8):
            cell = ws.cell(row=current_row, column=col)
            cell.font = Font(bold=True)
            if col in [2, 3, 4, 5, 6, 7]:  # Numeric columns
                cell.number_format = '#,##0.00'
        current_row += 1

    # Add one empty row as separator
    ws.append([''] * 7)  # Empty row separator

    # Add GRAND TOTAL row (sum of both internal and external)
    grand_total_row = current_row + 1

    # Calculate grand totals
    grand_total_liter = safe_sum(internal_data, 'ลิตร') + safe_sum(external_data, 'ลิตร')
    grand_total_baht = safe_sum(internal_data, 'บาท') + safe_sum(external_data, 'บาท')
    grand_eg_total_liter = safe_sum(internal_data, 'ลิตร.') + safe_sum(external_data, 'ลิตร.')
    grand_eg_total_baht = safe_sum(internal_data, 'บาท.') + safe_sum(external_data, 'บาท.')
    grand_total_other = safe_sum(internal_data, 'ค่าอะไหล่') + safe_sum(external_data, 'ค่าอะไหล่')
    grand_total_amount = safe_sum(internal_data, 'รวม') + safe_sum(external_data, 'รวม')

    # Create grand total row (showing all sums)
    ws[f'A{grand_total_row}'] = "รวมทั้งหมด"
    ws[f'B{grand_total_row}'] = grand_total_liter
    ws[f'C{grand_total_row}'] = grand_total_baht
    ws[f'D{grand_total_row}'] = grand_eg_total_liter
    ws[f'E{grand_total_row}'] = grand_eg_total_baht
    ws[f'F{grand_total_row}'] = grand_total_other
    ws[f'G{grand_total_row}'] = grand_total_amount

    # Apply formatting to grand total row
    for col in range(1, 8):
        cell = ws.cell(row=grand_total_row, column=col)
        cell.font = Font(bold=True)
        if col in [2, 3, 4, 5, 6, 7]:  # Numeric columns
            cell.number_format = '#,##0.00'
        if col == 1:  # Label column
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Apply special border style
    for col in range(1, 8):
        ws.cell(row=grand_total_row, column=col).border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='double')  # Double line for emphasis
        )

    # ใส่กรอบให้ทุกเซลล์
    for row in ws.iter_rows(min_row=1, max_row=current_row, max_col=7):
        for cell in row:
            cell.border = border

    # Adjust column widths if needed
    column_widths = {'A': 40, 'B': 15, 'C': 15, 'D': 15, 'E': 15, 'F': 20, 'G': 15}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # สร้าง HttpResponse
    output = BytesIO()
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Expense by Department({active}).xlsx'
    
    wb.save(response)
    return response

def exportToExcelRegistrationAndRepair(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)
    b_com = BaseBranchCompany.objects.get(code = active)

    start_created = request.GET.get('start_created') or None
    end_created = request.GET.get('end_created') or None

    start_date, end_date = get_month_start_end()
    if not start_created:
        start_created = start_date
    if not end_date:
        end_created = end_date

    my_q = Q()
    if start_created is not None:
        my_q &= Q(docdat__gte = start_created)
    if end_created is not None:
        my_q &=Q(docdat__lte = end_created)

    my_q &=Q(comcod = b_com.affiliated.name)
    my_q &=Q(docnum__startswith=b_com.invoice_code)

    if isinstance(start_created, str):
        date_start = datetime.datetime.strptime(start_created, '%Y-%m-%d').date()

    if isinstance(end_created, str):
        date_end = datetime.datetime.strptime(end_created, '%Y-%m-%d').date()

    # Generate (month, year) pairs within the selected range
    month_year_range = [
        (dt.month, dt.year) for dt in rrule(MONTHLY, dtstart=date_start, until=date_end)
    ]

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    thai_months = {
        '1': 'ม.ค.', '2': 'ก.พ.', '3': 'มี.ค.', '4': 'เม.ย.', '5': 'พ.ค.', '6': 'มิ.ย.',
        '7': 'ก.ค.', '8': 'ส.ค.', '9': 'ก.ย.', '10': 'ต.ค.', '11': 'พ.ย.', '12': 'ธ.ค.'
    }

    base_car = ExOESTNH.objects.using('pg_db').filter(
       my_q, 
       remark__isnull = False,
    ).values(
        'remark'
    ).distinct().order_by('remark')

    docdat_subquery = ExOESTNH.objects.using('pg_db').filter(
        remark__isnull = False
    ).filter(
        my_q
    ).filter(
        docnum=OuterRef('docnum'),
        comcod=OuterRef('comcod'),
    ).values('docdat')[:1]

    note2_subquery = ExOESTNH.objects.using('pg_db').filter(
        remark__isnull = False
    ).filter(
        my_q
    ).filter(
        docnum=OuterRef('docnum'),
        comcod=OuterRef('comcod'),
    ).values('note2')[:1]

    if not base_car.exists():
        return HttpResponse("No data to export.")

    workbook = openpyxl.Workbook()
    if base_car:
        for car in base_car:
            queryset = ExOESTND.objects.using('pg_db').filter(
                Q(docnum__startswith=b_com.invoice_code),
                comcod=b_com.affiliated.name,
                stkdes=car['remark']
            ).annotate(
                docdat=Subquery(docdat_subquery)
            ).values(
                'docdat'
            ).annotate(
                note2=Subquery(note2_subquery)
            ).values(
                'note2'
            ).annotate(
                year=ExtractYear('docdat'),
                month=ExtractMonth('docdat'),
            ).annotate(
                month_year=Concat(
                    Cast('month', models.CharField()),
                    Value('-'),
                    Cast('year', models.CharField()),
                    output_field=models.CharField()
                )
            ).values(
                'note2',
                'month_year', 'year', 'month'
            ).annotate(
                sum_amount=Sum(Case(
                    When(~Q(stkcod__startswith='OL'), then='trnval'),
                    output_field=models.DecimalField()
                ))
            ).order_by('year', 'month')

            try:
                car_name, car_code = car['remark'].split(":")
            except:
                car_code = ''
                car_name = car['remark']
            
            sheet_name = sanitize_sheet_title(car_name + " " + car_code)
            sheet = workbook.create_sheet(title=sheet_name)

            headers1 = [ car_name + " " + car_code ] + [f"{thai_months[str(m)]} {y}" for m, y in month_year_range]
            sheet.append(headers1)

            headers2 = [''] + ['ค่าอะไหล่ + ค่าแรง'] * len(month_year_range)
            sheet.append(headers2)

            sheet.merge_cells(f'A1:A2')

            header_row = next(sheet.iter_rows(min_row=1, max_row=1))
            header_map = {cell.value: idx for idx, cell in enumerate(header_row)}

            # Group data by repair type
            grouped_data = defaultdict(dict)
            for item in queryset:
                if item['year'] and item['month'] and item['note2']:
                    month = str(int(item['month']))  # convert to int, then back to str
                    year = str(item['year'])

                    #ดึงประเภทการซ่อมจากในระบบ
                    try:
                        rt = BaseRepairType.objects.get(id = item['note2'].strip())
                        repair_type = rt.name
                    except BaseRepairType.DoesNotExist:
                        repair_type = str(item['note2'])

                    month_th = f"{thai_months[month]} {year}"
                    amount = float(item['sum_amount'] or 0)
                    grouped_data[repair_type][month_th] = amount

            for repair_type, month_data in grouped_data.items():
                row_data = [''] * len(header_row)
                row_data[0] = repair_type
                for month, amount in month_data.items():
                    col_idx = header_map.get(month)
                    if col_idx is not None:
                        row_data[col_idx] = float(amount)
                sheet.append(row_data)

            column_index = len(month_year_range) + 2
            row_index = sheet.max_row + 1

            #คำนวนรวมทั้งสิ้น
            sheet.cell(row=row_index, column=1, value='รวมทั้งสิ้น')
            sum_by_col = Decimal('0.00')
            for col in range(2, column_index):
                for row in range(3, row_index):
                    sum_by_col = sum_by_col + Decimal( sheet.cell(row=row, column=col).value or '0.00' )
                sheet.cell(row=row_index, column=col, value=sum_by_col).number_format = '#,##0.00'
                sheet.cell(row=row_index, column=col).font = Font(bold=True)
                sum_by_col = Decimal('0.00')

            # Set column widths , number format and กรอบ
            for column_cells in sheet.columns:
                max_length = 0
                column = column_cells[2].column_letter
                for cell in column_cells:
                    cell.border = thin_border
                    try:
                        if isinstance(cell.value, float):
                            cell.number_format = '#,##0.00'
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = max_length + 2
                sheet.column_dimensions[column].width = adjusted_width
                sheet.column_dimensions[column].height = 20

        workbook.remove(workbook['Sheet'])
    else:
        worksheet = workbook.active
        worksheet.cell(row = 1, column = 1, value = f'ไม่มีข้อมูลรายงานสรุปตามทะเบียนรถและประเภทการซ่อม')

    output = BytesIO()
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Type of Repair({active}).xlsx'

    workbook.save(response)
    return response


def exportToExcelAllExpensesRegistration(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)
    b_com = BaseBranchCompany.objects.get(code = active)

    start_created = request.GET.get('start_created') or None
    end_created = request.GET.get('end_created') or None

    start_date, end_date = get_month_start_end()
    if not start_created:
        start_created = start_date
    if not end_date:
        end_created = end_date

    my_q = Q()
    if start_created is not None:
        my_q &= Q(docdat__gte = start_created)
    if end_created is not None:
        my_q &=Q(docdat__lte = end_created)

    my_q &=Q(comcod = b_com.affiliated.name)
    my_q &=Q(docnum__startswith=b_com.invoice_code) | Q(docnum__startswith=b_com.oi_invoice_code)

    if isinstance(start_created, str):
        date_start = datetime.datetime.strptime(start_created, '%Y-%m-%d').date()

    if isinstance(end_created, str):
        date_end = datetime.datetime.strptime(end_created, '%Y-%m-%d').date()

    # Generate (month, year) pairs within the selected range
    month_year_range = [
        (dt.month, dt.year) for dt in rrule(MONTHLY, dtstart=date_start, until=date_end)
    ]

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    thai_months = {
        "1": "ม.ค.", "2": "ก.พ.", "3": "มี.ค.", "4": "เม.ย.",
        "5": "พ.ค.", "6": "มิ.ย.", "7": "ก.ค.", "8": "ส.ค.",
        "9": "ก.ย.", "10": "ต.ค.", "11": "พ.ย.", "12": "ธ.ค."
    }

    all_car = ExOESTNH.objects.using('pg_db').filter(
       my_q
    ).values('remark').distinct().order_by('remark')

    docdat_subquery = ExOESTNH.objects.using('pg_db').filter(
        my_q
    ).filter(
        docnum=OuterRef('docnum'),
        comcod=OuterRef('comcod'),
    ).values('docdat')[:1]

    if not all_car.exists():
        return HttpResponse("No data to export.")

    # Create Excel workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Generate dynamic headers
    headers1 = ['ลำดับ', 'รหัส', 'ทะเบียนรถ/เครื่องจักร'] + [
        f"{thai_months[str(m)]} {y}" for m, y in month_year_range for _ in range(5)
    ]
    sheet.append(headers1)

    headers2 = ['ลำดับ', 'รหัส', 'ทะเบียนรถ/เครื่องจักร'] + ['น้ำมันดีเซล', 'น้ำมันดีเซล', 'น้ำมันหล่อลื่น', 'น้ำมันหล่อลื่น', 'ค่าอะไหล่+'] * len(month_year_range)
    sheet.append(headers2)

    headers3 = ['ลำดับ', 'รหัส', 'ทะเบียนรถ/เครื่องจักร'] + ['ลิตร', 'บาท', 'ลิตร.', 'บาท.', 'ค่าแรง'] * len(month_year_range)
    sheet.append(headers3)

    sheet.merge_cells('A1:A3')
    sheet.merge_cells('B1:B3')
    sheet.merge_cells('C1:C3')

    if all_car:
        for index, car in enumerate(all_car, start=1):
            iv = ExOESTND.objects.using('pg_db').filter(
                Q(docnum__startswith=b_com.invoice_code) | Q(docnum__startswith=b_com.oi_invoice_code),
                comcod=b_com.affiliated.name,
                stkdes=car['remark']
            ).annotate(
                docdat=Subquery(docdat_subquery)
            ).values('docdat').annotate(
                year=ExtractYear('docdat'),
                month=ExtractMonth('docdat'),
            ).annotate(
                month_year=Concat(
                    Cast('month', models.CharField()),
                    Value('-'),
                    Cast('year', models.CharField()),
                    output_field=models.CharField()
                )
            ).values('month_year', 'year', 'month').annotate(
                sum_amount=Sum('trnval'),
                oil_lir=Sum(Case(
                    When(stktyp ='น้ำมันดีเซล', then='ordqty'),
                    output_field=models.DecimalField()
                )),
                eg_o_amount=Sum(Case(
                    When(stktyp ='น้ำมันหล่อลื่น', then='trnval'),
                    output_field=models.DecimalField()
                )),
                eg_o_oil_lir=Sum(Case(
                    When(stktyp ='น้ำมันหล่อลื่น', then='ordqty'),
                    output_field=models.DecimalField()
                )),
                oil_amount=Sum(Case(
                    When(stktyp ='น้ำมันดีเซล', then='trnval'),
                    output_field=models.DecimalField()
                )),
                sum_other=Sum(Case(
                    When(stktyp ='อะไหล่', then='trnval'),
                    output_field=models.DecimalField()
                ))
            ).order_by('year', 'month')

            iv = [item for item in iv if item['month'] is not None and item['year'] is not None]
            iv.sort(key=itemgetter('year', 'month'))

            sum_dict = defaultdict(Decimal)
            sum_oil_l_dict = defaultdict(Decimal)
            sum_oil_a_dict = defaultdict(Decimal)
            sum_eg_l_dict = defaultdict(Decimal)
            sum_eg_a_dict = defaultdict(Decimal)
            

            for key, group in groupby(iv, key=itemgetter('month', 'year')):
                for item in group:
                    sum_dict[key] += item['sum_other'] or Decimal('0')
                    sum_oil_l_dict[key] += item['oil_lir'] or Decimal('0')
                    sum_oil_a_dict[key] += item['oil_amount'] or Decimal('0')
                    sum_eg_l_dict[key] += item['eg_o_oil_lir'] or Decimal('0')
                    sum_eg_a_dict[key] += item['eg_o_amount'] or Decimal('0')

            try:
                car_name, car_code = car['remark'].split(":")
            except:
                car_code = ''
                car_name = car['remark']

            row1 = [index, car_code, car_name] + [
                val for m, y in month_year_range for val in (
                    sum_oil_l_dict.get((m, y), ''),
                    sum_oil_a_dict.get((m, y), ''),
                    sum_eg_l_dict.get((m, y), ''),
                    sum_eg_a_dict.get((m, y), ''),
                    sum_dict.get((m, y), ''),
                )
            ]
            sheet.append(row1)

        #merge_cells เดือนเดียวกัน
        count = 3
        for i in range(len(month_year_range)):
            #merge_cells เดือนเดียวกัน
            sheet.merge_cells(start_row = 1, start_column = 1 + count, end_row = 1, end_column = 5 + count)
            #merge_cells น้ำมันดีเซล
            sheet.merge_cells(start_row = 2, start_column = 1 + count, end_row = 2, end_column = 2 + count)
            #merge_cells น้ำมันหล่อลื่น
            sheet.merge_cells(start_row = 2, start_column = 3 + count, end_row = 2, end_column = 4 + count)
            count = count + 5


        #คำนวนรวมทั้งสิ้น
        column_index = sheet.max_column + 1
        row_index = sheet.max_row + 1

        sheet.cell(row=row_index, column=1, value='รวมทั้งสิ้น')
        sum_by_col = Decimal('0.00')

        for col in range(4, column_index):
            for row in range(4, row_index):
                sum_by_col = sum_by_col + Decimal( sheet.cell(row=row, column=col).value or '0.00' )
            sheet.cell(row=row_index, column=col, value=sum_by_col).number_format = '#,##0.00'
            sheet.cell(row=row_index, column=col).font = Font(bold=True)
            sum_by_col = Decimal('0.00')
        
        # ===== Sum by row (เฉพาะเงิน → คอลัมน์เดียว) =====
        start_row = 4
        end_row = sheet.max_row
        start_col = 4
        cols_per_month = 5
        month_count = len(month_year_range)

        total_col = sheet.max_column + 1

        sheet.cell(row=1, column=total_col, value='รวมเงิน')
        sheet.cell(row=2, column=total_col, value='รวมเงิน')
        sheet.cell(row=3, column=total_col, value='บาท')
        sheet.merge_cells(start_row=1, start_column=total_col, end_row=2, end_column=total_col)

        for r in range(start_row, end_row + 1):
            total_money = Decimal('0')
            for i in range(month_count):
                base = start_col + (i * cols_per_month)
                total_money += Decimal(sheet.cell(r, base + 1).value or 0)
                total_money += Decimal(sheet.cell(r, base + 3).value or 0)
                total_money += Decimal(sheet.cell(r, base + 4).value or 0)

            cell = sheet.cell(row=r, column=total_col, value=total_money)
            cell.number_format = '#,##0.00'
            cell.font = Font(bold=True)

        # Set column widths , number format and กรอบ
        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[3].column_letter
            for cell in column_cells:
                cell.border = thin_border
                try:
                    if cell.value == Decimal('0') :
                        cell.value = ''
                    if isinstance(cell.value, Decimal):
                        cell.number_format = '#,##0.00'
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = 14
            if column == 'C':
                sheet.column_dimensions[column].width = 25
            else:
                sheet.column_dimensions[column].width = adjusted_width
            sheet.column_dimensions[column].height = 20

        sheet.freeze_panes = "D4"
    else:
        sheet.cell(row = 1, column = 1, value = f'ไม่มีข้อมูลรายงานสรุปค่าใช่จ่ายทะเบียนรถทั้งหมด')

    output = BytesIO()
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=expenses_registration_({active}).xlsx'

    workbook.save(response)
    return response

def viewExInvoice(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)
    b_com = BaseBranchCompany.objects.get(code = active)


    if b_com.oi_invoice_code:
        data = ExOESTNH.objects.using('pg_db').filter(comcod = b_com.affiliated.name, docnum__startswith = b_com.invoice_code).order_by('-docdat', '-docnum')
    else:
        data = ExOESTNH.objects.using('pg_db').none()

    #กรองข้อมูล
    myFilter = ExOESTNHFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)
    
    context = {
        'ois': dataPage,
        'filter':myFilter,
        'ex_iv_page': "tab-active",
        'ex_iv_show': "show",
        active :"active show",
        "colorNav":"enableNav"
    }
    return render(request, "express/viewExInvoice.html", context)

def viewExOiInvoice(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)
    b_com = BaseBranchCompany.objects.get(code = active)

    if b_com.oi_invoice_code:
        data = ExOESTNH.objects.using('pg_db').filter(comcod = b_com.affiliated.name, docnum__startswith = b_com.oi_invoice_code).order_by('-docdat', '-docnum')
    else:
        data = ExOESTNH.objects.using('pg_db').none()

    #กรองข้อมูล
    myFilter = ExOESTNHFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)
    
    context = {
        'ois': dataPage,
        'filter':myFilter,
        'ex_o_i_page': "tab-active",
        'ex_o_i_show': "show",
        active :"active show",
        "colorNav":"enableNav"
    }
    return render(request, "express/viewExOiInvoice.html", context)

def viewExSOC(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)
    b_com = BaseBranchCompany.objects.get(code = active)

    if b_com.soc_code:
        data = ExOEINVH.objects.using('pg_db').filter(comcod = b_com.affiliated.name, docnum__startswith = b_com.soc_code).order_by('-docdate', '-docnum')
    else:
        data = ExOEINVH.objects.using('pg_db').none()

    #กรองข้อมูล
    myFilter = ExOEINVHFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)
    
    context = {
        'ies': dataPage,
        'filter':myFilter,
        'ex_ie_page': "tab-active",
        'ex_ie_show': "show",
        active :"active show",
        "colorNav":"enableNav"
    }
    return render(request, "express/viewExSOC.html", context)

def viewExOiSOC(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)
    b_com = BaseBranchCompany.objects.get(code = active)

    if b_com.oi_soc_code:
        data = ExOEINVH.objects.using('pg_db').filter(comcod = b_com.affiliated.name, docnum__startswith = b_com.oi_soc_code).order_by('-docdate', '-docnum')
    else:
        data = ExOEINVH.objects.using('pg_db').none()

    #กรองข้อมูล
    myFilter = ExOEINVHFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)
    
    context = {
        'ies': dataPage,
        'filter':myFilter,
        'ex_o_ie_page': "tab-active",
        'ex_o_ie_show': "show",
        active :"active show",
        "colorNav":"enableNav"
    }
    return render(request, "express/viewExOiSOC.html", context)

def sanitize_sheet_title(title):
    # Remove invalid characters and truncate to 30 characters
    return re.sub(r'[\\/*?:[\]]', '_', title)[:30]

def export_products(request):
    wb = Workbook()
    wb.remove(wb.active)  # Remove the default sheet

    headers = ['รหัสสินค้า', 'ชื่อสินค้า', 'หน่วย', 'สังกัด', '']

    products = Product.objects.select_related('unit', 'category', 'affiliated').all()

    # Group products by category
    category_map = {}
    for product in products:
        raw_category_name = product.category.name if product.category else 'ไม่ระบุหมวดหมู่'
        safe_category_name = sanitize_sheet_title(raw_category_name)
        if safe_category_name not in category_map:
            category_map[safe_category_name] = []
        category_map[safe_category_name].append(product)

    for sheet_name, category_products in category_map.items():
        ws = wb.create_sheet(title=sheet_name)

        # Show full category name as a merged header in row 1
        full_category_name = category_products[0].category.name if category_products[0].category else 'ไม่ระบุหมวดหมู่'
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        ws.cell(row=1, column=1, value=f"หมวดหมู่: {full_category_name}")
        ws.row_dimensions[1].height = 25

        # Write column headers on row 2
        ws.append(headers)

        for idx, product in enumerate(category_products, start=2):
            ws.cell(row=idx, column=1, value=product.id)
            ws.cell(row=idx, column=2, value=product.name)
            ws.cell(row=idx, column=3, value=product.unit.name if product.unit else '')
            ws.cell(row=idx, column=4, value=product.affiliated.name if product.affiliated else '')

            ws.row_dimensions[idx].height = 40

            if product.qr_code:
                img_path = os.path.join(settings.MEDIA_ROOT, product.qr_code.name)
                if os.path.exists(img_path):
                    img = XLImage(img_path)
                    img.width = 48
                    img.height = 48
                    img.anchor = f"E{idx}"
                    ws.add_image(img)

        # Optional column width settings
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['E'].width = 10

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = StreamingHttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=products_by_category.xlsx'
    return response

def create_qr_code(request):
    products = Product.objects.all()

    for product in products:
        if not product.qr_code:
            # Create product QR code
            create_product_qr(product)

    return HttpResponse("QR codes generated.")


def create_product_qr(product):
    # Create QR code
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=5,
        border=1,
    )
    qr.add_data(product.id)
    qr.make(fit=True)
    qrcode_img = qr.make_image(fill_color="black", back_color="white")

    # Crop the QR code image to remove the border
    image_data = qrcode_img.get_image()
    image_data = image_data.crop((2, 2, image_data.size[0] - 2, image_data.size[1] - 2))

    # Create a white canvas and paste the QR code on it
    canvas = Image.new('RGB', image_data.size, 'white')
    canvas.paste(image_data)

    # Save image to in-memory buffer
    buffer = BytesIO()
    canvas.save(buffer, format='PNG')
    buffer.seek(0)  # Rewind the buffer

    # Save to model
    filename = f'qr_code_{product.id}.png'
    product.qr_code.save(filename, File(buffer), save=True)

    # Close resources
    buffer.close()
    canvas.close()

@csrf_exempt
def maintenance_appsheet(request):
    err_log = ''
    if request.method == 'POST':
        try:
            data = json.loads(request.POST.get('data', '{}'))# <-- จุดนี้ ถ้า body ไม่ใช่ JSON จะ error

            column_car = data.get("column_car")
            column_car_tail = data.get("column_car_tail")
            column_name = data.get("column_name")
            column_broke_re = data.get("column_broke_re")
            column_car_st = data.get("column_car_st")
            column_ap_name = data.get("column_ap_name")
            column_ap_status = data.get("column_ap_status")
            column_mile = data.get("column_mile")
            column_id = data.get("column_id")
            column_uniq_code = data.get("column_uniq_code")
            column_comp = data.get("column_comp")

            column_car =  column_car.strip()
            column_name =  column_name.strip()
            column_comp = column_comp.strip().split()[0]
            
            try:
                comp = BaseBranchCompany.objects.get(code = column_comp)
            except BaseBranchCompany.DoesNotExist:
                comp = BaseBranchCompany.objects.get(code = 'S1')

            try:
                car = BaseCar.objects.get(name = column_car)
            except BaseCar.DoesNotExist:
                car = None
                err_log += "car : " + str(column_car) + ",\n"


            try:
                car_tail = BaseCar.objects.get(name = column_car_tail)
            except BaseCar.DoesNotExist:
                car_tail = None
                err_log += "car_tail : " + str(column_car_tail) + ",\n"

            try:
                name = User.objects.annotate(
                    full_name=Concat('first_name', Value(' '), 'last_name')
                ).get(full_name=column_name)
            except ObjectDoesNotExist:
                name = None
                err_log += "name : " + str(column_name) + ",\n"


            have_id = Maintenance.objects.filter(id = column_id).exists()
            if not have_id:
                # บันทึกลง MySQL ตามต้องการ
                Maintenance.objects.create(
                    id = column_id,
                    car = car,
                    car_tail = car_tail,
                    name = name,
                    broke_reason = column_broke_re,
                    car_state = column_car_st,
                    mile = column_mile,
                    branch_company = comp,
                    approve_status = column_ap_status,
                    err_log = err_log,
                )
                
                obj = Maintenance.objects.get(id = column_id)

                if "image1" in request.FILES and request.FILES["image1"]:
                    obj.image1 = request.FILES["image1"]

                if "image2" in request.FILES and request.FILES["image2"]:
                    obj.image2 = request.FILES["image2"]

                if "image3" in request.FILES and request.FILES["image3"]:
                    obj.image3 = request.FILES["image3"]

                if "image4" in request.FILES and request.FILES["image4"]:
                    obj.image4 = request.FILES["image4"]

                obj.save()

                return JsonResponse({"status": "success", "id": obj.id}, status=200)
            else:
                return JsonResponse({"status": "error", "message": "dupilcate id"}, status=400)
            
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)
        
    return JsonResponse({"status": "error", "message": "Invalid request"}, status=400)

@csrf_exempt
def update_maintenance_appsheet(request):
    err_log = ''
    if request.method == 'POST':
        try:
            data = json.loads(request.POST.get('data', '{}'))# <-- จุดนี้ ถ้า body ไม่ใช่ JSON จะ error

            column_ap_name = data.get("column_ap_name")
            column_ap_status = data.get("column_ap_status")
            column_id = data.get("column_id")
            column_uniq_code = data.get("column_uniq_code")

            column_ap_name =  column_ap_name.strip()

            try:
                approve_name = User.objects.annotate(
                    full_name=Concat('first_name', Value(' '), 'last_name')
                ).get(full_name=column_ap_name)
            except ObjectDoesNotExist:
                approve_name = None
                err_log += "approve_name : " + str(column_ap_name) + ",\n"
            
            try:
                if column_ap_status == 'อนุมัติซ่อมบำรุง':
                    ma = Maintenance.objects.get(id = column_id)
                    ma.approve_name = approve_name
                    ma.approve_status = column_ap_status
                    ma.approve_update = datetime.datetime.now()
                    err_log += ma.err_log
                    ma.err_log = err_log
                    ma.save()
                    return JsonResponse({"status": "success", "id":"update approve status"}, status=200)
                
            except Maintenance.DoesNotExist:
                pass
            
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)
        
    return JsonResponse({"status": "error", "message": "Invalid request"}, status=400)

@login_required(login_url='signIn')
def viewMA(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    data = Maintenance.objects.filter(is_cancel = False, is_complete__isnull = True, branch_company__code__in = company_in)

    #กรองข้อมูล
    myFilter = MaintenanceFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)

    context = {
        'mas':dataPage,
        'filter':myFilter,
        'ma_page': "tab-active",
        'ma_show': "show",
        active :"active show",
        "colorNav":"enableNav"
    }
    return render(request, "maintenance/viewMA.html", context)

@login_required(login_url='signIn')
def showMA(request, ma_id, mode):
    active = request.session['company_code']
    base_ma_type = BaseMAType.objects.all()

    ma = Maintenance.objects.get(id = ma_id)
    if ma.car:
        base_rp_type = BaseRepairType.objects.filter(Q(rq_type = ma.car.rq_type)| Q(rq_type = 4))
    else:
        base_rp_type = BaseRepairType.objects.filter(Q(rq_type = 2)| Q(rq_type = 4))

    page = MAPageMode(mode)
    show = MAShowMode(mode)

    maa_form = PurchaseOrderAddressCompanyForm(instance=ma)
    if request.method == 'POST' and 'btnformMAa' in request.POST:
        maa_form = PurchaseOrderAddressCompanyForm(request.POST, request.FILES, instance=ma)
        if maa_form.is_valid():
            maa_form.save()
            return redirect('showMA', ma_id = ma_id , mode = mode)
        
    #ถ้า user login เป็นจัดซื้อ
    isPurchasing = is_purchasing(request.user)
    #ถ้า user login เป็นจัดซื้อ
    isSupplies = is_supplies(request.user)
    isUploadeReceipt = False
    if isPurchasing or isSupplies:
        isUploadeReceipt = True

    context = {
        'ma':ma,
        'base_ma_type': base_ma_type,
        'base_rp_type': base_rp_type,
        'isUploadeReceipt':isUploadeReceipt,
        'maa_form': maa_form,
        page: "tab-active",
        show: "show",
        active :"active show",
		"disableTab":"disableTab",
		"colorNav":"disableNav"
    }
    return render(request, 'maintenance/showMA.html',context)

@login_required(login_url='signIn')
def editMA(request, ma_id):
    active = request.session['company_code']

    ma = Maintenance.objects.get(id=ma_id)

    #ดึงร้านค้าและราคาใบสั่งซื้อ
    po_dist = None
    po_amount = None
    try:
        rq_id = Requisition.objects.filter(ma_id = ma_id)[:1].values('id')
        po_id = PurchaseOrderItem.objects.filter(item__requisition_id = rq_id)[:1].values('po__id')
        po = PurchaseOrder.objects.filter(id = po_id).first()
        if po:
            po_dist = po.distributor
            po_amount = po.amount
    except Requisition.DoesNotExist or PurchaseOrder.DoesNotExist:
        pass

    if request.method == 'POST':
        form = MaintenanceForm(request.POST, request.FILES, instance=ma)
        if form.is_valid():
            ma_form = form.save()
            try:
               rqs = Requisition.objects.filter(ma_id = ma_form.pk)
               for i in rqs:
                   i.repair_type = ma_form.repair_type
                   i.save()
            except Requisition.DoesNotExist:
                pass

            #กรณีที่ยังไม่บันทึกผู้จัดทำ
            if ma.organizer is None:
                ma.organizer = request.user
                ma.save()
            CDPmRoundItem(ma.id)
            return redirect('viewMA')
    else:
        form = MaintenanceForm(instance=ma, initial={'distributor': po_dist, 'amount': po_amount})

    context = {
        'ma':ma,
        'form':form,
        'ma_page': "tab-active",
        'ma_show': "show",
        active :"active show",
		"disableTab":"disableTab",
		"colorNav":"disableNav"
    }

    return render(request, "maintenance/editMA.html", context)

def cancelMA(request, ma_id):
    ma = Maintenance.objects.get(id = ma_id)
    #set ma ยกเลิกรายการจะไปอยู่ในประวัติ
    ma.is_cancel = True
    ma.save()
    return redirect('viewMA')

def autocompalte_maintenance(request):
    active = request.session['company_code']

    if 'term' in request.GET:
        term = request.GET.get('term')
        qs = Maintenance.objects.filter(Q(ref_no__icontains = term), branch_company__code = active, is_cancel = False, is_complete__isnull = True)[:15]
        titles = list()
        for obj in qs:
            titles.append(obj.ref_no)
    return JsonResponse(titles, safe=False)

def searchDataMaintenance(request):
    id = None
    ref_no = None
    rq_type = None
    car_id = None
    car_text = None
    repair_type_id = None
    repair_type_name = None
    mile = None
    have_id = False
    if 'id_input' in request.GET:
        id_input = request.GET.get('id_input')

        try:
            qs = Maintenance.objects.get(ref_no = id_input)
            id = qs.id
            ref_no = qs.ref_no

            if qs.car.rq_type.id:
                rq_type = qs.car.rq_type.id
            if qs.car:
                car_id = qs.car.id
                car_text = f"{qs.car.code} : {qs.car.name}"
            if qs.repair_type:
                repair_type_id = qs.repair_type.id
                repair_type_name = qs.repair_type.name
            if qs.mile:
                mile = qs.mile

            have_id = True
        except Maintenance.DoesNotExist:
            pass

    data = {
        'id': id,
        'ref_no': ref_no,
        'rq_type': rq_type,
        'car_id': car_id,
        'car_text': car_text,
        'repair_type_id': repair_type_id,
        'repair_type_name': repair_type_name,
        'mile': mile,
        'have_id': have_id,
    }
    return JsonResponse(data)

def parse_sheet_time(value: str, tz="Asia/Bangkok"):
    if value:
        try:
            # 1) parse ISO string เป็น datetime UTC
            dt = datetime.datetime.fromisoformat(value.replace("Z", "+00:00"))

            # 2) แปลงจาก UTC → timezone ของ Spreadsheet
            tzinfo = pytz.timezone(tz)
            local_dt = dt.astimezone(tzinfo)

            # 3) ดึงเฉพาะเวลา
            t = local_dt.time()

            # 4) แปลงเป็น total_seconds ของวัน แล้วปัดเป็น nearest minute
            seconds = t.hour * 3600 + t.minute * 60 + t.second
            rounded_seconds = round(seconds, -1)   # ปัดเป็น 10 วินาที (หรือ -2 เพื่อปัดเป็นนาที)

            # 5) สร้าง time object ใหม่
            corrected_time = (datetime.datetime.min + datetime.timedelta(seconds=rounded_seconds)).time()

            return corrected_time
        except Exception as e:
            return None
        
def parse_decimal(value):
    try:
        if value in (None, '', ' ', 'null'):
            return Decimal(0)
        return Decimal(str(value))
    except Exception:
        return Decimal(0)
    
def parse_int(value):
    try:
        if value in (None, '', ' ', 'null'):
            return int(0)
        return int(str(value))
    except Exception:
        return int(0)
    
@csrf_exempt
def carLogBook_appsheet(request):
    err_log = ''
    if request.method == 'POST':
        try:
            data = json.loads(request.POST.get('data', '{}'))# <-- จุดนี้ ถ้า body ไม่ใช่ JSON จะ error
            #print('data =========== '+ str(data))

            column_series = data.get("column_series")
            column_uniq_code = data.get("column_uniq_code")
            column_comp = data.get("column_comp")
            column_name = data.get("column_name")
            column_car_code = data.get("column_car_code")
            column_car = data.get("column_car")

            column_comp = column_comp.strip().split()[0]
            column_name =  column_name.strip()
            column_car_code =  column_car_code.strip()
            column_car =  column_car.strip()

            column_oil = parse_decimal(data.get("column_oil"))
            column_gas = parse_decimal(data.get("column_gas"))
            column_engine = parse_decimal(data.get("column_engine"))
            column_hydraulic = parse_decimal(data.get("column_hydraulic"))
            column_grease = parse_decimal(data.get("column_grease"))

            column_mile_start = data.get("column_mile_start") 
            column_mile_end = data.get("column_mile_end")

            column_job1 = data.get("column_job1")
            column_start_job1 = parse_sheet_time(data.get("column_start_job1"))
            column_end_job1 = parse_sheet_time(data.get("column_end_job1"))

            column_job2 = data.get("column_job2")
            column_start_job2 = parse_sheet_time(data.get("column_start_job2"))
            column_end_job2 = parse_sheet_time(data.get("column_end_job2"))

            column_job3 = data.get("column_job3")
            column_start_job3 = parse_sheet_time(data.get("column_start_job3"))
            column_end_job3 = parse_sheet_time(data.get("column_end_job3"))

            column_job4 = data.get("column_job4")
            column_start_job4 = parse_sheet_time(data.get("column_start_job4"))
            column_end_job4 = parse_sheet_time(data.get("column_end_job4"))

            column_note = data.get("column_note")

            try:
                comp = BaseBranchCompany.objects.get(code = column_comp)
            except BaseBranchCompany.DoesNotExist:
                comp = BaseBranchCompany.objects.get(code = 'S1')
            
            car = BaseCar.objects.filter(
                Q(code=column_car_code) | Q(name=column_car)
            ).first()
            if not car:
                err_log += f"car code: {column_car_code}, car name: {column_car}\n"

            try:
                name = User.objects.annotate(
                    full_name=Concat('first_name', Value(' '), 'last_name')
                ).get(full_name=column_name)
            except ObjectDoesNotExist:
                name = None
                err_log += "name : " + str(column_name) + ",\n"


            have_id = CarLogbook.objects.filter(series = column_series , branch_company = comp).exists()
            if not have_id:
                try:
                    # บันทึกลง MySQL ตามต้องการ
                    CarLogbook.objects.create(
                        series = column_series,
                        car = car,
                        name = name,
                        branch_company = comp,

                        oil = column_oil,
                        gas = column_gas,
                        engine = column_engine,
                        hydraulic = column_hydraulic,
                        grease = column_grease,

                        mile_start = column_mile_start,
                        mile_end = column_mile_end,

                        job1 = column_job1,
                        start_job1 = column_start_job1,
                        end_job1 = column_end_job1,

                        job2 = column_job2,
                        start_job2 = column_start_job2,
                        end_job2 = column_end_job2,

                        job3 = column_job3,
                        start_job3 = column_start_job3,
                        end_job3 = column_end_job3,

                        job4 = column_job4,
                        start_job4 = column_start_job4,
                        end_job4 = column_end_job4,

                        note = column_note,
                        err_log = err_log,
                    )

                    obj = CarLogbook.objects.get(series = column_series , branch_company = comp)
                    if "image_mile" in request.FILES and request.FILES["image_mile"]:
                        obj.image_mile = request.FILES["image_mile"]
                    obj.save()

                except Exception as ex:
                    pass

                return JsonResponse({"status": "success", "series": column_series}, status=200)
            else:
                return JsonResponse({"status": "error", "message": "dupilcate id"}, status=400)
            
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)
        
    return JsonResponse({"status": "error", "message": "Invalid request"}, status=400)
@csrf_exempt
def roi_carLogBook_appsheet(request):
    err_log = ''
    if request.method == 'POST':
        try:
            data = json.loads(request.POST.get('data', '{}'))# <-- จุดนี้ ถ้า body ไม่ใช่ JSON จะ error
            #print('data =========== '+ str(data))

            column_series = data.get("column_series")
            column_uniq_code = data.get("column_uniq_code")
            column_comp = data.get("column_comp")
            column_name = data.get("column_name")
            column_car = data.get("column_car")
            column_car_tail = data.get("column_car_tail")

            column_comp = column_comp.strip().split()[0]
            column_name =  column_name.strip()
            column_car =  column_car.strip()
            column_car_tail =  column_car_tail.strip()

            column_oil = parse_decimal(data.get("column_oil"))
            column_engine = parse_decimal(data.get("column_engine"))
            column_hydraulic = parse_decimal(data.get("column_hydraulic"))
            column_grease = parse_decimal(data.get("column_grease"))

            column_job1 = data.get("column_job1")
            column_mile_start_job1 = parse_int(data.get("column_mile_start_job1"))
            column_mile_end_job1 = parse_int(data.get("column_mile_end_job1"))

            column_job2 = data.get("column_job2")
            column_mile_start_job2 = parse_int(data.get("column_mile_start_job2"))
            column_mile_end_job2 = parse_int(data.get("column_mile_end_job2"))

            column_job3 = data.get("column_job3")
            column_mile_start_job3 = parse_int(data.get("column_mile_start_job3"))
            column_mile_end_job3 = parse_int(data.get("column_mile_end_job3"))

            column_job4 = data.get("column_job4")
            column_mile_start_job4 = parse_int(data.get("column_mile_start_job4"))
            column_mile_end_job4 = parse_int(data.get("column_mile_end_job4"))

            column_job5 = data.get("column_job5")
            column_mile_start_job5 = parse_int(data.get("column_mile_start_job5"))
            column_mile_end_job5 = parse_int(data.get("column_mile_end_job5"))

            column_job6 = data.get("column_job6")
            column_mile_start_job6 = parse_int(data.get("column_mile_start_job6"))
            column_mile_end_job6 = parse_int(data.get("column_mile_end_job6"))

            try:
                comp = BaseBranchCompany.objects.get(code = column_comp)
            except BaseBranchCompany.DoesNotExist:
                comp = BaseBranchCompany.objects.get(code = 'S1')

            try:
                car = BaseCar.objects.get(name = column_car)
            except BaseCar.DoesNotExist:
                car = None
                err_log += "car : " + str(column_car) + ",\n"

            try:
                car_tail = BaseCar.objects.get(name = column_car_tail)
            except BaseCar.DoesNotExist:
                car_tail = None
                err_log += "car_tail : " + str(column_car_tail) + ",\n"

            try:
                name = User.objects.annotate(
                    full_name=Concat('first_name', Value(' '), 'last_name')
                ).get(full_name=column_name)
            except ObjectDoesNotExist:
                name = None
                err_log += "name : " + str(column_name) + ",\n"

            have_id = CarLogbook.objects.filter(series = column_series , branch_company = comp).exists()
            if not have_id:
                try:
                    # บันทึกลง MySQL ตามต้องการ
                    CarLogbook.objects.create(
                        series = column_series,
                        car = car,
                        car_tail = car_tail,
                        name = name,
                        branch_company = comp,

                        oil = column_oil,
                        engine = column_engine,
                        hydraulic = column_hydraulic,
                        grease = column_grease,

                        job1 = column_job1,
                        mile_start_job1 = column_mile_start_job1,
                        mile_end_job1 = column_mile_end_job1,

                        job2 = column_job2,
                        mile_start_job2 = column_mile_start_job2,
                        mile_end_job2 = column_mile_end_job2,

                        job3 = column_job3,
                        mile_start_job3 = column_mile_start_job3,
                        mile_end_job3 = column_mile_end_job3,

                        job4 = column_job4,
                        mile_start_job4 = column_mile_start_job4,
                        mile_end_job4 = column_mile_end_job4,

                        job5 = column_job5,
                        mile_start_job5 = column_mile_start_job5,
                        mile_end_job5 = column_mile_end_job5,

                        job6 = column_job6,
                        mile_start_job6 = column_mile_start_job6,
                        mile_end_job6 = column_mile_end_job6,
                        err_log = err_log,

                    )

                    obj = CarLogbook.objects.get(series = column_series , branch_company = comp)
                    if "image_mile" in request.FILES and request.FILES["image_mile"]:
                        obj.image_mile = request.FILES["image_mile"]
                    obj.save()
                    
                except Exception as ex:
                    pass

                return JsonResponse({"status": "success", "series": column_series}, status=200)
            else:
                return JsonResponse({"status": "error", "message": "dupilcate id"}, status=400)
            
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)
        
    return JsonResponse({"status": "error", "message": "Invalid request"}, status=400)

@login_required(login_url='signIn')
def viewCL(request):
    active = request.session['company_code']
    data = CarLogbook.objects.filter(branch_company__code = active)

    #กรองข้อมูล
    myFilter = CarLogbookFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)

    context = {
        'cls':dataPage,
        'filter':myFilter,
        'cl_page': "tab-active",
        'cl_show': "show",
        active :"active show",
        "colorNav":"enableNav"
    }
    return render(request, "carLogbook/viewCL.html", context)

@login_required(login_url='signIn')
def viewCLMB(request):
    ucd = UserCarDepartment.objects.get(user = request.user.id)
    company = BaseBranchCompany.objects.get(code = ucd.in_comp.code)
    active = company.code
    user_in_comp = company.code

    data = CarLogbook.objects.filter(branch_company__code = active)

    #กรองข้อมูล
    myFilter = CarLogbookFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)

    context = {
        'cls':dataPage,
        'user_in_comp':user_in_comp,
        'filter':myFilter,
        'cl_page': "tab-active",
        'cl_show': "show",
        active :"active show",
        "colorNav":"enableNav"
    }
    return render(request, "carLogbook/viewCLMB.html", context)

#หาวันแรกของเดือนนี้
def startDateInMonth(day):
    dt = datetime.datetime.strptime(f"{day}", '%Y-%m-%d')
    result = dt.replace(day=1).date()
    return f"{result}"

def excelDailyCL(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    comp = BaseBranchCompany.objects.get(code = active)
    
    start_created = request.GET.get('start_created') or None
    end_created = request.GET.get('end_created') or None

    my_q = Q()
    if start_created is not None:
        my_q &= Q(created__gte = start_created)
    if end_created is not None:
        my_q &=Q(created__lte = end_created)
    my_q &=Q(branch_company__code__in = company_in)

    current_date_time = datetime.datetime.today()

    startDate = datetime.datetime.strptime(start_created or startDateInMonth(current_date_time.strftime('%Y-%m-%d')), "%Y-%m-%d").date()
    endDate = datetime.datetime.strptime(end_created or current_date_time.strftime('%Y-%m-%d'), "%Y-%m-%d").date()

    list_date = [startDate+timedelta(days=x) for x in range((endDate-startDate).days + 1)]

    #data = CarLogbook.objects.filter(my_q)
    
    car_all = CarLogbook.objects.filter(my_q).values_list('car', flat=True).distinct()
    cars = BaseCar.objects.filter(id__in = car_all)

    workbook = openpyxl.Workbook()

    # Define custom date style only once
    if "custom_datetime" not in workbook.named_styles:
        date_style = NamedStyle(name='custom_datetime', number_format='DD/MM/YYYY')
        workbook.add_named_style(date_style)

    red_font = Font(color="FF0000")  # Hex color for red

    # Define a thin border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    startDate_dt = make_aware(datetime.datetime.combine(startDate, datetime.time.min))
    endDate_dt   = make_aware(datetime.datetime.combine(endDate, datetime.time.max))

    if cars:
        for car in cars:
            driver = (
                CarLogbook.objects.filter(
                    car=car,
                    created__range=(startDate_dt, endDate_dt),
                    branch_company__code__in=company_in
                )
                .annotate(
                    full_name=Concat(
                        F('name__first_name'),
                        Value(' '),
                        F('name__last_name'),
                        output_field=CharField()
                    )
                )
                .values('full_name')
                .annotate(count=Count('id'))
                .order_by('-count')
                .first()
            )
            driver_name = driver['full_name'] if driver else ''
            
            qs = CarLogbook.objects.filter(
                car=car, 
                created__range=(startDate_dt, endDate_dt), 
                branch_company__code__in=company_in
            )
            data_sum = qs.aggregate(
                sum_oil=Sum('oil'),
                sum_gas=Sum('gas'),
                sum_engine=Sum('engine'),
                sum_hydraulic=Sum('hydraulic'),
                sum_grease=Sum('grease'),
                sum_coolant=Sum('coolant'),
                sum_DW_water=Sum('DW_water'),
            )

            first_row = qs.order_by('created').first()
            last_row = qs.order_by('-created').first()

            if first_row and last_row:
                total_mile = last_row.mile_end - first_row.mile_start
            else:
                total_mile = 0
                            
            sheet = workbook.create_sheet(title=f"{car.code} {car.name}")
            sheet.cell(row=1, column=1, value='รายงานการใช้รถ').font = Font(bold=True)
            sheet.cell(row=2, column=1, value='หน่วยงาน')
            sheet.cell(row=3, column=1, value='ทะเบียนรถ')
            sheet.cell(row=4, column=1, value='ผู้ขับ')
            sheet.cell(row=6, column=1, value='วันที่ ' + str(startDate.strftime('%d/%m/%Y')) + " ถึง " + str(endDate.strftime('%d/%m/%Y'))) 

            sheet.cell(row=2, column=2, value=f"{comp.name}")
            sheet.cell(row=3, column=2, value=f"{car.name}")
            sheet.cell(row=4, column=2, value=f"{driver_name}")
            sheet.cell(row=3, column=12, value=f"{car.code}").font = Font(bold=True)
            
            sheet.cell(row=7, column=1, value='วันที่')
            sheet.cell(row=7, column=2, value='น้ำมันโซล่า')
            sheet.cell(row=7, column=3, value='น้ำมันเบนซิล')
            sheet.cell(row=7, column=4, value='น้ำมันเครื่อง')
            sheet.cell(row=7, column=5, value='น้ำมันไฮโดรลิค')
            sheet.cell(row=7, column=6, value='จารบี')
            sheet.cell(row=7, column=5, value='น้ำมันไฮโดรลิค')
            sheet.cell(row=7, column=6, value='จารบี')
            sheet.cell(row=7, column=7, value='น้ำยาหม้อน้ำ')
            sheet.cell(row=7, column=8, value='น้ำกลั่น')

            sheet.cell(row=7, column=9, value='ระยะทาง (เลขไมล์)')
            sheet.cell(row=7, column=11, value='รวม กม.')
            sheet.cell(row=7, column=12, value='รายละเอียดงาน')
            sheet.cell(row=7, column=13, value='หมายเหตุ')
            sheet.cell(row=7, column=14, value='ชื่อผู้ขับ')


            sheet.cell(row=8, column=9, value='เริ่มต้น')
            sheet.cell(row=8, column=10, value='สิ้นสุด')

            sheet.merge_cells(f'A1:L1')
            sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')

            sheet.merge_cells(f'A6:L6')
            sheet.merge_cells(f'A7:A8')
            sheet.merge_cells(f'B7:B8')
            sheet.merge_cells(f'C7:C8')
            sheet.merge_cells(f'D7:D8')
            sheet.merge_cells(f'E7:E8')
            sheet.merge_cells(f'F7:F8')
            sheet.merge_cells(f'G7:G8')
            sheet.merge_cells(f'H7:H8')

            sheet.merge_cells(f'I7:J7')
            sheet.merge_cells(f'K7:K8')
            sheet.merge_cells(f'L7:L8')
            sheet.merge_cells(f'M7:M8')
            sheet.merge_cells(f'N7:N8')

            total_distance = 0
            row_index = 9
            for idl, ldate in enumerate(list_date):
                row_index += 1
                start_dt = make_aware(datetime.datetime.combine(ldate, datetime.time.min))
                end_dt   = make_aware(datetime.datetime.combine(ldate, datetime.time.max))

                # aggregate sums
                data = CarLogbook.objects.filter(
                    car=car, 
                    created__range=(start_dt, end_dt), 
                    branch_company__code__in=company_in
                ).aggregate(
                    sum_oil=Sum('oil'),
                    sum_gas=Sum('gas'),
                    sum_engine=Sum('engine'),
                    sum_hydraulic=Sum('hydraulic'),
                    sum_grease=Sum('grease'),
                    sum_coolant=Sum('coolant'),
                    sum_DW_water=Sum('DW_water'),
                )

                # first and last records
                data_first = CarLogbook.objects.filter(
                    car=car, created__range=(start_dt, end_dt),
                    branch_company__code__in=company_in
                ).order_by('created').first()

                data_last = CarLogbook.objects.filter(
                    car=car, created__range=(start_dt, end_dt),
                    branch_company__code__in=company_in
                ).order_by('created').last()


                notes = CarLogbook.objects.filter(
                    car=car,
                    created__range=(start_dt, end_dt),
                    branch_company__code__in=company_in
                ).annotate(
                    full_name=Concat(
                        F('name__first_name'),
                        Value(' '),
                        F('name__last_name'),
                        output_field=CharField()
                    )
                ).values_list('note', 'full_name', 'job1','job2','job3','job4','job5','job6')

                # clean job text
                cleaned_notes = []
                for note, full_name, *jobs in notes:
                    job_text = ", ".join([j for j in jobs if j])  # filter None/empty
                    cleaned_notes.append((note, full_name, job_text))

                # แยกเป็น string สำหรับเขียน Excel
                notes_text       = ", ".join([row[0] for row in cleaned_notes if row[0]])
                notes_name_text  = ", ".join([row[1] for row in cleaned_notes if row[1]])
                job_text         = ", ".join([row[2] for row in cleaned_notes if row[2]])

                # write date
                sheet.cell(row=idl+9, column=1, value=ldate).style = "custom_datetime"
                sheet.cell(row=idl+9, column=1).alignment = Alignment(horizontal='center')

                # write sums
                if data:
                    sheet.cell(row=idl+9, column=2, value=data['sum_oil'] or '')
                    sheet.cell(row=idl+9, column=3, value=data['sum_gas'] or '')
                    sheet.cell(row=idl+9, column=4, value=data['sum_engine'] or '')
                    sheet.cell(row=idl+9, column=5, value=data['sum_hydraulic'] or '')
                    sheet.cell(row=idl+9, column=6, value=data['sum_grease'] or '')
                    sheet.cell(row=idl+9, column=7, value=data['sum_coolant'] or '')
                    sheet.cell(row=idl+9, column=8, value=data['sum_DW_water'] or '')

                    sheet.cell(row=idl+9, column=12, value=job_text)
                    sheet.cell(row=idl+9, column=13, value=notes_text)
                    sheet.cell(row=idl+9, column=14, value=notes_name_text)

                # write mile_start and mile_end
                if data_first:
                    sheet.cell(row=idl+9, column=9, value=data_first.mile_start)
                if data_last:
                    sheet.cell(row=idl+9, column=10, value=data_last.mile_end)

                # write distance
                if data_first and data_last:
                    distance = data_last.mile_end - data_first.mile_start
                    total_distance += distance
                    sheet.cell(row=idl+9, column=11, value=distance)
            
            sheet.cell(row=row_index, column=1, value='รวม')
            sheet.cell(row=row_index, column=2, value=data_sum['sum_oil'] or '')
            sheet.cell(row=row_index, column=3, value=data_sum['sum_gas'] or '')
            sheet.cell(row=row_index, column=4, value=data_sum['sum_engine'] or '')
            sheet.cell(row=row_index, column=5, value=data_sum['sum_hydraulic'] or '')
            sheet.cell(row=row_index, column=6, value=data_sum['sum_grease'] or '')
            sheet.cell(row=row_index, column=7, value=data_sum['sum_coolant'] or '')
            sheet.cell(row=row_index, column=8, value=data_sum['sum_DW_water'] or '')

            sheet.cell(row=row_index, column=10, value=total_mile)
            sheet.cell(row=row_index, column=11, value=total_distance)

            sheet.cell(row=2, column=4, value='รวมกม.')
            sheet.cell(row=3, column=4, value='เลขไมล์เริ่ม - เลขไมล์สิ้นสุด')
            sheet.cell(row=4, column=4, value='อัตราการสิ้นเปลืองน้ำมันจริง')

            sheet.cell(row=2, column=6, value=f'{total_mile}')
            sheet.cell(row=3, column=6, value=f'{total_distance}')

            if data_sum['sum_oil'] and  total_distance:
                num = data_sum['sum_oil'] / total_distance
                rate = round(num, 2)
                sheet.cell(row=4, column=6, value=f'{rate}')

            sheet.cell(row=2, column=9, value='กม.')
            sheet.cell(row=3, column=9, value='กม.')
            sheet.cell(row=4, column=9, value='ลิตร/กม.')

            for col in range(1, 15):  # columns 1 to 11 (A to K)
                col_letter = get_column_letter(col)
                if col > 11:
                    sheet.column_dimensions[col_letter].width = 40
                else:
                    sheet.column_dimensions[col_letter].width = 10
            
            # 2. Apply border to all cells from row 7 to last row (row_index)
            for row in range(7, row_index + 1):
                for col in range(1, 15):
                    cell = sheet.cell(row=row, column=col)
                    cell.border = thin_border

            # 3. Apply red font to the total row
            for col in [1,2,3,4,5,6,7,8,10,11]:  # only columns with totals
                sheet.cell(row=row_index, column=col).font = red_font

        workbook.remove(workbook['Sheet'])
    else:
        worksheet = workbook.active
        worksheet.cell(row=1, column=1, value='ไม่มีข้อมูลบันทึกปฎิบัติการโรงโม่ดือนนี้')

    # Save workbook into memory
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    size = output.getbuffer().nbytes

    # Generator to stream file in chunks
    def file_iterator(buffer, chunk_size=8192):
        while True:
            data = buffer.read(chunk_size)
            if not data:
                break
            yield data

    response = StreamingHttpResponse(
        file_iterator(output),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="Driver is report({active}).xlsx"'
    response["Content-Length"] = str(size)
    return response

@login_required(login_url='signIn')
def mobileMenu(request):
    try:
        user_profile = UserProfile.objects.get(user_id = request.user.id)
        visible_tab = BaseVisible.objects.filter(userprofile = user_profile)
    except:
        visible_tab = None
    
    isStaff = False
    isMobile = False
    isPermission = False

    ###################### Mobile Menu ##################
    #มีสิทธิอนุมัติ pr, cp หรือ po
    try:
        user_profile = UserProfile.objects.get(user_id = request.user.id)
        isPermission = PositionBasePermission.objects.filter(position_id = user_profile.position_id).exists()
    except UserProfile.DoesNotExist or PositionBasePermission.DoesNotExist:
        pass

    for tab in visible_tab:
        if 'Category' in tab.name:
            isStaff = True
            break
        if 'Mobile' in tab.name:
            isMobile = True
            break

    #สังกัดบริษัทของคนนั้นๆ
    user_in_comp = None
    try:
        ucd = UserCarDepartment.objects.get(user = request.user.id)
        company = BaseBranchCompany.objects.get(code = ucd.in_comp.code)
        user_in_comp = company.code
    except:
        pass

    isApproveMA = False
    try:
        isApproveMA = ApproveCarDepartment.objects.filter(user = request.user.id).exists()
    except ApproveCarDepartment.DoesNotExist:
        pass

    isMobileMenu = False
    if isMobile and not isPermission and not isStaff:
        isMobileMenu = True

    context = {
        "disableTab":"disableTab",
        "colorNav":"disableNav",
        "user_in_comp": user_in_comp,
        "isApproveMA": isApproveMA,
        "isMobileMenu": isMobileMenu,
    }
    return render(request, "mobileApp/menu.html", context)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def createCL(request):
    ucd = UserCarDepartment.objects.get(user = request.user.id)
    company = BaseBranchCompany.objects.get(code = ucd.in_comp.code)
    active = company.code
    
    try:
        form = CarLogbookForm(request.POST or None, initial={'branch_company': company, 'name': request.user,})
        if form.is_valid():
            form = CarLogbookForm(request.POST or None, request.FILES)
            new_contact = form.save(commit=False)
            new_contact.save()

            messages.success(request, "เพิ่มบันทึกการใช้รถ " + str(new_contact.ref_no) + " เรียบร้อยแล้ว")
            return redirect('mobileMenu')
    except ValueError:
        pass

    context = {
        'form':form,
        'cl_page': "tab-active",
        'cl_show': "show",
        active :"active show",
        "disableTab":"disableTab",
        "colorNav":"disableNav"
    }
    return render(request, "carLogbook/createCL.html", context)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def editCL(request, cl_id):
    active = request.session['company_code']
    cl = CarLogbook.objects.get(id=cl_id)

    if request.method == 'POST':
        form = CarLogbookForm(
            request.POST,
            request.FILES,
            instance=cl
        )
        if form.is_valid():
            new_contact = form.save()
            return redirect('viewCL')
    else:
        # ✅ ต้องส่ง request ด้วย
        form = CarLogbookForm(instance=cl)

    context = {
        'form': form,
        'cl_page': "tab-active",
        'cl_show': "show",
        'cl_data': cl,
        active: "active show",
        "disableTab": "disableTab",
        "colorNav": "disableNav"
    }
    return render(request, "carLogbook/editCL.html", context)

def cancelCL(request, cl_id):
    cl = CarLogbook.objects.get(id = cl_id)
    #set ma ยกเลิกรายการจะไปอยู่ในประวัติ
    cl.is_cancel = True
    cl.save()
    return redirect('viewCL')


def job_car_dep_autocomplete(request):
    term = request.GET.get("term")
    car_id = request.GET.get("car")
    #ucd = UserCarDepartment.objects.get(user = request.user.id)
    car = BaseCar.objects.get(id = car_id)
    
    jobs = BaseJobCarDep.objects.filter(name__icontains=term , car_dep = car.car_dep)[:20]
    results =   [
        {"label": p.name, "value": p.name} for p in jobs
    ]
    return JsonResponse(results, safe=False)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def createCLRoi(request):
    ucd = UserCarDepartment.objects.get(user = request.user.id)
    company = BaseBranchCompany.objects.get(code = ucd.in_comp.code)
    active = company.code
    
    try:
        form = RoiCarLogbookForm(request, request.POST or None, initial={'branch_company': company, 'name': request.user, 'car': ucd.car})
        if form.is_valid():
            form = RoiCarLogbookForm(request, request.POST or None, request.FILES)
            new_contact = form.save(commit=False)
            new_contact.save()

            messages.success(request, "เพิ่มบันทึกการใช้รถ " + str(new_contact.ref_no) + " เรียบร้อยแล้ว")
            return redirect('mobileMenu')
    except ValueError:
        pass

    context = {
        'form':form,
        'cl_page': "tab-active",
        'cl_show': "show",
        active :"active show",
        "disableTab":"disableTab",
        "colorNav":"disableNav"
    }
    return render(request, "carLogbook/createCLRoi.html", context)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def createMA(request):
    ucd = UserCarDepartment.objects.get(user = request.user.id)
    company = BaseBranchCompany.objects.get(code = ucd.in_comp.code)
    active = company.code
    
    try:
        form = CrMaintenanceForm(request, request.POST or None, initial={'branch_company': company, 'name': request.user, 'approve_status': 'ขออนุมัติซ่อมบำรุง', 'car': ucd.car})
        if request.method == 'POST':
            if form.is_valid():
                name_pm = request.POST['name_pm'] or None
                
                form = CrMaintenanceForm(request, request.POST or None, request.FILES)
                new_contact = form.save(commit=False)
                new_contact.save()

                if name_pm == '1':
                    ma_type = BaseMAType.objects.get(id = '04')
                    new_contact.ma_type = ma_type
                    new_contact.save()
                    createPmRoundItem(new_contact.id)

            messages.success(request, "เพิ่มบันทึกการแจ้งซ่อม " + str(new_contact.ref_no) + " เรียบร้อยแล้ว")
            return redirect('mobileMenu')
    except ValueError:
        pass

    #ถ้า user มีสิทธิ edit pr
    isChangeCarMa = is_change_car_ma(request.user)

    context = {
        'form':form,
        'ma_page': "tab-active",
        'ma_show': "show",
        'isChangeCarMa': isChangeCarMa,
        active :"active show",
        "disableTab":"disableTab",
        "colorNav":"disableNav"
    }
    return render(request, "maintenance/createMA.html", context)


def CDPmRoundItem(ma_id):
    have_pm = False
    try:
        have_pm = PmRoundItem.objects.filter(ma_id = ma_id).exists()
    except PmRoundItem.DoesNotExist:
        pass

    ma = Maintenance.objects.get(id = ma_id)
    if ma.ma_type.id == '04':
        if not have_pm:
            createPmRoundItem(ma_id)
    else:
        try:
            pm = PmRoundItem.objects.filter(ma_id = ma_id)
            pm.delete()
        except PmRoundItem.DoesNotExist:
            pass

def createPmRoundItem(ma_id):
    ma = Maintenance.objects.get(id = ma_id)
    obj = PmRoundItem.objects.create(
        car = ma.car,
        created = ma.created,
        update = datetime.datetime.now(),
        num_pm = ma.mile,
        ma_id = ma.id,
        ma_ref_no = ma.ref_no,
    )
    obj.save()

@login_required(login_url='signIn')
def viewMAApprove(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    # ได้รายการ car_dep และ branch_company ของ user นี้
    acdr = ApproveCarDepartment.objects.filter(user=request.user.id).values_list('car_dep', 'branch_company')

    # ดึงเฉพาะ car_dep ทั้งหมดเป็น list
    car_deps = [row[0] for row in acdr]

    # ดึงเฉพาะ branch_company ทั้งหมดเป็น list
    branch_companies = [row[1] for row in acdr]

    # ดึง user ที่อยู่ใน car_dep เหล่านั้น
    ucd = BaseCar.objects.filter(car_dep__in=car_deps).values_list('id', flat=True)

    # filter Maintenance ตามเงื่อนไข
    data = Maintenance.objects.filter(
        approve_status='ขออนุมัติซ่อมบำรุง',
        branch_company__in=branch_companies,
        car__in=ucd
    )

    '''
    #กรองข้อมูล
    myFilter = MaintenanceFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)    
    '''

    context = {
        'mas':data,
        'ma_page': "tab-active",
        'ma_show': "show",
        "disableTab":"disableTab",
        "colorNav":"disableNav"
    }
    return render(request, "maintenanceApprove/viewMAApprove.html", context)

@login_required(login_url='signIn')
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def editMAApprove(request, ma_id, mode):
    active = request.session['company_code']
    base_ma_type = BaseMAType.objects.all()

    ma = Maintenance.objects.get(id = ma_id)
    if ma.car:
        base_rp_type = BaseRepairType.objects.filter(Q(rq_type = ma.car.rq_type)| Q(rq_type = 4))
    else:
        base_rp_type = BaseRepairType.objects.filter(Q(rq_type = 2)| Q(rq_type = 4))

    page = MAPageMode(mode)
    show = MAShowMode(mode)

    if request.method == 'POST':
        post_status = request.POST['status'] or None

        if post_status == 'อนุมัติ':
            obj = Maintenance.objects.get(id = ma_id)
            obj.approve_status = 'อนุมัติซ่อมบำรุง'
            obj.approve_name_id = request.user.id
            obj.approve_update = datetime.datetime.now()
            obj.save()
            messages.success(request, "อนุมัติแจ้งซ่อม " + str(obj.ref_no) + " เรียบร้อยแล้ว")
        return redirect('viewMAApprove')

    context = {
        'ma':ma,
        'base_ma_type': base_ma_type,
        'base_rp_type': base_rp_type,
        'ma_page': "tab-active",
        'ma_show': "show",
        "disableTab":"disableTab",
        "colorNav":"disableNav"
    }
    return render(request, 'maintenanceApprove/editMAApprove.html',context)

def searchExpenseDeptByJob(request):
    if 'job_id' in request.GET:
        job_id = request.GET.get('job_id')

        exd_job = BaseJobCarDep.objects.filter(id = job_id).values('expense_dept__id', 'expense_dept__name')
    data = {
        'exd_job_list': list(exd_job),
    }
    return JsonResponse(data)

def searchJobByCarDep(request):
    if 'car' in request.GET:
        car = request.GET.get('car')

        bc_dep = BaseCar.objects.get(id = car)
        job = BaseJobCarDep.objects.filter(car_dep = bc_dep.car_dep).values('id', 'name')
        
    data = {
        'job_list': list(job),
    }
    return JsonResponse(data)

def formatHourMinute(time):
    result = None
    if time:
       result = f'{time}'[:-3]
    return result

def format_duration(duration):
    result = None
    if duration:
        hours = duration // timedelta(hours=1)
        minutes = (duration % timedelta(hours=1)) // timedelta(minutes=1)
        result = f"{hours:02d}:{minutes:02d}"
    else:
        result = f"-"
    return result

def excelExpensesByCarLog(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    comp = BaseBranchCompany.objects.get(code = active)
    
    start_created = request.GET.get('start_created') or None
    end_created = request.GET.get('end_created') or None


    current_date_time = datetime.datetime.today()

    startDate = datetime.datetime.strptime(start_created or startDateInMonth(current_date_time.strftime('%Y-%m-%d')), "%Y-%m-%d").date()
    endDate = datetime.datetime.strptime(end_created or current_date_time.strftime('%Y-%m-%d'), "%Y-%m-%d").date()

    startDate_dt = make_aware(datetime.datetime.combine(startDate, datetime.time.min))
    endDate_dt = make_aware(datetime.datetime.combine(endDate, datetime.time.max))

    my_q = Q()
    if start_created is not None:
        my_q &= Q(created__gte = startDate_dt)
    if end_created is not None:
        my_q &=Q(created__lte = endDate_dt)
    my_q &=Q(branch_company__code__in = company_in)


    #data = CarLogbook.objects.filter(my_q)
    data = CarLogbook.objects.filter(my_q).values('car__id', 'car__code', 'car__name').annotate(s_oil = Sum('oil'), s_gas = Sum('gas'), s_engine = Sum('engine'), s_hydraulic = Sum('hydraulic'), s_grease = Sum('grease'), s_coolant = Sum('coolant'), s_DW_water = Sum('DW_water'))

    qs = CarLogbook.objects.filter(
        my_q,
        Q(exd_job1__isnull=False) |
        Q(exd_job2__isnull=False) |
        Q(exd_job3__isnull=False) |
        Q(exd_job4__isnull=False)
    ).values_list('exd_job1', 'exd_job2', 'exd_job3', 'exd_job4')

    # Flatten and remove None
    exd_ids = {exd for row in qs for exd in row if exd}

    exd_all = BaseExpenseDepartment.objects.filter(id__in=exd_ids).values('id', 'name')

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Define custom date style only once
    if "custom_datetime" not in workbook.named_styles:
        date_style = NamedStyle(name='custom_datetime', number_format='DD/MM/YYYY')
        workbook.add_named_style(date_style)

    # Define a thin border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    if data:
        sheet.cell(row=1, column=1, value=f'สรุปค่าใช้จ่ายแต่ละหน่วยงาน ตามทะเบียนรถ {comp.affiliated.name_th}').font = Font(bold=True)
        sheet.cell(row=2, column=1, value='วันที่ ' + str(startDate.strftime('%d/%m/%Y')) + " ถึง " + str(endDate.strftime('%d/%m/%Y')))
        sheet.cell(row=3, column=1, value='')

        row = []
        row.extend(['ลำดับที่', 'รหัส', 'ทะเบียนรถ', 'ปริมาณที่ใช้', '', '', '', '', '', '', 'อะไหล่ + ค่าซ่อม'])
        row.extend(['จำนวนชั่วโมงทำงาน' for dept in exd_all])
        row.extend(['รวม'])
        sheet.append(row)

        row2 = []
        row2.extend(['ลำดับที่', 'รหัส', 'ทะเบียนรถ', 'น้ำมันโซล่า', 'น้ำเบนซิล', 'น้ำมันเครื่อง', 'น้ำมันไฮโดรลิค', 'จารบี', 'น้ำยาหม้อน้ำ', 'น้ำกลั่น', 'อะไหล่ + ค่าซ่อม'])
        row2.extend([dept['name'] for dept in exd_all])
        row2.extend(['รวม'])
        sheet.append(row2)

        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14 + len(exd_all))
        sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=14 + len(exd_all))
        sheet.cell(row=2, column=1).alignment = Alignment(horizontal='center')
        sheet.merge_cells(f'A4:A5')
        sheet.merge_cells(f'B4:B5')
        sheet.merge_cells(f'C4:C5')
        sheet.merge_cells(f'D4:J4')
        sheet.merge_cells(f'K4:K5')

        sheet.merge_cells(start_row=4, start_column=12, end_row=4, end_column=12 + len(exd_all))

        for idx, car in enumerate(data, start=1):
            row3 = []
            row3.extend([idx, car['car__code'], car['car__name'], car['s_oil'], car['s_gas'], car['s_engine'], car['s_hydraulic'], car['s_grease'], car['s_coolant'], car['s_DW_water']])
            
            #ใบจ่ายสินค้าภายใน - จ่ายอะไหล่ 
            l_iv = list(
                ExOESTNH.objects.using('pg_db')
                .filter(
                    docnum__startswith=comp.invoice_code,
                    comcod=comp.affiliated.name,
                    docdat__range=(start_created, end_created),
                    remark = f"{car['car__name']}:{car['car__code']}"
                )
                .values_list('docnum', flat=True)
            )
            sum_trnval = ExOESTND.objects.using('pg_db').filter(
                comcod=comp.affiliated.name,
                docnum__in=l_iv
            ).aggregate(
                sum_other=Sum(Case(
                    When(stktyp ='อะไหล่', then='trnval'),
                    output_field=models.DecimalField()
                ))
            )['sum_other'] or Decimal(0) 
            row3.extend([sum_trnval])

            for dept in exd_all:
                total = (
                    CarLogbook.objects.filter(
                        my_q,
                        Q(exd_job1=dept['id']) | Q(exd_job2=dept['id']) | Q(exd_job3=dept['id']) | Q(exd_job4=dept['id']),
                        car=car['car__id'],
                    )
                    .aggregate(
                        total_diff=Sum(
                            Case(
                                When(exd_job1=dept['id'], then=F('diff_time_job1')),
                                When(exd_job2=dept['id'], then=F('diff_time_job2')),
                                When(exd_job3=dept['id'], then=F('diff_time_job3')),
                                When(exd_job4=dept['id'], then=F('diff_time_job4')),
                                output_field=models.DurationField(),
                            )
                        )
                    )['total_diff'] or None
                )

                row3.append(format_duration(total))
            

            total_diff = (
                CarLogbook.objects
                .filter(my_q, car=car['car__id'])
                .aggregate(
                    total=(
                        Sum(
                            Case(
                                When(exd_job1__isnull=False, then=Coalesce(F('diff_time_job1'), Value(0))),
                                default=Value(0),
                                output_field=models.DurationField(),
                            )
                        )
                        + Sum(
                            Case(
                                When(exd_job2__isnull=False, then=Coalesce(F('diff_time_job2'), Value(0))),
                                default=Value(0),
                                output_field=models.DurationField(),
                            )
                        )
                        + Sum(
                            Case(
                                When(exd_job3__isnull=False, then=Coalesce(F('diff_time_job3'), Value(0))),
                                default=Value(0),
                                output_field=models.DurationField(),
                            )
                        )
                        + Sum(
                            Case(
                                When(exd_job4__isnull=False, then=Coalesce(F('diff_time_job4'), Value(0))),
                                default=Value(0),
                                output_field=models.DurationField(),
                            )
                        )
                    )
                )['total'] or None 
            )
            row3.append(format_duration(total_diff))    
            sheet.append(row3)

        col_index = 13 + len(exd_all)
        row_index = 4 + len(data)
        for col in range(1, col_index):  # columns 1 to 11 (A to K)
            col_letter = get_column_letter(col)
            if col > 10:
                sheet.column_dimensions[col_letter].width = 20
            else:
                sheet.column_dimensions[col_letter].width = 10
            
        # 2. Apply border to all cells from row 7 to last row (row_index)
        for row in range(4, row_index + 2):
            for col in range(1, col_index):
                cell = sheet.cell(row=row, column=col)
                cell.border = thin_border
                if row == 4 or row == 5:
                    cell.alignment = Alignment(horizontal='center')
    else:
        sheet.cell(row = 1, column = 1, value = f'ไม่มีข้อมูลรายงานสรุปค่าใช่จ่ายแต่ละหน่วยงาน ตามทะเบียนรถ')

    output = BytesIO()
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(), 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Driver is report summary({active}).xlsx'

    workbook.save(response)
    return response

def searchPmRound(request):
    alert_pm = False
    if 'id_car' in request.GET and 'mile_end' in request.GET:
        id_car = request.GET.get('id_car')
        mile_end = request.GET.get('mile_end')

        try:
            car = BaseCar.objects.get(id = id_car)
            pm_item = PmRoundItem.objects.filter(car = id_car).order_by('-created').first()

            if car and pm_item:
                next_pm = car.pm_round + pm_item.num_pm
                if parse_int(mile_end) >= next_pm:
                    alert_pm = True
        except BaseCar.DoesNotExist or PmRoundItem.DoesNotExist:
            pass
             
    data = {
        'alert_pm': alert_pm,
        'next_pm': next_pm,
    }
    return JsonResponse(data)

@login_required(login_url='signIn')
def viewCLReport(request):
    active = request.session['company_code']
    data = CarLogbook.objects.filter(branch_company__code = active)

    #กรองข้อมูล
    myFilter = CarLogbookFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    dataPage = p.get_page(page)

    context = {
        'cls':dataPage,
        'filter':myFilter,
        'rp_cl_page': "tab-active",
        'rp_cl_show': "show",
        active :"active show",
        "colorNav":"enableNav"
    }
    return render(request, "report/viewCL.html", context)